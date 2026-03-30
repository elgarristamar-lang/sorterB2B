# Version: 0.01
# gantt_1h_visual.py
# ---------------------------------------------------------
# Gantt 1H VISUAL (semana DOMINGO->SÁBADO)
# - Texto en celdas: agrupación playa (sin [B2B], sin bloque, sin día)
# - Comentario/nota: reducido (para que se vea bien) y sin bloque/día
# - BLOQUES_DESTINOS: 1 fila por destino, + columnas:
#       POSTEX_N | POSTEX_POS_SUBRAMPA | POSTEX_POS_LIST (COMPLETA, sin +5)
# - used = SOLO POSTEX (posiciones suelo)
# ---------------------------------------------------------

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional, Dict, Tuple, Set, List
from collections import defaultdict
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


# -----------------------------
# Paths / config
# -----------------------------
import sys as _sys

def _parse_args():
    """gantt_1h.py <capacity.csv> <grupo_destinos.xlsx> <bloques_horarios.xlsx> <output.xlsx> [grupo_sheet]"""
    args = _sys.argv[1:]
    if len(args) < 4:
        print("Uso: python gantt_1h.py <ramp_capacity.csv> <grupo_destinos.xlsx> <bloques_horarios.xlsx> <output.xlsx> [hoja]")
        _sys.exit(1)
    return Path(args[0]), Path(args[1]), Path(args[2]), Path(args[3]), args[4] if len(args) > 4 else "Hoja1"

CAPACITY_CSV, GRUPO_XLSX, BLOQUES_XLSX, OUTPUT_XLSX, GRUPO_SHEET = _parse_args()

EXCLUDE_PREFIXES = ("R01", "R03")
WEEK_MINUTES = 7 * 1440

AUTHOR = "gantt_1h_visual"
MAX_COMMENT = 500  # <-- reduce tamaño de nota / comentario


# -----------------------------
# Helpers
# -----------------------------
def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = {c.lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.lower()
        if key in cols:
            return cols[key]
    return None


def ramp_sort_key(subramp: str):
    subramp = subramp.strip().upper()
    m = re.match(r"^R(\d{2})([A-D])$", subramp)
    if not m:
        return (999, "Z", subramp)
    return (int(m.group(1)), m.group(2), subramp)


def filter_rows_by_block(df: pd.DataFrame, block: str, desc_col: str) -> pd.DataFrame:
    s = df[desc_col].astype(str).fillna("")
    block = block.strip().upper()

    if "BLO" in block:
        mask_literal = s.str.contains(re.escape(block), case=False, na=False)
        if mask_literal.any():
            return df[mask_literal].copy()

    mask_priority = s.str.contains(fr"BLO{re.escape(block)}", case=False, na=False)
    if mask_priority.any():
        return df[mask_priority].copy()

    mask_fallback = s.str.contains(re.escape(block), case=False, na=False)
    return df[mask_fallback].copy()


def clean_desc(raw_desc: str) -> str:
    s = (raw_desc or "").strip()
    s = re.sub(r"^\s*\[[^\]]+\]\s*", "", s)  # quita "[B2B] "
    return s.strip()


DAY_NAMES = ["DOMINGO", "LUNES", "MARTES", "MIERCOLES", "MIÉRCOLES", "JUEVES", "VIERNES", "SABADO", "SÁBADO"]


def clean_agr_playa(raw_desc: str) -> str:
    """
    Entrada típica: "[B2B] 1BLOD4_LUNES_ESPANA_BENAVENTE_TSA"
    Salida: "ESPANA_BENAVENTE_TSA"
    """
    s = clean_desc(raw_desc)  # quita [B2B]
    s = re.sub(r"^\s*\d+BLO[A-Z]\d+_", "", s)  # quita "1BLOD4_"
    s = re.sub(
        r"^\s*(DOMINGO|LUNES|MARTES|MIERCOLES|MIÉRCOLES|JUEVES|VIERNES|SABADO|SÁBADO)_",
        "",
        s,
        flags=re.IGNORECASE,
    )
    return s.strip()


def short_label(full_text: str) -> str:
    """
    full_text aquí ya viene como:
      "ESPANA_BENAVENTE_TSA + ESPANA_BALEARES_TSA"
    """
    if not full_text:
        return ""

    parts = [p.strip() for p in full_text.split(" + ") if p.strip()]
    first = parts[0] if parts else full_text
    if len(parts) > 1:
        return f"{first} (+{len(parts)-1})"
    return first


# -----------------------------
# Schedule logic (DOMINGO->SÁBADO)
# -----------------------------
DAY_TO_INDEX = {
    "DOMINGO": 0,
    "LUNES": 1,
    "MARTES": 2,
    "MIERCOLES": 3,
    "MIÉRCOLES": 3,
    "JUEVES": 4,
    "VIERNES": 5,
    "SABADO": 6,
    "SÁBADO": 6,
}

INDEX_TO_DAY = {
    0: "DOMINGO",
    1: "LUNES",
    2: "MARTES",
    3: "MIERCOLES",
    4: "JUEVES",
    5: "VIERNES",
    6: "SABADO",
}


def _time_to_minutes(t) -> int:
    if pd.isna(t):
        return 0
    if hasattr(t, "hour") and hasattr(t, "minute"):
        return int(t.hour) * 60 + int(t.minute)
    s = str(t).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    return int(m.group(1)) * 60 + int(m.group(2)) if m else 0


def load_bloques_horarios(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    df = standardize_columns(df)

    col_block = find_col(df, ["NUEVO BLOQUE", "BLOQUE", "BLOCK"])
    col_start_day = find_col(df, ["Día LIBERACIÓN BLOQUES", "DIA LIBERACIÓN BLOQUES", "DIA LIBERACION BLOQUES", "DIA LIBERACION"])
    col_start_time = find_col(df, ["Hora LIBERACIÓN BLOQUES", "HORA LIBERACIÓN BLOQUES", "HORA LIBERACION BLOQUES", "HORA LIBERACION"])
    col_end_day = find_col(df, ["Día DESACTIVACIÓN", "DIA DESACTIVACIÓN", "DIA DESACTIVACION"])
    col_end_time = find_col(df, ["Hora DESACTIVACIÓN", "HORA DESACTIVACIÓN", "HORA DESACTIVACION"])

    if not all([col_block, col_start_day, col_start_time, col_end_day, col_end_time]):
        raise ValueError(f"bloques_horarios inválido. Columnas: {list(df.columns)}")

    df[[col_block, col_start_day, col_start_time, col_end_day, col_end_time]] = df[
        [col_block, col_start_day, col_start_time, col_end_day, col_end_time]
    ].ffill()

    out = df.rename(columns={
        col_block: "BLOCK",
        col_start_day: "START_DAY",
        col_start_time: "START_TIME",
        col_end_day: "END_DAY",
        col_end_time: "END_TIME",
    })

    out["BLOCK"] = out["BLOCK"].astype(str).str.strip().str.upper()
    out["START_DAY"] = out["START_DAY"].astype(str).str.strip().str.upper()
    out["END_DAY"] = out["END_DAY"].astype(str).str.strip().str.upper()

    return out[["BLOCK", "START_DAY", "START_TIME", "END_DAY", "END_TIME"]]


def build_schedule_intervals(bh: pd.DataFrame) -> Dict[str, Tuple[int, int]]:
    intervals: Dict[str, Tuple[int, int]] = {}
    for _, row in bh.iterrows():
        block = str(row["BLOCK"]).strip().upper()
        sd = str(row["START_DAY"]).strip().upper()
        ed = str(row["END_DAY"]).strip().upper()

        if sd not in DAY_TO_INDEX or ed not in DAY_TO_INDEX:
            continue

        sd_idx = DAY_TO_INDEX[sd]
        ed_idx = DAY_TO_INDEX[ed]

        start = sd_idx * 1440 + _time_to_minutes(row["START_TIME"])
        end = ed_idx * 1440 + _time_to_minutes(row["END_TIME"])

        if ed_idx < sd_idx:
            end += WEEK_MINUTES
        elif ed_idx == sd_idx and end <= start:
            end += 1440

        intervals[block] = (start, end)

    return intervals


def minutes_to_day_hour_label(minute: int) -> str:
    hour_min = (minute // 60) * 60
    day_idx = (hour_min // 1440) % 7
    hour = (hour_min % 1440) // 60
    day_name = INDEX_TO_DAY.get(day_idx, "DIA")
    return f"{day_name}_{hour:02d}:00"


def build_hour_grid(min_start: int, max_end: int) -> List[int]:
    start_hour = (min_start // 60) * 60
    end_hour = ((max_end + 59) // 60) * 60
    return list(range(start_hour, end_hour, 60))


def hour_overlaps_interval(hour_start: int, interval: Tuple[int, int]) -> bool:
    a0 = hour_start
    a1 = hour_start + 60
    b0, b1 = interval
    return a0 < b1 and b0 < a1


# -----------------------------
# Parse "Elemento" to (subramp, slot)
# -----------------------------
def parse_subramp_and_slot(value: object) -> Tuple[Optional[str], Optional[int]]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, None
    s = str(value).strip().upper()
    if not s:
        return None, None

    m = re.search(r"\b(R\d{2})\s*[_\- ]\s*([A-D])\s*-\s*(\d+)\b", s)
    if m:
        return f"{m.group(1)}{m.group(2)}", int(m.group(3))

    m2 = re.search(r"\b(R\d{2})([A-D])\s*-\s*(\d+)\b", s)
    if m2:
        return f"{m2.group(1)}{m2.group(2)}", int(m2.group(3))

    m3 = re.search(r"\b(R\d{2})([A-D])\b", s)
    if m3:
        return f"{m3.group(1)}{m3.group(2)}", None

    return None, None


# -----------------------------
# Loaders
# -----------------------------
def load_capacity(path: Path) -> Dict[str, int]:
    df = pd.read_csv(path, sep=None, engine="python")
    df = standardize_columns(df)

    col_ramp = find_col(df, ["RAMP", "RAMPA", "SUBRAMPA"])
    col_pallets = find_col(df, ["PALLETS", "CAPACITY", "CAPACIDAD", "POSICIONES", "PALLETS_CAPACITY"])
    if not col_ramp or not col_pallets:
        raise ValueError(f"CSV ramp_capacity inválido. Columnas: {list(df.columns)}")

    df[col_ramp] = df[col_ramp].astype(str).str.strip().str.upper()
    df[col_pallets] = pd.to_numeric(df[col_pallets], errors="coerce")

    cap_map: Dict[str, int] = {}
    for _, row in df.iterrows():
        r = str(row[col_ramp]).strip().upper()
        if not r or any(r.startswith(pref) for pref in EXCLUDE_PREFIXES):
            continue
        if pd.isna(row[col_pallets]):
            continue
        cap_map[r] = int(row[col_pallets])

    return cap_map


def load_grupo(path: Path, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet)
    return standardize_columns(df)


# -----------------------------
# Build per-block data:
# text_by_block[block][sub] = "AGR1 + AGR2"
# used_slots_by_block[block][sub] = set(slot) SOLO POSTEX
# postex_pos_by_block_dest[block][agr] = set((sub, slot)) SOLO POSTEX
# -----------------------------
def build_block_maps(
    grupo: pd.DataFrame,
    all_blocks: Set[str],
) -> Tuple[
    Dict[str, Dict[str, str]],
    Dict[str, Dict[str, Set[int]]],
    Dict[str, Dict[str, Set[Tuple[str, int]]]],
    int
]:
    desc_col = find_col(grupo, ["Descripción Grupos de destino", "Descripcion Grupos de destino", "DESCRIPCION GRUPOS DE DESTINO"])
    elem_col = find_col(grupo, ["Elemento", "ELEMENTO"])
    tipo_col = find_col(grupo, ["Tipo de zona", "TIPO DE ZONA", "TIPO_ZONA", "TIPO ZONA"])

    if not desc_col or not elem_col:
        raise ValueError(f"GRUPO_DESTINOS: faltan columnas. Columnas: {list(grupo.columns)}")

    warnings = 0
    text_by_block: Dict[str, Dict[str, str]] = defaultdict(dict)
    used_slots_by_block: Dict[str, Dict[str, Set[int]]] = defaultdict(lambda: defaultdict(set))
    postex_pos_by_block_dest: Dict[str, Dict[str, Set[Tuple[str, int]]]] = defaultdict(lambda: defaultdict(set))

    for block in sorted(all_blocks):
        rows = filter_rows_by_block(grupo, block, desc_col)
        if rows.empty:
            continue

        agrup_by_sub: Dict[str, Set[str]] = defaultdict(set)

        for _, r in rows.iterrows():
            raw_desc = str(r[desc_col]).strip()
            agr_playa = clean_agr_playa(raw_desc) if raw_desc and raw_desc.lower() != "nan" else ""

            sub, slot = parse_subramp_and_slot(r[elem_col])

            # SOLO POSTEX para slots
            if tipo_col is not None and slot is not None and sub is not None:
                tipo = str(r[tipo_col]).strip().upper()
                if tipo == "POSTEX" and not any(sub.startswith(p) for p in EXCLUDE_PREFIXES):
                    used_slots_by_block[block][sub].add(slot)
                    if agr_playa:
                        postex_pos_by_block_dest[block][agr_playa].add((sub, slot))

            # Texto por subrampa
            if not sub:
                v = r[elem_col]
                if v is not None and str(v).strip():
                    warnings += 1
                continue
            if any(sub.startswith(p) for p in EXCLUDE_PREFIXES):
                continue

            if agr_playa:
                agrup_by_sub[sub].add(agr_playa)

        for sub, agrups in agrup_by_sub.items():
            text_by_block[block][sub] = " + ".join(sorted(agrups))

    return dict(text_by_block), used_slots_by_block, dict(postex_pos_by_block_dest), warnings


# -----------------------------
# Color assignment by block, unique per START_DAY
# -----------------------------
PALETTE = [
    "BDD7EE",  # azul claro
    "C6E0B4",  # verde claro
    "F8CBAD",  # salmón
    "FFF2CC",  # amarillo
    "D9E1F2",  # azul/lila
    "EAD1F2",  # lila
    "E2F0D9",  # verde suave
    "FCE4D6",  # naranja claro
    "E7E6E6",  # gris
    "DDEBF7",  # azul muy suave
]


def assign_colors_by_day(bh: pd.DataFrame) -> Dict[str, str]:
    start_day_by_block: Dict[str, str] = {}
    for _, row in bh.iterrows():
        b = str(row["BLOCK"]).strip().upper()
        sd = str(row["START_DAY"]).strip().upper()
        if b and sd:
            start_day_by_block[b] = sd

    blocks_by_day: Dict[str, List[str]] = defaultdict(list)
    for b, sd in start_day_by_block.items():
        blocks_by_day[sd].append(b)

    color_by_block: Dict[str, str] = {}
    for sd, blist in blocks_by_day.items():
        blist_sorted = sorted(set(blist))
        for i, b in enumerate(blist_sorted):
            color_by_block[b] = PALETTE[i % len(PALETTE)]

    return color_by_block


# -----------------------------
# Cell meta
# -----------------------------
def build_cell_meta(
    intervals: Dict[str, Tuple[int, int]],
    hour_minutes: List[int],
    text_by_block: Dict[str, Dict[str, str]],
    used_slots_by_block: Dict[str, Dict[str, Set[int]]],
    cap_map: Dict[str, int],
) -> Dict[Tuple[str, int], Dict[str, object]]:
    cell_meta: Dict[Tuple[str, int], Dict[str, object]] = {}

    for block, (start, end) in intervals.items():
        if block not in text_by_block:
            continue

        for hm in hour_minutes:
            if not hour_overlaps_interval(hm, (start, end)):
                continue

            sub_map = text_by_block[block]
            for sub, full in sub_map.items():
                key = (sub, hm)

                used = len(used_slots_by_block.get(block, {}).get(sub, set()))
                cap = int(cap_map.get(sub, 0))

                if key not in cell_meta:
                    cell_meta[key] = {
                        "blocks": [block],
                        "full": full,           # agr_playas concatenadas
                        "label": short_label(full),
                        "used": used,
                        "cap": cap,
                    }
                else:
                    cell_meta[key]["blocks"].append(block)
                    prev_full = str(cell_meta[key]["full"])
                    if prev_full != full:
                        cell_meta[key]["full"] = f"{prev_full} || {full}"
                    cell_meta[key]["used"] = int(cell_meta[key]["used"]) + used
                    cell_meta[key]["label"] = str(cell_meta[key]["label"]) + " *"

    for k, v in cell_meta.items():
        v["blocks"] = tuple(sorted(set(v["blocks"])))

    return cell_meta


# -----------------------------
# Excel sheets
# -----------------------------
def write_legend_sheet(wb: Workbook, color_by_block: Dict[str, str], bh: pd.DataFrame) -> None:
    ws = wb.create_sheet("LEYENDA")

    ws["A1"] = "LEYENDA - Colores por BLOQUE"
    ws["A1"].font = Font(bold=True, size=14)

    blocks_by_day: Dict[str, List[str]] = defaultdict(list)
    for _, row in bh.iterrows():
        b = str(row["BLOCK"]).strip().upper()
        sd = str(row["START_DAY"]).strip().upper()
        if b and sd:
            blocks_by_day[sd].append(b)

    r = 3
    for day in ["DOMINGO", "LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO"]:
        blist = sorted(set(blocks_by_day.get(day, [])))
        if not blist:
            continue

        ws.cell(row=r, column=1, value=day).font = Font(bold=True, size=12)
        r += 1

        ws.cell(row=r, column=1, value="Color").font = Font(bold=True)
        ws.cell(row=r, column=2, value="Bloque").font = Font(bold=True)
        r += 1

        for b in blist:
            hexrgb = color_by_block.get(b, "FFFFFF")
            box = ws.cell(row=r, column=1, value="")
            box.fill = PatternFill("solid", fgColor=hexrgb)
            ws.cell(row=r, column=2, value=b)
            r += 1

        r += 2

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22


def write_bloques_destinos_sheet(
    wb: Workbook,
    bh: pd.DataFrame,
    postex_pos_by_block_dest: Dict[str, Dict[str, Set[Tuple[str, int]]]],
) -> None:
    """
    Una fila por DESTINO(agr_playa) dentro de cada BLOQUE.
    Columnas:
      BLOCK | START_DAY | START_TIME | END_DAY | END_TIME | DESTINO | POSTEX_N | POSTEX_POS_SUBRAMPA | POSTEX_POS_LIST
    POSTEX_POS_LIST: LISTA COMPLETA (sin +N)
    """
    ws = wb.create_sheet("BLOQUES_DESTINOS")

    headers = [
        "BLOCK", "START_DAY", "START_TIME", "END_DAY", "END_TIME",
        "DESTINO", "POSTEX_N", "POSTEX_POS_SUBRAMPA", "POSTEX_POS_LIST"
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    bh_unique = bh.drop_duplicates(subset=["BLOCK"]).copy()
    bh_unique["BLOCK"] = bh_unique["BLOCK"].astype(str).str.strip().str.upper()

    r = 2
    for _, row in bh_unique.sort_values(["START_DAY", "START_TIME", "BLOCK"]).iterrows():
        block = str(row["BLOCK"]).strip().upper()
        sd = str(row["START_DAY"]).strip().upper()
        st = str(row["START_TIME"])
        ed = str(row["END_DAY"]).strip().upper()
        et = str(row["END_TIME"])

        destinos_map = postex_pos_by_block_dest.get(block, {})
        destinos_sorted = sorted(destinos_map.keys())

        if not destinos_sorted:
            ws.cell(row=r, column=1, value=block)
            ws.cell(row=r, column=2, value=sd)
            ws.cell(row=r, column=3, value=st)
            ws.cell(row=r, column=4, value=ed)
            ws.cell(row=r, column=5, value=et)
            ws.cell(row=r, column=6, value="")
            ws.cell(row=r, column=7, value=0)
            ws.cell(row=r, column=8, value="")
            ws.cell(row=r, column=9, value="")
            r += 1
            continue

        for destino in destinos_sorted:
            pairs = destinos_map.get(destino, set())
            pairs_sorted = sorted(pairs, key=lambda x: (ramp_sort_key(x[0]), x[1]))

            formatted = [f"{sub}-{slot:02d}" for sub, slot in pairs_sorted]
            n = len(formatted)
            pos_str = ", ".join(formatted)  # <-- COMPLETA

            subs_only = sorted({sub for sub, _ in pairs_sorted}, key=ramp_sort_key)
            subs_str = ", ".join(subs_only)

            ws.cell(row=r, column=1, value=block)
            ws.cell(row=r, column=2, value=sd)
            ws.cell(row=r, column=3, value=st)
            ws.cell(row=r, column=4, value=ed)
            ws.cell(row=r, column=5, value=et)
            ws.cell(row=r, column=6, value=destino)
            ws.cell(row=r, column=7, value=n)
            ws.cell(row=r, column=8, value=subs_str)

            cpos = ws.cell(row=r, column=9, value=pos_str)
            cpos.alignment = Alignment(wrap_text=True, vertical="top")

            r += 1

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 24
    ws.column_dimensions["I"].width = 90

    for rr in range(2, r):
        ws.row_dimensions[rr].height = 28


def export_visual_sheet(
    wb: Workbook,
    title: str,
    subramps: List[str],
    hour_minutes: List[int],
    cell_meta: Dict[Tuple[str, int], Dict[str, object]],
    color_by_block: Dict[str, str],
):
    ws = wb.create_sheet(title=title)

    header_row = 1
    ws.cell(row=header_row, column=1, value="Subrampa").font = Font(bold=True)
    for j, hm in enumerate(hour_minutes, start=2):
        c = ws.cell(row=header_row, column=j, value=minutes_to_day_hour_label(hm))
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Body
    for i, sub in enumerate(subramps, start=header_row + 1):
        ws.cell(row=i, column=1, value=sub).font = Font(bold=True)
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="left", vertical="center")

        for j, hm in enumerate(hour_minutes, start=2):
            key = (sub, hm)
            if key not in cell_meta:
                continue

            meta = cell_meta[key]
            blocks = meta["blocks"]
            full = str(meta.get("full", ""))     # agr_playa concatenada
            label = str(meta.get("label", ""))

            used = int(meta.get("used", 0))
            cap = int(meta.get("cap", 0))

            visible = f"{label} ({used}/{cap})" if cap > 0 else label
            cell = ws.cell(row=i, column=j, value=visible)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Comentario reducido (para que no se corte feo / sea legible)
            comment_text = full[:MAX_COMMENT] + (" ..." if len(full) > MAX_COMMENT else "")
            cell.comment = Comment(comment_text, AUTHOR)

            # Color por bloque (si solo hay 1)
            if len(blocks) == 1:
                hexrgb = color_by_block.get(blocks[0], "F2F2F2")
                cell.fill = PatternFill("solid", fgColor=hexrgb)
            else:
                cell.fill = PatternFill("solid", fgColor="E7E6E6")

    # Merge por tramos
    for i in range(header_row + 1, header_row + 1 + len(subramps)):
        run_start = None
        run_key = None
        sub = ws.cell(row=i, column=1).value

        for j in range(2, 2 + len(hour_minutes) + 1):
            if j <= 1 + len(hour_minutes):
                v = ws.cell(row=i, column=j).value
                hm = hour_minutes[j - 2]
                m = cell_meta.get((sub, hm))
                blocks = tuple(sorted(m["blocks"])) if (m and v) else ()
                cur_key = (v, blocks) if v else None
            else:
                cur_key = None

            if cur_key and cur_key == run_key:
                continue

            if run_key and run_start is not None:
                run_end = j - 1
                if run_end > run_start:
                    ws.merge_cells(start_row=i, start_column=run_start, end_row=i, end_column=run_end)
                    ws.cell(row=i, column=run_start).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if cur_key:
                run_start = j
                run_key = cur_key
            else:
                run_start = None
                run_key = None

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate

    ws.column_dimensions["A"].width = 12
    for j in range(2, 2 + len(hour_minutes)):
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.row_dimensions[header_row].height = 22
    for i in range(header_row + 1, header_row + 1 + len(subramps)):
        ws.row_dimensions[i].height = 26


def export_operativo_sheet(
    wb: Workbook,
    title: str,
    subramps: List[str],
    hour_minutes: List[int],
    cell_meta: Dict[Tuple[str, int], Dict[str, object]],
):
    ws = wb.create_sheet(title=title)

    header_row = 1
    ws.cell(row=header_row, column=1, value="Subrampa").font = Font(bold=True)
    for j, hm in enumerate(hour_minutes, start=2):
        ws.cell(row=header_row, column=j, value=minutes_to_day_hour_label(hm)).font = Font(bold=True)

    for i, sub in enumerate(subramps, start=header_row + 1):
        ws.cell(row=i, column=1, value=sub).font = Font(bold=True)
        for j, hm in enumerate(hour_minutes, start=2):
            key = (sub, hm)
            if key not in cell_meta:
                continue
            full = str(cell_meta[key].get("full", ""))
            if full:
                c = ws.cell(row=i, column=j, value=full)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate
    ws.column_dimensions["A"].width = 12
    for j in range(2, 2 + len(hour_minutes)):
        ws.column_dimensions[get_column_letter(j)].width = 20
    ws.row_dimensions[header_row].height = 22


# -----------------------------
# Main
# -----------------------------
def main():
    print("=== Generador GANTT 1H VISUAL (script general actualizado) ===")

    cap_map = load_capacity(CAPACITY_CSV)
    subramps = sorted(cap_map.keys(), key=ramp_sort_key)

    grupo = load_grupo(GRUPO_XLSX, GRUPO_SHEET)
    bh = load_bloques_horarios(BLOQUES_XLSX)
    intervals = build_schedule_intervals(bh)

    if not intervals:
        raise RuntimeError("No se han podido construir intervalos desde bloques_horarios.xlsx")

    all_blocks = set(intervals.keys())

    min_start = min(s for s, _ in intervals.values())
    max_end = max(e for _, e in intervals.values())
    hour_minutes = build_hour_grid(min_start, max_end)

    text_by_block, used_slots_by_block, postex_pos_by_block_dest, warnings = build_block_maps(grupo, all_blocks)
    color_by_block = assign_colors_by_day(bh)

    cell_meta = build_cell_meta(
        intervals=intervals,
        hour_minutes=hour_minutes,
        text_by_block=text_by_block,
        used_slots_by_block=used_slots_by_block,  # used SOLO POSTEX
        cap_map=cap_map,
    )

    wb = Workbook()
    wb.remove(wb.active)

    write_legend_sheet(wb, color_by_block=color_by_block, bh=bh)
    write_bloques_destinos_sheet(wb, bh=bh, postex_pos_by_block_dest=postex_pos_by_block_dest)

    export_visual_sheet(
        wb=wb,
        title="GANTT_VISUAL",
        subramps=subramps,
        hour_minutes=hour_minutes,
        cell_meta=cell_meta,
        color_by_block=color_by_block,
    )

    export_operativo_sheet(
        wb=wb,
        title="GANTT_OPERATIVO",
        subramps=subramps,
        hour_minutes=hour_minutes,
        cell_meta=cell_meta,
    )

    wb.save(OUTPUT_XLSX)
    print(f"✅ Generado: {OUTPUT_XLSX}")

    if warnings:
        print(f"⚠️ Warnings: {warnings} valores de 'Elemento' no parseables (ignorados).")


if __name__ == "__main__":
    main()