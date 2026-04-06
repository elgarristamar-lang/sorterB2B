#!/usr/bin/env python3
"""
Herramienta de configuración del Sorter VDL B2B
Procesa semanas especiales: asigna rampas y posiciones libres a destinos
que cambian de día de salida, respetando los horarios de bloque.

Admite dos formatos de GRUPO_DESTINOS:
  - Formato clásico (GRUPO_DESTINOS.xlsx)
  - Formato DXC export (resultadosConsulta*.xlsx): con columnas Estado, Secuencia, etc.

Uso:
    python process_parrilla.py <parrilla.xlsx> <grupo_destinos.xlsx> <ramp_capacity.csv> <sheet_name> [semana]

Ejemplo:
    python process_parrilla.py parrilla_de_salidas.xlsx resultadosConsulta.xlsx ramp_capacity.csv parrilla_test_s14 S14
"""

import sys, re, csv, os, shutil, tempfile, json
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

BLOQUE_RE  = re.compile(r'^(\d+BLO[A-Z]\d+)_(.+)$')
DAYS       = ['DOMINGO','LUNES','MARTES','MIERCOLES','MIÉRCOLES','JUEVES','VIERNES','SABADO']
E2_RE      = re.compile(r'^(MAN|EXDOCK|DOCK)', re.IGNORECASE)
SORTER_RE  = re.compile(r'^R\d+')

# Rampas reserved for manipulado — never use for automatic especial assignment
EXCLUDED_RAMPAS = {
    'R03A', 'R03B', 'R03C', 'R03D',   # Rampa 3: manipulado exclusivo
}

# Minutes-from-start-of-week (Sunday 00:00)
DAY_MIN = {'DOMINGO':0,'LUNES':1440,'MARTES':2880,'MIERCOLES':4320,
           'MIÉRCOLES':4320,'JUEVES':5760,'VIERNES':7200,'SABADO':8640}

# Cluster-letter → day prefix mapping (L=Lunes, M=Martes, X=Miercoles, J=Jueves, V=Viernes, S=Sabado, D=Domingo)
CLUSTER_DAY = {'D':'DOMINGO','L':'LUNES','M':'MARTES','X':'MIERCOLES',
               'J':'JUEVES','V':'VIERNES','S':'SABADO'}

def bloque_letra(bloque_str):
    """Extract letter from bloque code: '2BLOL1' → 'L', '4BLOX3' → 'X'"""
    import re as _re
    m = _re.match(r'^\d+BLO([A-Z])\d+$', str(bloque_str or ''))
    return m.group(1) if m else None


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def parse_gd_desc(desc):
    if not desc: return None, None, None
    core = re.sub(r'^\[B2B\]\s*', '', str(desc)).strip()
    core = re.sub(r'\s+PARA BAJAR POR.*$', '', core)
    core = re.sub(r'_CANCELADA.*$', '', core)
    m = BLOQUE_RE.match(core)
    if not m: return None, None, None
    bloque, rest = m.group(1), m.group(2)
    for d in DAYS:
        if rest.upper().startswith(d + '_'):
            return bloque, d.replace('MIÉRCOLES','MIERCOLES'), rest[len(d)+1:]
    return bloque, None, rest

def safe_str(val):
    if val is None: return ''
    s = str(val).strip()
    return '' if s.startswith('=') or s == '#N/A' else s

def parse_rampa(elemento):
    m = re.match(r'^R0*(\d+)_([A-Z])-(\d+)$', str(elemento or ''))
    if not m: return None, None
    return f"R{int(m.group(1)):02d}{m.group(2)}", int(m.group(3))

def postex_elem(rampa, pos):
    return f"R{rampa[1:-1]}_{rampa[-1]}-{pos:02d}"

def is_sorter_elem(elem):
    return bool(SORTER_RE.match(str(elem or '')))

def is_e2_elem(elem):
    return bool(E2_RE.match(str(elem or '')))

def timing_to_min(s):
    """'LUNES_16:00' → minutes from Sunday 00:00"""
    if not s or '_' not in s: return None
    day, time = s.split('_', 1)
    h, m = map(int, time.split(':'))
    return DAY_MIN.get(day.upper(), 0) + h * 60 + m

def bloques_overlap(b1, b2, bloque_timings):
    """True if the two bloques have overlapping [liberacion, desactivacion] windows."""
    t1, t2 = bloque_timings.get(b1), bloque_timings.get(b2)
    if not t1 or not t2: return True  # conservative
    l1, d1 = timing_to_min(t1['lib']), timing_to_min(t1['desac'])
    l2, d2 = timing_to_min(t2['lib']), timing_to_min(t2['desac'])
    if None in (l1, d1, l2, d2): return True
    return not (d1 <= l2 or d2 <= l1)

def resolve_bloque_for_new_day(id_cluster_str, new_dia, bloque_timings):
    """
    Given 'L4-J4' and new_dia='LUNES', return the matching bloque code '2BLOL4'.
    Parses each cluster token, finds the one whose day matches new_dia, returns its bloque.
    Falls back to the first token if no exact match.
    """
    if not id_cluster_str: return None
    # Build reverse: cluster_code → bloque
    cluster_to_bloque = {v['cluster']: k for k, v in bloque_timings.items()}
    
    tokens = [t.strip() for t in id_cluster_str.split('-') if t.strip()]
    # Find token matching new_dia
    new_dia_norm = new_dia.upper().replace('MIÉRCOLES','MIERCOLES')
    for token in tokens:
        if not token: continue
        letter = token[0].upper()
        token_day = CLUSTER_DAY.get(letter)
        if token_day == new_dia_norm:
            bloque = cluster_to_bloque.get(token)
            if bloque: return bloque
    # Fallback: first valid token
    for token in tokens:
        bloque = cluster_to_bloque.get(token.strip())
        if bloque: return bloque
    return None


# ─── LOADERS ──────────────────────────────────────────────────────────────────

def load_capacity(path):
    cap = {}
    with open(path, newline='') as f:
        reader = csv.reader(f, delimiter=';')
        next(reader)
        for row in reader:
            if len(row) >= 2 and row[0].strip():
                try: cap[row[0].strip()] = int(row[1].strip())
                except ValueError: pass
    return cap

def load_bloque_timings(parrilla_path):
    """Load Resumen Bloques sheet → {bloque: {lib, cutoff, desac, cluster}}
    Supports:
    - S14: sheet named 'Resumen Bloques' with cols [Bloque, Cluster, Liberacion, Cutoff, Desac]
    - S15: any other sheet whose first two columns look like bloque+cluster codes
    """
    wb = load_workbook(parrilla_path, read_only=True)
    timings = {}

    def _try_load(ws):
        result = {}
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return result
        # Detect header vs data rows: skip rows where col0 matches 'Bloque'/'BLOQUE'/etc.
        for r in rows:
            v0 = str(r[0] or '').strip()
            if not v0 or v0.lower() in ('bloque', 'block', '') or v0.startswith('='): continue
            # Detect S15 format: cols are Bloque, Cluster, Lib, Cutoff, Desac
            # Col indices may vary — find them by content pattern
            # Bloque looks like '1BLOD0', cluster like 'D0'
            if not re.match(r'^\d+BLO[A-Z]\d+$', v0): continue
            lib   = str(r[2] or '').strip() if len(r) > 2 else ''
            cutoff= str(r[3] or '').strip() if len(r) > 3 else ''
            desac = str(r[4] or '').strip() if len(r) > 4 else ''
            cluster = str(r[1] or '').strip() if len(r) > 1 else ''
            if lib:
                result[v0] = {'cluster': cluster, 'lib': lib, 'cutoff': cutoff, 'desac': desac}
        return result

    # Priority 1: named 'Resumen Bloques'
    if 'Resumen Bloques' in wb.sheetnames:
        timings = _try_load(wb['Resumen Bloques'])
        if timings: return timings

    # Priority 2: any sheet (except the main data sheet) that has bloque data
    main_sheets = set()
    for sh in wb.sheetnames:
        ws_tmp = wb[sh]
        hdr = next(ws_tmp.iter_rows(values_only=True, max_row=1), ())
        if any(str(h or '').strip().upper() == 'TIPO_SALIDA' for h in hdr):
            main_sheets.add(sh)
    for sh in wb.sheetnames:
        if sh in main_sheets: continue
        t = _try_load(wb[sh])
        if t:
            timings = t
            print(f"  Timings de bloques leídos de hoja alternativa: '{sh}'")
            break
    return timings

def load_grupo_destinos(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    file_header = all_rows[0]
    raw_data = all_rows[1:]

    hdrs = [str(h).strip() if h else '' for h in file_header]
    is_dxc = 'Estado' in hdrs or 'Secuencia' in hdrs

    if is_dxc:
        idx_grupo = hdrs.index('Grupo de destinos')
        idx_desc  = hdrs.index('Descripción Grupos de destino')
        idx_zona  = hdrs.index('Tipo de zona')
        idx_dest  = hdrs.index('Destino')
        idx_alm   = hdrs.index('Almacén')
        idx_elem  = hdrs.index('Elemento')
        idx_est   = hdrs.index('Estado') if 'Estado' in hdrs else None
        out_header = ('BLOQUE','Grupo de destinos','Descripción Grupos de destino',
                      'Tipo de zona','Destino','Almacén','Elemento')
    else:
        idx_grupo, idx_desc, idx_zona, idx_dest, idx_alm, idx_elem = 1, 2, 3, 4, 5, 6
        idx_est = None
        out_header = file_header

    print(f"  Formato: {'DXC export' if is_dxc else 'clásico'}")

    by_dia_playa = defaultdict(list)
    tagged = []

    for row in raw_data:
        grupo     = str(row[idx_grupo]) if row[idx_grupo] else ''
        desc      = str(row[idx_desc])  if row[idx_desc]  else ''
        tipo_zona = row[idx_zona] or ''
        destino   = str(row[idx_dest])  if row[idx_dest]  else ''
        almacen   = row[idx_alm]  or ''
        elemento  = str(row[idx_elem])  if row[idx_elem]  else ''
        estado    = str(row[idx_est])   if idx_est is not None and row[idx_est] else 'N/A'

        pb, pd, pp = parse_gd_desc(desc)
        entry = {
            'grupo': grupo, 'desc': desc, 'tipo_zona': tipo_zona,
            'destino': destino, 'almacen': almacen, 'elemento': elemento,
            'bloque': pb, 'dia': pd, 'playa': pp, 'estado': estado,
            'is_weekly': bool(pb and pd and pp),
            'is_sorter': is_sorter_elem(elemento),
            'is_e2': is_e2_elem(elemento),
            '_raw': (None, grupo, desc, tipo_zona, destino, almacen, elemento),
        }
        tagged.append(entry)
        if pb and pd and pp:
            by_dia_playa[(pd, pp)].append(entry)

    return out_header, tagged, by_dia_playa

def load_parrilla(path, sheet_name):
    wb = load_workbook(path, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Hoja '{sheet_name}' no encontrada. Disponibles: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    col = {str(h).strip().upper(): i for i, h in enumerate(all_rows[0]) if h}
    records = []
    for row in all_rows[1:]:
        def g(c, d=''):
            i = col.get(c)
            return safe_str(row[i]) if i is not None and row[i] is not None else d
        tipo_sal = g('TIPO_SALIDA').upper()
        dpn  = g('DIA_PLAYA_NEW')
        dpo  = g('DIA_PLAYA_ORIGINAL')
        # Extract playa: for especiales use DIA_PLAYA_ORIGINAL (GD is keyed on orig day)
        # For cancelladas and regulares use DIA_PLAYA_NEW / PLAYA col
        playa = g('PLAYA') or g('AGRUPACION_PLAYA')
        if not playa:
            src_field = dpo if tipo_sal == 'ESPECIAL DIA CAMBIO' and dpo else dpn
            if src_field:
                _su = src_field.upper()
                if _su.startswith('CANCELADA_') or _su.startswith('CANCELADO_'):
                    playa = src_field[src_field.index('_')+1:].strip()
                else:
                    _dm = re.match(
                        r'^(?:DOMINGO|LUNES|MARTES|MIERCOLES|MIÉRCOLES|JUEVES|VIERNES|SABADO)_(.+)$',
                        src_field, re.IGNORECASE)
                    playa = _dm.group(1).strip() if _dm else src_field
        # Derive dia_salida from DIA_PLAYA_ORIGINAL when DIA_SALIDA_ORIGINAL missing
        dia_salida = g('DIA_SALIDA') or g('DIA_SALIDA_ORIGINAL')
        if not dia_salida and dpo:
            _ddm = re.match(
                r'^(DOMINGO|LUNES|MARTES|MIERCOLES|MIÉRCOLES|JUEVES|VIERNES|SABADO)_',
                dpo, re.IGNORECASE)
            if _ddm: dia_salida = _ddm.group(1).upper()
        r = {'dia_playa': g('DIA_PLAYA') or dpn,
             'playa': playa,
             'dia_salida': dia_salida,
             'cutoff': g('CUTOFF') or g('CUTOFF_NEW'),
             'dia_salida_new': g('DIA_SALIDA_NEW'),
             'bloque': g('BLOQUE'),
             'nomenclatura': g('NOMENCLATURA'),
             'tipo_salida': tipo_sal,
             'id_cluster': g('ID_CLUSTER'),
             'id_cluster_new': g('ID_CLUSTER_NEW'),
             'mantener_original': g('MANTENER_ORIGINAL','').strip().upper() == 'SI'}
        if r['playa'] and r['tipo_salida']:
            records.append(r)
    return records


# ─── RAMPA ASSIGNMENT ─────────────────────────────────────────────────────────

def build_day_occ(tagged, dia, new_bloque, bloque_timings,
                   exclude_playas=None, run_occ=None):
    """
    Build {rampa: {pos: grupo}} for dia, merging:
    - positions from tagged GD that overlap in time with new_bloque
      (regardless of which day the entry is tagged to)
    - positions already assigned in this run (run_occ) — always conflicting
    """
    ex = exclude_playas or set()
    occ = defaultdict(dict)
    for e in tagged:
        if e['tipo_zona'] != 'POSTEX': continue
        if e['playa'] in ex or not e['is_sorter']: continue
        existing_bloque = e['bloque'] or ''
        # Skip if timing is known and does NOT overlap with new_bloque
        if new_bloque and existing_bloque and new_bloque in bloque_timings and existing_bloque in bloque_timings:
            if not bloques_overlap(new_bloque, existing_bloque, bloque_timings):
                continue
        # If no timing info available, only include entries from the same day (conservative)
        elif e['dia'] != dia:
            continue
        r, p = parse_rampa(e['elemento'])
        if r and p:
            occ[r][p] = e['grupo']
    # Always add within-run assignments — no timing bypass
    for rampa, pos_map in (run_occ or {}).items():
        for pos, grupo in pos_map.items():
            occ[rampa][pos] = grupo
    return occ



def load_superplaya(path) -> dict:
    """Load superplaya.xlsx → {playa: superplaya_group}"""
    if not path or not os.path.exists(str(path)):
        return {}
    from openpyxl import load_workbook as _lwb
    wb = _lwb(str(path), read_only=True)
    ws = wb.active
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[1] and str(row[0]).strip() != 'AGRUPACION_PLAYA':
            mapping[str(row[0]).strip().upper()] = str(row[1]).strip().upper()
    return mapping

def lexical_prefix(playa: str) -> str:
    """Extract family prefix: ESPANA_LEGANES_EXT → ESPANA_LEGANES"""
    p = playa.upper()
    p = re.sub(r'_\d+$', '', p)
    for sfx in ['_TSA_EXT','_TSA','_EXT','_CPT_OL','_CPT','_AIR','_OL','_CIP']:
        if p.endswith(sfx):
            p = p[:-len(sfx)]
            break
    return p


def _ramp_number(r: str) -> int:
    """Extract numeric part: R11C → 11"""
    import re as _re
    m = _re.match(r'^R(\d+)', r)
    return int(m.group(1)) if m else 99

def _ramp_group(r: str) -> str:
    """'par' if RXX is even, 'impar' if odd"""
    n = _ramp_number(r)
    return 'par' if n % 2 == 0 else 'impar'

def _ramp_proximity_key(r: str, anchor_numbers: set) -> int:
    """Distance to nearest already-used rampa number in the superplaya."""
    n = _ramp_number(r)
    if not anchor_numbers:
        return n  # no anchor yet → sort by number
    return min(abs(n - a) for a in anchor_numbers)

def find_free_slots(occ, capacity, n_needed,
                    preferred_rampas=None, committed_group=None, anchor_numbers=None,
                    full_occ=None):
    """
    Asigna n_needed slots libres siguiendo esta prioridad:

    1. RAMPAS HERMANAS  — mismas rampas que usó el último miembro de la superplaya
    2. PAR-PAREJA       — A↔B o C↔D de las rampas hermanas (mismo par físico)
    3. POOL GENERAL     — ordenado por:
         a. Grupo (par/impar) comprometido para la superplaya
         b. Rampas VACÍAS primero (sin bloques solapantes del GD estándar)
            - Con anchor: las más cercanas numéricamente al anchor
            - Sin anchor: las de número más bajo (rampas "del centro")
         c. Rampas con par-pareja también libre (mejor para la siguiente playa)
         d. Más slots libres en total (rampa + pareja)
    """
    def get_free(r):
        if r in EXCLUDED_RAMPAS: return []
        # Only block time-conflicting slots (full_occ is used for is_empty scoring only)
        used = set(occ.get(r, {}).keys())
        return sorted(p for p in range(1, capacity.get(r, 0) + 1) if p not in used)

    def pair_of(r):
        m = {'A':'B','B':'A','C':'D','D':'C'}.get(r[-1] if r else '')
        return r[:-1] + m if m else None

    assigned, rem = [], n_needed

    # Paso 1: rampas hermanas (mismo grupo, ya usadas por sibling anterior)
    # Only use preferred rampas if together they have enough slots for n_needed.
    # If not, skip them so the playa lands on a single rampa with enough space.
    filtered_pref = [r for r in (preferred_rampas or [])
                     if committed_group is None or _ramp_group(r) == committed_group]
    pref_total_free = sum(len(get_free(r)) for r in filtered_pref)
    if pref_total_free >= n_needed:
        for r in sorted(filtered_pref, key=_ramp_number):
            if rem <= 0: break
            free = get_free(r)
            if free:
                assigned.extend((r, p) for p in free[:rem]); rem -= len(free[:rem])

    # Paso 2: par-pareja de las hermanas (only if paso 1 was used)
    pair_pool = set()
    if pref_total_free >= n_needed:
        pair_pool = {pair_of(r) for r in filtered_pref if pair_of(r) and pair_of(r) not in filtered_pref
                     and (committed_group is None or _ramp_group(pair_of(r)) == committed_group)}
        for r in sorted(pair_pool, key=_ramp_number):
            if rem <= 0: break
            free = get_free(r)
            if free:
                assigned.extend((r, p) for p in free[:rem]); rem -= len(free[:rem])

    if rem <= 0:
        return assigned, rem

    # Paso 3: pool general ordenado
    tried = (set(preferred_rampas or [])) | pair_pool
    _anchor = anchor_numbers or set()

    def sort_key(r):
        free = get_free(r)
        if not free: return (999,) * 6
        # Prefer rampas that have enough slots for the whole playa (avoid splitting)
        has_enough = len(free) >= rem
        _any_occ = full_occ if full_occ is not None else occ
        is_empty   = len(_any_occ.get(r, {})) == 0
        in_group   = committed_group is None or _ramp_group(r) == committed_group
        prox       = _ramp_proximity_key(r, _anchor) if _anchor else _ramp_number(r)
        mate       = pair_of(r)
        mate_free  = len(get_free(mate)) if mate and mate not in EXCLUDED_RAMPAS else 0
        eff_prox = prox if (is_empty and _anchor) else (-len(free) if is_empty else 1000 + prox)
        # has_enough=True sorts first (0 < 1)
        return (-int(in_group), -int(has_enough), eff_prox, -int(mate_free > 0), -(len(free) + mate_free))

    pool = sorted([r for r in capacity if r not in EXCLUDED_RAMPAS and r not in tried
                   and get_free(r)], key=sort_key)

    for r in pool:
        if rem <= 0: break
        free = get_free(r)
        assigned.extend((r, p) for p in free[:rem]); rem -= len(free[:rem])

    return assigned, rem

def playa_is_e2(playa, by_dia_playa):
    """
    Check if a playa is E2-only (all its elements across ALL days are MAN/EXDOCK).
    Returns True only if at least one entry exists and ALL of them are E2 elements.
    """
    all_entries = [e for entries in by_dia_playa.values()
                   for e in entries if e['playa'] == playa]
    if not all_entries:
        return False
    return all(e['is_e2'] or not e['elemento'] for e in all_entries)

def find_best_source_day(playa, orig_dia, by_dia_playa):
    """Find the day with most sorter POSTEX entries for this playa."""
    candidates = []
    for (d, p), entries in by_dia_playa.items():
        if p != playa: continue
        postex_sorter = [e for e in entries if e['tipo_zona'] == 'POSTEX' and e['is_sorter']]
        if postex_sorter:
            candidates.append((d, postex_sorter))
    if not candidates: return None, []
    day_order = ['DOMINGO','LUNES','MARTES','MIERCOLES','JUEVES','VIERNES','SABADO']
    orig_idx = day_order.index(orig_dia) if orig_dia in day_order else 0
    candidates.sort(key=lambda x: (abs(day_order.index(x[0])-orig_idx) if x[0] in day_order else 99, -len(x[1])))
    return candidates[0]

def get_slot_structure(postex_entries, sorexp_entries):
    """
    Group POSTEX entries by their shared physical (rampa, pos) slot.
    Multiple destinos can share the same slot — e.g. MEXICO packs 63 store codes
    into a single sorter position. The slot is the true unit of assignment.
    Returns: slots list, sorexp_by_rampa dict, alm_postex str, alm_sorexp str.
    """
    pos_dests   = defaultdict(list)
    rampa_dests = defaultdict(list)
    for e in postex_entries:
        r, p = parse_rampa(e['elemento'])
        if r and p:
            pos_dests[(r, p)].append(e['destino'])
    for e in sorexp_entries:
        elem = e['elemento']
        if re.match(r'^R\d+[A-Z]$', elem):
            rampa_dests[elem].append(e['destino'])
    slots = [{'dests': v, 'orig_rampa': k[0], 'orig_pos': k[1]}
             for k, v in sorted(pos_dests.items())]
    alm_p = postex_entries[0]['almacen'] if postex_entries else 'CONRAM'
    alm_s = sorexp_entries[0]['almacen'] if sorexp_entries else 'SOREXP'
    return slots, dict(rampa_dests), alm_p, alm_s


def assign_especial(orig_dia, orig_playa, new_dia, raw_bloque, id_cluster,
                    by_dia_playa, tagged, capacity, bloque_timings, freed_in_new_dia,
                    run_occ_new_dia=None, preferred_rampas=None, all_especial_playas=None,
                    committed_group=None, anchor_numbers=None):
    """
    Assign slots from (orig_dia, orig_playa) to free sorter positions in new_dia,
    respecting bloque timing and preserving the slot→destinos structure.

    The assignment unit is a physical SLOT (rampa, pos) that may hold multiple
    destinos — e.g. MEXICO packs 63 store codes per slot. We find N free slots
    (where N = number of unique positions in the original config) and replicate
    the same destino grouping onto each new slot.
    """
    if playa_is_e2(orig_playa, by_dia_playa):
        return [], {'status':'E2_ROUTE','playa':orig_playa,'dia_orig':orig_dia,'dia_new':new_dia,
                    'msg':'Ruta E2/manual (MAN/EXDOCK) — no pasa por rampas del sorter'}

    new_bloque = None
    if raw_bloque and raw_bloque not in ('#N/A','NO_BLOQUE','?',''):
        new_bloque = raw_bloque
    if not new_bloque:
        new_bloque = resolve_bloque_for_new_day(id_cluster, new_dia, bloque_timings)

    orig        = by_dia_playa.get((orig_dia, orig_playa), [])
    postex_orig = [e for e in orig if e['tipo_zona'] == 'POSTEX' and e['is_sorter']]
    sorexp_orig = [e for e in orig if e['tipo_zona'] == 'SOREXP' and e['is_sorter']]
    source_day  = orig_dia

    if not postex_orig:
        best_day, fallback = find_best_source_day(orig_playa, orig_dia, by_dia_playa)
        if not best_day:
            all_p  = [e for v in by_dia_playa.values() for e in v if e['playa'] == orig_playa]
            status = 'E2_ROUTE' if all_p else 'NO_CONFIG'
            msg    = ('Ruta E2/manual — elementos no son rampas estándar' if all_p
                      else 'Sin configuración GD en ningún día')
            return [], {'status':status,'playa':orig_playa,'dia_orig':orig_dia,'dia_new':new_dia,'msg':msg}
        postex_orig = fallback
        sorexp_orig = [e for e in by_dia_playa.get((best_day, orig_playa),[])
                       if e['tipo_zona'] == 'SOREXP' and e['is_sorter']]
        source_day  = best_day

    # Build slot structure: N unique (rampa,pos) → list of destinos each
    slots, _, alm_postex, alm_sorexp = get_slot_structure(postex_orig, sorexp_orig)
    n_slots = len(slots)

    # Exclude all especial playas: they're being moved so their positions are freed
    freed    = (all_especial_playas or set()) | freed_in_new_dia
    occ      = build_day_occ(tagged, new_dia, new_bloque, bloque_timings,
                             exclude_playas=freed, run_occ=run_occ_new_dia)
    # full_occ: entries from blocks that overlap OR share the same physical day.
    # Blocks sharing same day = same day-letter prefix (e.g. M for MARTES, X for MIERCOLES).
    # This prevents especiales from taking slots used by same-day standard blocks,
    # even if they don't conflict timewise. Falls back to occ if no bloque.
    import re as _re_fo
    if new_bloque:
        _day_match = _re_fo.match(r'\d+BLO([A-Z])\d+', new_bloque or '')
        _day_letters = {_day_match.group(1)} if _day_match else set()
        # Also include adjacent overlapping day letter (end day of block)
        for _bk in bloque_timings:
            _bm = _re_fo.match(r'\d+BLO([A-Z])\d+', _bk)
            if _bm and bloques_overlap(new_bloque, _bk, bloque_timings):
                _day_letters.add(_bm.group(1))
        # Build full_occ for same-day blocks
        _full_occ_map = defaultdict(dict)
        for _e in tagged:
            if _e['tipo_zona'] != 'POSTEX' or _e['playa'] in freed or not _e['is_sorter']: continue
            _eb = _e.get('bloque','') or ''
            _ebm = _re_fo.match(r'\d+BLO([A-Z])\d+', _eb)
            if _ebm and _ebm.group(1) in _day_letters:
                _r, _p = parse_rampa(_e['elemento'])
                if _r and _p and _r not in EXCLUDED_RAMPAS:
                    _full_occ_map[_r][_p] = _eb
        full_occ = dict(_full_occ_map)
    else:
        full_occ = occ
    assigned, unmet = find_free_slots(occ, capacity, n_slots, preferred_rampas=preferred_rampas, committed_group=committed_group, anchor_numbers=anchor_numbers, full_occ=full_occ)

    playa_tag = orig_playa.replace('ESPANA_','')[:6].replace('_','')
    new_grupo = f"ESP_{new_dia[:3]}_{playa_tag}"[:18]
    fb_note   = f" [datos {source_day}]" if source_day != orig_dia else ""
    new_desc  = f"[B2B] {new_bloque or '?'}_{new_dia}_{orig_playa} (ESPECIAL{fb_note})"

    rows_out, rampas_used = [], defaultdict(list)
    n_destinos_total = 0
    for slot, (new_rampa, new_pos) in zip(slots, assigned):
        for destino in slot['dests']:
            rows_out.append((None, new_grupo, new_desc, 'POSTEX',
                             destino, alm_postex, postex_elem(new_rampa, new_pos)))
            rows_out.append((None, new_grupo, new_desc, 'SOREXP',
                             destino, alm_sorexp, new_rampa))
        rampas_used[new_rampa].append(new_pos)
        n_destinos_total += len(slot['dests'])

    return rows_out, {
        'status':      'OK' if unmet == 0 else 'PARTIAL',
        'playa':        orig_playa,
        'dia_orig':     orig_dia,
        'dia_new':      new_dia,
        'bloque_new':   new_bloque or '?',
        'n_slots':      n_slots,
        'n_destinos':   n_destinos_total,
        'n_assigned':   n_slots - unmet,
        'n_unmet':      unmet,
        'grupo_new':    new_grupo,
        'desc_new':     new_desc,
        'source_day':   source_day,
        'rampas':       {r: sorted(p) for r, p in rampas_used.items()},
        'n_rows':       len(rows_out),
    }


# ─── PROCESS ──────────────────────────────────────────────────────────────────


def load_especial_bloque_map(parrilla_path):
    """
    Read especial bloque assignments from the parrilla workbook.
    Supports two formats:
    - S14: reads 'SEMANA SANTA W*' sheet (has BLOQUE column directly)
    - S18: reads parrilla_test_* sheet (derives bloque from ID_CLUSTER_NEW + Resumen Bloques)
    Returns {(dia_new_upper, playa): bloque}
    """
    from openpyxl import load_workbook as _lw
    import re as _re
    result = {}
    _DAY_PFX = _re.compile(
        r'^(DOMINGO|LUNES|MARTES|MIERCOLES|JUEVES|VIERNES|SABADO)_', _re.IGNORECASE)
    try:
        wb = _lw(str(parrilla_path), read_only=True)

        # --- Try S14 format: SEMANA SANTA sheet with BLOQUE column ---
        ss_sheet = next((s for s in wb.sheetnames
                         if 'SEMANA SANTA' in s.upper() or 'SEMANA_SANTA' in s.upper()), None)
        if ss_sheet:
            ws = wb[ss_sheet]
            rows = list(ws.iter_rows(values_only=True))
            col = {str(h).strip().upper(): i for i, h in enumerate(rows[0]) if h}
            for r in rows[1:]:
                bloque = r[col.get('BLOQUE', -1)] if 'BLOQUE' in col else None
                if not bloque or str(bloque) in ('None', '#N/A', 'NO_BLOQUE'): continue
                dia_new = str(r[col.get('DIA_SALIDA_NEW', 3)] or '')
                playa_field = str(r[col.get('DIA_PLAYA_NEW', 1)] or '')
                tipo = str(r[col.get('TIPO_SALIDA', 12)] or '')
                if 'CANCELADA' in tipo: continue
                m = _DAY_PFX.match(playa_field)
                if m and dia_new:
                    result[(dia_new.upper(), playa_field[m.end():].strip())] = str(bloque)
            if result:
                return result

        # --- S15/S18 format: derive bloque from the cluster part that CHANGED ---
        # ID_CLUSTER_NEW = ID_CLUSTER with the old day's cluster replaced by new day's
        # The new special block is the part in ID_CLUSTER_NEW not present in ID_CLUSTER
        _cb = {}  # cluster_code → bloque_name
        # Find the bloques sheet: prefer 'Resumen Bloques', else any sheet with bloque codes
        _bloques_sheet = None
        if 'Resumen Bloques' in wb.sheetnames:
            _bloques_sheet = 'Resumen Bloques'
        else:
            for _sh in wb.sheetnames:
                _ws_b = wb[_sh]
                _hdr_b = next(_ws_b.iter_rows(values_only=True, max_row=1), ())
                _hdr_b_up = [str(h or '').strip().upper() for h in _hdr_b]
                if any('BLOQUE' in h or 'CLUSTER' in h for h in _hdr_b_up):
                    # Verify it has actual bloque data (1BLO... pattern)
                    _sample = list(_ws_b.iter_rows(values_only=True, max_row=5))
                    if any(_r and _r[0] and re.match(r'^\d+BLO', str(_r[0] or '')) for _r in _sample[1:]):
                        _bloques_sheet = _sh
                        break
        if _bloques_sheet:
            for _r in wb[_bloques_sheet].iter_rows(values_only=True):
                v0 = str(_r[0] or '').strip()
                v1 = str(_r[1] or '').strip()
                if v0 and v1 and v0.lower() not in ('bloque', 'block', '') and not v0.startswith('='):
                    _cb[v1.upper()] = v0  # cluster → bloque

        def _get_new_cluster(id_cluster, id_cluster_new):
            """The new special block = the part in ID_CLUSTER_NEW not in ID_CLUSTER."""
            old_parts = set(str(id_cluster or '').upper().split('-'))
            new_parts = str(id_cluster_new or '').upper().split('-')
            changed = [p for p in new_parts if p not in old_parts]
            return changed[0] if changed else (new_parts[-1] if new_parts else '')

        # Find parrilla sheet (has TIPO_SALIDA + DIA_PLAYA_NEW)
        par_sheet = next((s for s in wb.sheetnames
                          if s.lower().startswith('parrilla_test')), None)
        if not par_sheet:
            par_sheet = next((s for s in wb.sheetnames
                              if any(str(h or '').strip().upper() == 'TIPO_SALIDA'
                                     for h in next(wb[s].iter_rows(values_only=True, max_row=1), []))), None)
        if par_sheet:
            ws2 = wb[par_sheet]
            rows2 = list(ws2.iter_rows(values_only=True))
            col2 = {str(h).strip().upper(): i for i, h in enumerate(rows2[0]) if h}
            for r in rows2[1:]:
                tipo = str(r[col2.get('TIPO_SALIDA', 10)] or '').strip().upper()
                if 'ESPECIAL DIA CAMBIO' not in tipo: continue
                if 'CANCELADA' in tipo: continue
                dia_new        = str(r[col2.get('DIA_SALIDA_NEW', 4)] or '').strip().upper()
                dpn            = str(r[col2.get('DIA_PLAYA_NEW', 1)] or '').strip()
                id_cluster     = str(r[col2.get('ID_CLUSTER', 8)] or '').strip()
                id_cluster_new = str(r[col2.get('ID_CLUSTER_NEW', 9)] or '').strip()
                dpo2 = str(r[col2.get('DIA_PLAYA_ORIGINAL', 2)] or '').strip()
                # For especiales: extract playa from DIA_PLAYA_ORIGINAL (authoritative)
                if dpo2:
                    _mo = _DAY_PFX.match(dpo2)
                    playa = dpo2[_mo.end():].strip() if _mo else ''
                if not playa:
                    m = _DAY_PFX.match(dpn)
                    playa = dpn[m.end():].strip() if m else str(r[col2.get('AGRUPACION_PLAYA', 3)] or '').strip()
                if not playa or not dia_new: continue
                # New block = the cluster part that changed (not in original ID_CLUSTER)
                new_cluster = _get_new_cluster(id_cluster, id_cluster_new)
                bloque = _cb.get(new_cluster.upper(), '')
                if bloque:
                    result[(dia_new, playa)] = bloque
    except Exception as e:
        print(f"Warning: could not load especial bloque map: {e}")
    return result


def load_cancelled_especiales(parrilla_path):
    """
    Return set of playa names with tipo=CANCELADA.
    Supports S14 (SEMANA SANTA sheet) and S15 (unified Hoja1).
    Also handles CANCELADO_ (masculine) prefix variant.
    """
    from openpyxl import load_workbook as _lwb_c
    cancelled = set()
    try:
        wb = _lwb_c(parrilla_path, read_only=True)
        # Find best sheet: prefer SEMANA SANTA, else any sheet with TIPO_SALIDA
        sheet = next((s for s in wb.sheetnames
                      if re.search(r'SEMANA.SANTA', s, re.IGNORECASE)), None)
        if not sheet:
            sheet = next((s for s in wb.sheetnames
                          if any(str(h or '').strip().upper() == 'TIPO_SALIDA'
                                 for h in next(wb[s].iter_rows(values_only=True, max_row=1), []))), None)
        if not sheet:
            return cancelled
        rows = list(wb[sheet].iter_rows(values_only=True))
        hdr = {str(h).strip().upper(): i for i, h in enumerate(rows[0]) if h}
        for r in rows[1:]:
            if str(r[hdr.get('TIPO_SALIDA', 99)] or '').strip().upper() != 'CANCELADA':
                continue
            dpn = str(r[hdr.get('DIA_PLAYA_NEW', 1)] or '').strip()
            dpo = str(r[hdr.get('DIA_PLAYA_ORIGINAL', 2)] or '').strip()
            su  = dpn.upper()
            if su.startswith('CANCELADA_') or su.startswith('CANCELADO_'):
                cancelled.add(dpn[dpn.index('_')+1:].strip().upper())
            elif dpo:
                # Extract from DIA_PLAYA_ORIGINAL: "LUNES_ESPANA_X" → "ESPANA_X"
                _m = re.match(
                    r'^(?:DOMINGO|LUNES|MARTES|MIERCOLES|JUEVES|VIERNES|SABADO)_(.+)$',
                    dpo, re.IGNORECASE)
                if _m: cancelled.add(_m.group(1).strip().upper())
            else:
                _m = re.match(r'(?:\w+)_(.+)', dpn)
                if _m: cancelled.add(_m.group(1).strip().upper())
    except Exception:
        pass
    return cancelled


def process(parrilla_records, tagged, by_dia_playa, capacity, bloque_timings, filter_days=None, superplaya_map=None, especial_bloque_map=None, cancelled_especiales=None):
    canceladas, especiales, habituales = {}, {}, {}

    for r in parrilla_records:
        playa, tipo = r['playa'], r['tipo_salida']
        dia_orig = r['dia_salida']
        dia_new  = r['dia_salida_new'] or dia_orig
        if tipo == 'CANCELADA':
            canceladas[(dia_orig, playa)] = r
        elif tipo == 'ESPECIAL DIA CAMBIO' and dia_new and dia_new != dia_orig:
            especiales[(dia_orig, playa)] = (dia_new, r)
        elif tipo in ('HABITUAL','REGULAR','ESPECIAL CUTOFF'):
            habituales[(dia_orig, playa)] = r

    freed_per_day = defaultdict(set)
    mantener_set: set = set()  # (dia_orig, playa) → keep original entries
    for (dia, playa) in canceladas: freed_per_day[dia].add(playa)
    for (dia_orig, playa), (dia_new, r) in especiales.items():
        if r.get('mantener_original'):
            mantener_set.add((dia_orig, playa))
        else:
            freed_per_day[dia_orig].add(playa)

    output_rows = []
    rem_cancel = rem_moved = kept = kept_nw = 0

    for entry in tagged:
        if not entry['is_weekly']:
            output_rows.append(entry['_raw']); kept_nw += 1; continue
        key = (entry['dia'], entry['playa'])
        if key in canceladas: rem_cancel += 1; continue
        if key in especiales and key not in mantener_set:
            rem_moved += 1; continue
        elif key in especiales and key in mantener_set:
            pass  # MANTENER_ORIGINAL=SI → keep original AND add new-day entries
        # Bloque filter: skip entries whose bloque letter is not selected
        if filter_days:
            _letra = bloque_letra(entry.get('bloque',''))
            if _letra and _letra not in filter_days: continue
            elif not _letra and entry['dia'] not in [CLUSTER_DAY.get(l,l) for l in filter_days]: continue
        output_rows.append(entry['_raw']); kept += 1

    # Accumulate slots assigned in this run per new_dia so subsequent
    # playas don't collide (e.g. MEXICO + MEXICO_2 both going to MIERCOLES)
    run_occ = defaultdict(dict)   # dia -> {rampa: {pos: grupo}}

    results, added = [], 0
    especial_rows = []
    _last_rampas_by_prefix = {}   # (dia_new, prefix) → set of rampas used by last sibling
    _committed_group = {}         # (dia_new, prefix) → 'par' or 'impar'
    _anchor_numbers  = {}         # (dia_new, prefix) → set of rampa numbers used so far
    all_esp_playas = {playa for (_, playa) in especiales}  # all playas being moved (any day)
    # Sort especiales so same-family playas (shared lexical prefix) are consecutive
    # This lets siblings share rampas naturally via run_occ
    def _esp_sort_key(item):
        (dia_orig, playa), (dia_new, record) = item
        sp = (superplaya_map or {}).get(playa.upper(), lexical_prefix(playa))
        return (dia_new, record.get('bloque',''), sp, playa)
    especiales_sorted = sorted(especiales.items(), key=_esp_sort_key)

    for (dia_orig, playa), (dia_new, record) in especiales_sorted:
        # Bloque filter: skip especiales whose bloque letter is not selected
        if filter_days:
            _blq = record.get('bloque','')
            _letra_esp = bloque_letra(_blq)
            if not _letra_esp or _letra_esp not in filter_days: continue
        # Preferred rampas: rampas already used by the last sibling (same prefix, same new day)
        prefix = (superplaya_map or {}).get(playa.upper(), lexical_prefix(playa))
        pref_rampas = _last_rampas_by_prefix.get((dia_new, prefix))

        cg  = _committed_group.get((dia_new, prefix))
        an  = _anchor_numbers.get((dia_new, prefix))

        # Resolve the correct bloque using especial_bloque_map first (same logic as assign_especial)
        _raw_bloque_presel = record.get('bloque', '')
        _is_na_presel = not _raw_bloque_presel or str(_raw_bloque_presel).strip().upper() in ('#N/A','N/A','NONE','NO_BLOQUE','')
        if _is_na_presel and especial_bloque_map:
            _raw_bloque_presel = especial_bloque_map.get((dia_new.upper(), playa), _raw_bloque_presel)
        _presel_bloque = (_raw_bloque_presel if _raw_bloque_presel and str(_raw_bloque_presel).strip().upper() not in ('#N/A','N/A','NONE','NO_BLOQUE','')
                         else resolve_bloque_for_new_day(record.get('id_cluster',''), dia_new, bloque_timings) or '')

        # If no committed group yet, pre-select the group with most available slots
        if cg is None:
            occ_snap = build_day_occ(
                tagged, dia_new,
                _presel_bloque,
                bloque_timings, exclude_playas=all_esp_playas,
                run_occ=run_occ.get(dia_new, {}))
            par_free_n = sum(
                len([p for p in range(1, capacity.get(r,0)+1) if p not in occ_snap.get(r,{})])
                for r in capacity if r not in EXCLUDED_RAMPAS and _ramp_number(r) % 2 == 0)
            imp_free_n = sum(
                len([p for p in range(1, capacity.get(r,0)+1) if p not in occ_snap.get(r,{})])
                for r in capacity if r not in EXCLUDED_RAMPAS and _ramp_number(r) % 2 != 0)
            cg = 'par' if par_free_n >= imp_free_n else 'impar'

        # Resolve correct bloque: parrilla BLOQUE if set, else SEMANA SANTA map, else auto-resolve
        _raw_bloque = record.get('bloque', '')
        _is_na = not _raw_bloque or str(_raw_bloque).strip().upper() in ('#N/A','N/A','NONE','NO_BLOQUE','')
        if _is_na and especial_bloque_map:
            _raw_bloque = especial_bloque_map.get((dia_new.upper(), playa), _raw_bloque)
        # S18: _raw_bloque may be "2BLOL0,4BLOX3" — run assign_especial for each bloque
        _raw_bloques = [b.strip() for b in str(_raw_bloque).split(',') if b.strip()]
        if not _raw_bloques:
            _raw_bloques = ['']
        new_rows, info = [], {'status': 'NO_CONFIG', 'rampas': {}, 'playa': playa, 'dia_orig': dia_orig, 'dia_new': dia_new, 'n_assigned': 0, 'n_destinos': 0, 'msg': 'sin datos en ningún día'}
        for _single_bloque in _raw_bloques:
            _rows_b, _info_b = assign_especial(
                dia_orig, playa, dia_new,
                _single_bloque, record['id_cluster'],
                by_dia_playa, tagged, capacity, bloque_timings,
                freed_per_day.get(dia_new, set()),
                run_occ.get(dia_new, {}),
                preferred_rampas=pref_rampas,
                all_especial_playas=all_esp_playas,
                committed_group=cg,
                anchor_numbers=an,
            )
            new_rows.extend(_rows_b)
            if _info_b.get('status') == 'OK':
                info = _info_b  # keep last OK info for tracking
            elif _info_b.get('status') in ('E2_ROUTE', 'NO_CONFIG') and info.get('status') == 'NO_CONFIG':
                info = _info_b  # propagate E2_ROUTE/NO_CONFIG if nothing better yet
        # Update sibling proximity tracking
        if info.get('status') == 'OK' and info.get('rampas'):
            used_rampas = set(info['rampas'].keys())
            _last_rampas_by_prefix[(dia_new, prefix)] = used_rampas
            # Determine/lock committed group (majority vote on first assignment)
            if (dia_new, prefix) not in _committed_group:
                par_count   = sum(1 for r in used_rampas if _ramp_group(r) == 'par')
                impar_count = len(used_rampas) - par_count
                _committed_group[(dia_new, prefix)] = 'par' if par_count >= impar_count else 'impar'
            # Update anchor numbers
            existing = _anchor_numbers.get((dia_new, prefix), set())
            _anchor_numbers[(dia_new, prefix)] = existing | {_ramp_number(r) for r in used_rampas}

        # Register newly assigned positions so next playa in same run sees them
        for row in new_rows:
            if row[3] == 'POSTEX':
                import re as _re
                m = _re.match(r'^R0*(\d+)_([A-Z])-(\d+)$', str(row[6] or ''))
                if m:
                    rampa = f"R{int(m.group(1)):02d}{m.group(2)}"
                    pos   = int(m.group(3))
                    run_occ.setdefault(dia_new, {}).setdefault(rampa, {})[pos] = info.get('grupo_new','ESP')
        # Rename description if playa is cancelled in SEMANA SANTA
        if cancelled_especiales and playa.upper() in cancelled_especiales and new_rows:
            semana = especial_bloque_map and next(
                (k[1] for k in (especial_bloque_map or {}) if k[0]==dia_new and k[1]==playa), None)
            suffix = '_CANCELADA_SOLO_W14'
            import re as _re_canc
            renamed = []
            for _r in new_rows:
                _desc = str(_r[2] or '')
                if playa in _desc and suffix not in _desc:
                    # Replace playa name with playa+suffix wherever it appears
                    # Handle both "(ESPECIAL" and "_ESPECIAL" suffixes
                    if f'{playa} ' in _desc or f'{playa}(' in _desc or f'{playa}_' in _desc:
                        _desc = _desc.replace(playa, playa + suffix, 1)
                renamed.append((_r[0], _r[1], _desc, _r[3], _r[4], _r[5], _r[6]))
            new_rows = renamed
        output_rows.extend(new_rows)
        especial_rows.extend(new_rows)
        added += len(new_rows)
        results.append(info)

    return output_rows, {
        'n_canceladas': len(canceladas), 'n_especiales': len(especiales),
        'rows_removed_cancelled': rem_cancel, 'rows_removed_moved': rem_moved,
        'rows_kept': kept, 'rows_kept_non_weekly': kept_nw,
        'rows_added_especial': added, 'total_output_rows': len(output_rows),
        'canceladas': canceladas, 'assignment_results': results,
        'especial_rows': especial_rows,
    }



# ─── CHART ENRICHMENT ─────────────────────────────────────────────────────────

_BLOQUE_DATA_JS = 'BLOQUE_DATA_JS = """[\n  {b:\'1BLOS1\',cl:\'S1\',lib:\'SÁB 01:00\',desac:\'SÁB 10:00\',std:13,s14std:13,s14esp:0,can:0},\n  {b:\'1BLOD0\',cl:\'D0\',lib:\'DOM 15:00\',desac:\'LUN 02:15\',std:0,s14std:0,s14esp:0,can:0},\n  {b:\'1BLOD1\',cl:\'D1\',lib:\'DOM 16:00\',desac:\'LUN 02:15\',std:165,s14std:165,s14esp:0,can:0},\n  {b:\'1BLOD2\',cl:\'D2\',lib:\'DOM 20:00\',desac:\'LUN 04:00\',std:97,s14std:59,s14esp:38,can:0},\n  {b:\'1BLOD3\',cl:\'D3\',lib:\'DOM 20:01\',desac:\'LUN 06:00\',std:72,s14std:63,s14esp:0,can:9},\n  {b:\'1BLOD4\',cl:\'D4\',lib:\'DOM 23:59\',desac:\'LUN 09:00\',std:41,s14std:22,s14esp:26,can:19},\n  {b:\'2BLOL0\',cl:\'L0\',lib:\'LUN 01:00\',desac:\'LUN 11:00\',std:14,s14std:14,s14esp:0,can:0},\n  {b:\'2BLOL1\',cl:\'L1\',lib:\'LUN 03:00\',desac:\'LUN 13:00\',std:169,s14std:169,s14esp:0,can:0},\n  {b:\'2BLOL2\',cl:\'L2\',lib:\'LUN 09:00\',desac:\'LUN 19:00\',std:69,s14std:69,s14esp:0,can:0},\n  {b:\'2BLOL3\',cl:\'L3\',lib:\'LUN 16:00\',desac:\'MAR 02:00\',std:67,s14std:67,s14esp:0,can:0},\n  {b:\'2BLOL4\',cl:\'L4\',lib:\'LUN 18:00\',desac:\'MAR 04:00\',std:189,s14std:88,s14esp:76,can:0},\n  {b:\'2BLOL5\',cl:\'L5\',lib:\'LUN 22:00\',desac:\'MAR 08:00\',std:81,s14std:54,s14esp:14,can:0},\n  {b:\'3BLOM1\',cl:\'M1\',lib:\'MAR 03:00\',desac:\'MAR 13:00\',std:108,s14std:108,s14esp:0,can:0},\n  {b:\'3BLOM2\',cl:\'M2\',lib:\'MAR 06:30\',desac:\'MAR 16:30\',std:43,s14std:0,s14esp:37,can:43},\n  {b:\'3BLOM3\',cl:\'M3\',lib:\'MAR 09:00\',desac:\'MIÉ 07:00\',std:29,s14std:19,s14esp:0,can:0},\n  {b:\'3BLOM4\',cl:\'M4\',lib:\'MAR 15:00\',desac:\'MIÉ 01:00\',std:95,s14std:95,s14esp:0,can:0},\n  {b:\'3BLOM5\',cl:\'M5\',lib:\'MAR 18:00\',desac:\'MIÉ 05:00\',std:166,s14std:17,s14esp:29,can:0},\n  {b:\'3BLOM6\',cl:\'M6\',lib:\'MAR 23:00\',desac:\'MIÉ 09:00\',std:137,s14std:131,s14esp:0,can:0},\n  {b:\'4BLOX1\',cl:\'X1\',lib:\'MIÉ 03:00\',desac:\'MIÉ 13:00\',std:65,s14std:4,s14esp:27,can:0},\n  {b:\'4BLOX2\',cl:\'X2\',lib:\'MIÉ 07:00\',desac:\'MIÉ 17:00\',std:38,s14std:38,s14esp:0,can:0},\n  {b:\'4BLOX3\',cl:\'X3\',lib:\'MIÉ 15:00\',desac:\'JUE 02:30\',std:177,s14std:136,s14esp:59,can:0},\n  {b:\'4BLOX4\',cl:\'X4\',lib:\'MIÉ 19:00\',desac:\'JUE 04:00\',std:100,s14std:75,s14esp:28,can:0},\n  {b:\'4BLOX5\',cl:\'X5\',lib:\'MIÉ 23:59\',desac:\'JUE 08:00\',std:80,s14std:52,s14esp:252,can:0},\n  {b:\'5BLOJ1\',cl:\'J1\',lib:\'JUE 03:00\',desac:\'JUE 13:00\',std:150,s14std:114,s14esp:0,can:0},\n  {b:\'5BLOJ2\',cl:\'J2\',lib:\'JUE 09:00\',desac:\'JUE 23:00\',std:72,s14std:37,s14esp:0,can:0},\n  {b:\'5BLOJ3\',cl:\'J3\',lib:\'JUE 15:00\',desac:\'VIE 01:00\',std:190,s14std:152,s14esp:28,can:0},\n  {b:\'5BLOJ4\',cl:\'J4\',lib:\'JUE 18:00\',desac:\'VIE 03:00\',std:139,s14std:68,s14esp:24,can:0},\n  {b:\'5BLOJ5\',cl:\'J5\',lib:\'JUE 23:00\',desac:\'VIE 09:00\',std:118,s14std:118,s14esp:0,can:0},\n  {b:\'6BLOV1\',cl:\'V1\',lib:\'VIE 05:00\',desac:\'VIE 15:00\',std:33,s14std:33,s14esp:0,can:0},\n  {b:\'6BLOV2\',cl:\'V2\',lib:\'VIE 09:00\',desac:\'SÁB 07:00\',std:16,s14std:16,s14esp:0,can:0},\n  {b:\'6BLOV3\',cl:\'V3\',lib:\'VIE 17:00\',desac:\'SÁB 05:00\',std:200,s14std:200,s14esp:0,can:0},\n  {b:\'6BLOV4\',cl:\'V4\',lib:\'VIE 23:00\',desac:\'SÁB 11:00\',std:41,s14std:41,s14esp:0,can:0},\n  {b:\'6BLOV5\',cl:\'V5\',lib:\'SÁB 03:00\',desac:\'DOM 14:00\',std:46,s14std:46,s14esp:0,can:0}\n]"""\n'

_CHART_HTML = '<div class="sec no-print-break" style="margin-top:40px;">\n  <div class="sn">04 — Ocupación del sorter por bloque horario</div>\n  <h2>Análisis de capacidad: semana estándar vs S14 Semana Santa</h2>\n  <div class="sd">Haz clic en cualquier barra para ver las agrupaciones playa asignadas a ese bloque. Azul = salidas regulares mantenidas · Naranja = salidas especiales reasignadas · Rojo translúcido = posiciones canceladas liberadas.</div>\n\n  <div style="display:flex;gap:10px;margin-bottom:1.2rem;flex-wrap:wrap;">\n    <div style="background:#f5f5f5;border-radius:6px;padding:.75rem 1rem;flex:1;min-width:110px;text-align:center;">\n      <div style="font-size:11px;color:#888;margin-bottom:3px;">capacidad total</div>\n      <div style="font-size:20px;font-weight:500;">653</div>\n      <div style="font-size:10px;color:#888;">posiciones · 50 rampas</div>\n    </div>\n    <div style="background:#f5f5f5;border-radius:6px;padding:.75rem 1rem;flex:1;min-width:110px;text-align:center;">\n      <div style="font-size:11px;color:#888;margin-bottom:3px;">bloque más cargado S14</div>\n      <div style="font-size:20px;font-weight:500;color:#b45309;">4BLOX5</div>\n      <div style="font-size:10px;color:#888;">46.5% cap. (MEXICO + especiales)</div>\n    </div>\n    <div style="background:#f5f5f5;border-radius:6px;padding:.75rem 1rem;flex:1;min-width:110px;text-align:center;">\n      <div style="font-size:11px;color:#888;margin-bottom:3px;">día con más especiales</div>\n      <div style="font-size:20px;font-weight:500;color:#b45309;">Miércoles</div>\n      <div style="font-size:10px;color:#888;">19 salidas reasignadas aquí</div>\n    </div>\n    <div style="background:#f5f5f5;border-radius:6px;padding:.75rem 1rem;flex:1;min-width:110px;text-align:center;">\n      <div style="font-size:11px;color:#888;margin-bottom:3px;">máximo alcanzado</div>\n      <div style="font-size:20px;font-weight:500;color:#1d6a3e;">46.5%</div>\n      <div style="font-size:10px;color:#888;">ningún bloque al 100%</div>\n    </div>\n  </div>\n\n  <div style="display:flex;gap:14px;margin-bottom:10px;font-size:11px;color:#888;flex-wrap:wrap;align-items:center;">\n    <span style="display:flex;align-items:center;gap:5px;"><span style="width:11px;height:11px;border-radius:2px;background:#378ADD;display:inline-block;"></span>estándar (mantenido en S14)</span>\n    <span style="display:flex;align-items:center;gap:5px;"><span style="width:11px;height:11px;border-radius:2px;background:#EF9F27;display:inline-block;"></span>especiales añadidas</span>\n    <span style="display:flex;align-items:center;gap:5px;"><span style="width:11px;height:11px;border-radius:2px;background:rgba(226,75,74,0.35);border:1px solid #E24B4A;display:inline-block;"></span>canceladas liberadas</span>\n    <span style="display:flex;align-items:center;gap:5px;"><span style="display:inline-block;width:18px;border-top:2px dashed #E24B4A;"></span>100% cap.</span>\n    <span style="color:#1d4ed8;font-style:italic;">← haz clic en una barra</span>\n  </div>\n\n  <div style="display:grid;grid-template-columns:1fr 340px;gap:16px;align-items:start;" id="chart-grid">\n    <div style="position:relative;width:100%;height:720px;"><canvas id="occ-chart"></canvas></div>\n    <div id="detail-panel" style="border:0.5px solid #ebebeb;border-radius:8px;padding:14px;font-size:12px;min-height:200px;background:#fafafa;">\n      <div style="color:#999;font-size:11px;font-style:italic;text-align:center;margin-top:40px;">Selecciona un bloque en el gráfico para ver sus agrupaciones playa</div>\n    </div>\n  </div>\n\n  <div style="margin-top:1.4rem;border-top:0.5px solid #ebebeb;padding-top:1rem;">\n    <div style="font-size:11px;color:#888;margin-bottom:8px;font-weight:500;">salidas por día — semana S14 vs semana normal</div>\n    <div id="day-bars"></div>\n  </div>\n</div>\n\n<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>\n<script>\nconst TOTAL_CAP = 653;\nconst DETAIL = {DETAIL_JSON};\nconst BLOQUES = {BLOQUE_DATA_JS};\n\nconst labels = BLOQUES.map(d => d.b);\nconst stdPct  = BLOQUES.map(d => +(d.std  / TOTAL_CAP * 100).toFixed(1));\nconst espPct  = BLOQUES.map(d => +(d.s14esp / TOTAL_CAP * 100).toFixed(1));\nconst canPct  = BLOQUES.map(d => +(d.can  / TOTAL_CAP * 100).toFixed(1));\n\nconst ctx = document.getElementById(\'occ-chart\').getContext(\'2d\');\nconst chart = new Chart(ctx, {{\n  type: \'bar\',\n  data: {{\n    labels,\n    datasets: [\n      {{ label:\'Estándar\', data:stdPct, backgroundColor:\'#378ADD\', borderRadius:0, stack:\'s\', barThickness:16 }},\n      {{ label:\'Especiales\', data:espPct, backgroundColor:\'#EF9F27\', borderRadius:2, stack:\'s\', barThickness:16 }},\n      {{ label:\'Canceladas\', data:canPct, backgroundColor:\'rgba(226,75,74,0.3)\', borderColor:\'#E24B4A\', borderWidth:1, borderRadius:0, stack:\'s\', barThickness:16 }}\n    ]\n  }},\n  options: {{\n    indexAxis:\'y\', responsive:true, maintainAspectRatio:false,\n    onClick(evt, elements) {{\n      if (!elements.length) return;\n      const idx = elements[0].index;\n      showDetail(idx);\n    }},\n    plugins: {{\n      legend: {{ display:false }},\n      tooltip: {{\n        callbacks: {{\n          title: items => {{\n            const d = BLOQUES[items[0].dataIndex];\n            return d.b + \' · \' + d.lib + \' → \' + d.desac;\n          }},\n          label: item => {{\n            const d = BLOQUES[item.dataIndex];\n            const tot = d.s14std + d.s14esp;\n            const pct = (tot/TOTAL_CAP*100).toFixed(1);\n            if (item.datasetIndex===0) return \' Estándar mantenido: \' + d.s14std + \' pos\';\n            if (item.datasetIndex===1) return \' Especiales añadidas: \' + d.s14esp + \' pos — total S14: \' + tot + \' (\' + pct + \'%)\';\n            if (item.datasetIndex===2) return \' Canceladas liberadas: \' + d.can + \' pos\';\n            return \'\';\n          }}\n        }}\n      }}\n    }},\n    scales: {{\n      x: {{ stacked:true, min:0, max:50, grid:{{color:\'rgba(0,0,0,0.06)\'}},\n            ticks:{{color:\'#888\',font:{{size:11}},callback:v=>v+\'%\'}},\n            title:{{display:true,text:\'% posiciones del sorter (cap. total = 653)\',color:\'#888\',font:{{size:10}}}} }},\n      y: {{ stacked:true, grid:{{display:false}},\n            ticks:{{color:\'#555\',font:{{size:10,family:\'monospace\'}},autoSkip:false}} }}\n    }}\n  }},\n  plugins: [{{\n    id:\'tline\',\n    afterDraw(chart) {{\n      const {{ctx,scales}} = chart;\n      const xPos = scales.x.getPixelForValue(100);\n      if (xPos > scales.x.left && xPos < scales.x.right) {{\n        ctx.save(); ctx.beginPath(); ctx.setLineDash([4,3]);\n        ctx.strokeStyle=\'#E24B4A\'; ctx.lineWidth=1.5;\n        ctx.moveTo(xPos, scales.y.top); ctx.lineTo(xPos, scales.y.bottom);\n        ctx.stroke(); ctx.restore();\n      }}\n      BLOQUES.forEach((d,i) => {{\n        const tot = d.s14std + d.s14esp;\n        const pct = tot/TOTAL_CAP*100;\n        if (pct > 30) {{\n          const yPos = scales.y.getPixelForValue(d.b);\n          const xEnd = scales.x.getPixelForValue(pct);\n          ctx.save(); ctx.font=\'bold 10px monospace\';\n          ctx.fillStyle = pct > 45 ? \'#c2410c\' : \'#92400e\';\n          ctx.textAlign=\'left\';\n          ctx.fillText(pct.toFixed(1)+\'%\', xEnd+4, yPos+4);\n          ctx.restore();\n        }}\n      }});\n    }}\n  }}]\n}});\n\nlet selectedIdx = null;\nfunction showDetail(idx) {{\n  const d = BLOQUES[idx];\n  const det = DETAIL[d.b] || {{std:[],esp:[],can:[]}};\n  const panel = document.getElementById(\'detail-panel\');\n  const tot = d.s14std + d.s14esp;\n  const pct = (tot/TOTAL_CAP*100).toFixed(1);\n  const stdN = d.s14std, espN = d.s14esp, canN = d.can;\n\n  const makePill = (text, col) =>\n    `<span style="display:inline-block;font-size:9px;padding:1px 6px;border-radius:3px;background:${{col}}20;color:${{col}};font-family:monospace;margin:1px;">${{text}}</span>`;\n\n  const makeRow = (item, type) => {{\n    const rampaStr = item.r.slice(0,4).join(\' \') + (item.r.length>4 ? ` +${{item.r.length-4}}` : \'\');\n    const pctStr = item.pct.toFixed(1)+\'%\';\n    const bar = `<div style="height:4px;border-radius:2px;background:${{type===\'std\'?\'#378ADD\':type===\'esp\'?\'#EF9F27\':\'rgba(226,75,74,0.5)\'}};width:${{Math.min(item.pct/50*100,100)}}%;margin-top:2px;"></div>`;\n    return `<tr style="border-bottom:0.5px solid #f0f0f0;">\n      <td style="padding:5px 6px 5px 0;vertical-align:top;font-size:11px;font-weight:${{type===\'esp\'?\'500\':\'400\'}};color:${{type===\'can\'?\'#9a3412\':\'#1a1a1a\'}};white-space:nowrap;max-width:160px;overflow:hidden;text-overflow:ellipsis;" title="${{item.p}}">${{item.p.replace(/_/g,\'_\\u200b\')}}</td>\n      <td style="padding:5px 0 5px 4px;text-align:right;font-size:10px;color:#555;white-space:nowrap;font-family:monospace;">${{item.n}}p · ${{pctStr}}</td>\n    </tr>`;\n  }};\n\n  let html = `<div style="border-bottom:0.5px solid #ebebeb;padding-bottom:10px;margin-bottom:10px;">\n    <div style="font-size:13px;font-weight:500;font-family:monospace;">${{d.b}}</div>\n    <div style="font-size:11px;color:#888;margin-top:2px;">${{d.lib}} → ${{d.desac}}</div>\n    <div style="display:flex;gap:8px;margin-top:8px;flex-wrap:wrap;">\n      ${{makePill(stdN+\' std\',\'#1d4ed8\')}}\n      ${{espN>0?makePill(espN+\' esp\',\'#92400e\'):\'\'}}\n      ${{canN>0?makePill(canN+\' can\',\'#9a3412\'):\'\'}}\n      ${{makePill(pct+\'% cap.\', tot/TOTAL_CAP>0.4?\'#c2410c\':\'#1d6a3e\')}}\n    </div>\n  </div>`;\n\n  if (det.esp && det.esp.length) {{\n    html += `<div style="font-size:10px;font-weight:500;color:#92400e;margin-bottom:4px;text-transform:uppercase;letter-spacing:.06em;">Especiales añadidas S14</div>\n    <table style="width:100%;border-collapse:collapse;margin-bottom:10px;">${{det.esp.map(p=>makeRow(p,\'esp\')).join(\'\')}}</table>`;\n  }}\n  if (det.std && det.std.length) {{\n    html += `<div style="font-size:10px;font-weight:500;color:#1d4ed8;margin-bottom:4px;text-transform:uppercase;letter-spacing:.06em;">Regulares mantenidas</div>\n    <table style="width:100%;border-collapse:collapse;margin-bottom:10px;">${{det.std.map(p=>makeRow(p,\'std\')).join(\'\')}}</table>`;\n  }}\n  if (det.can && det.can.length) {{\n    html += `<div style="font-size:10px;font-weight:500;color:#9a3412;margin-bottom:4px;text-transform:uppercase;letter-spacing:.06em;">Canceladas (liberadas)</div>\n    <table style="width:100%;border-collapse:collapse;">${{det.can.map(p=>makeRow(p,\'can\')).join(\'\')}}</table>`;\n  }}\n  if (!det.std.length && !det.esp.length && !det.can.length) {{\n    html += `<div style="color:#999;font-size:11px;font-style:italic;text-align:center;margin-top:30px;">Sin datos de playa para este bloque</div>`;\n  }}\n\n  panel.innerHTML = html;\n  selectedIdx = idx;\n}}\n\nconst dayData = [\n  {{dia:\'Domingo\',reg:4,esp:0,can:0}},\n  {{dia:\'Lunes\',reg:84,esp:11,can:4}},\n  {{dia:\'Martes\',reg:52,esp:15,can:1}},\n  {{dia:\'Miércoles\',reg:39,esp:10,can:5}},\n  {{dia:\'Jueves\',reg:38,esp:16,can:19}},\n  {{dia:\'Viernes\',reg:30,esp:0,can:9}},\n  {{dia:\'Sábado\',reg:20,esp:0,can:0}},\n];\nconst maxS = 100;\ndocument.getElementById(\'day-bars\').innerHTML = dayData.map(d => {{\n  const wR = Math.round(d.reg/maxS*100), wE = Math.round(d.esp/maxS*100), wC = Math.round(d.can/maxS*100);\n  return `<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">\n    <div style="font-size:11px;color:#888;width:70px;text-align:right;flex-shrink:0;">${{d.dia}}</div>\n    <div style="flex:1;display:flex;align-items:center;height:16px;gap:1px;">\n      <div style="height:16px;width:${{wR}}%;background:#378ADD;border-radius:2px 0 0 2px;min-width:${{d.reg>0?2:0}}px;"></div>\n      ${{d.esp>0?`<div style="height:16px;width:${{wE}}%;background:#EF9F27;min-width:2px;"></div>`:\'\'}}\n      ${{d.can>0?`<div style="height:16px;width:${{wC}}%;background:rgba(226,75,74,0.4);border:1px solid #E24B4A;min-width:2px;"></div>`:\'\'}}\n    </div>\n    <div style="font-size:11px;color:#888;white-space:nowrap;min-width:140px;">\n      ${{d.reg}} reg${{d.esp>0?` + <b style="color:#92400e">${{d.esp}} esp</b>`:\'\'}}${{d.can>0?` − <span style="color:#9a3412">${{d.can}} can</span>`:\'\'}}\n    </div>\n  </div>`;\n}}).join(\'\');\n</script>\n"""'


def _build_detail_json(output_rows, total_cap=653):
    """Compute per-bloque playa detail JSON from the output rows for the chart panel."""
    BR = re.compile(r"(\d+BLO[A-Z]\d+)")
    SR = re.compile(r"^R\d+")
    DY = ["DOMINGO","LUNES","MARTES","MIERCOLES","JUEVES","VIERNES","SABADO"]

    def _playa(desc):
        if not desc: return None, None
        c = re.sub(r"^\[B2B\]\s*","",str(desc)).strip()
        c = re.sub(r"\s+PARA BAJAR POR.*$","",c)
        c = re.sub(r"_CANCELADA[^_]*$","",c)
        m = BR.match(c)
        if not m: return None, None
        rest = c[len(m.group(1))+1:]
        for d in DY:
            if rest.upper().startswith(d+"_"):
                return m.group(1), re.sub(r"\s*\(ESPECIAL.*","",rest[len(d)+1:]).strip()
        return m.group(1), None

    def _slot(store, b, p, elem):
        m = re.match(r"^R0*(\d+)_([A-Z])-(\d+)$", elem)
        if not m: return
        store[b][p].add((f"R{int(m.group(1)):02d}{m.group(2)}", int(m.group(3))))

    ss = defaultdict(lambda: defaultdict(set))
    se = defaultdict(lambda: defaultdict(set))

    for row in output_rows:
        _, grupo, desc, zona, dest, alm, elem = row
        if zona != "POSTEX" or not elem or not SR.match(str(elem)): continue
        b, p = _playa(desc)
        if not b or not p: continue
        _slot(se if "ESPECIAL" in str(desc) else ss, b, p, str(elem))

    def ml(d):
        return [{"p":p,"n":len(s),"r":sorted(set(x[0] for x in s)),
                 "pct":round(len(s)/total_cap*100,1)}
                for p,s in sorted(d.items(),key=lambda x:-len(x[1]))]

    result = {}
    for b in sorted(set(list(ss)+list(se))):
        result[b] = {"std":ml(ss[b]),"esp":ml(se[b]),"can":[]}
    return json.dumps(result, ensure_ascii=False)


def _enrich_html(html, output_rows, summary):
    """Inject chart section + compact cancel list into base HTML."""
    detail_json = _build_detail_json(output_rows)
    chart = _CHART_HTML.replace("{DETAIL_JSON}", detail_json).replace("{BLOQUE_DATA_JS}", _BLOQUE_DATA_JS)

    # Compact cancel list
    found = re.findall(
        r'<span class="b bd">([^<]+)</span></td><td>([^<]+)</td><span class="b bc">CANCELADA',
        html)
    if found:
        by_day = defaultdict(list)
        for dia, playa in found: by_day[dia].append(playa)
        items = ""
        for dia in ["LUNES","MARTES","MIERCOLES","JUEVES","VIERNES","SABADO","DOMINGO"]:
            for playa in by_day.get(dia,[]):
                items += (f'<li style="display:flex;gap:10px;align-items:baseline;padding:3px 0;">' +
                          f'<span style="font-size:9px;font-family:monospace;background:#f3f4f6;color:#374151;' +
                          f'padding:1px 5px;border-radius:2px;white-space:nowrap;flex-shrink:0;">{dia}</span>' +
                          f'<span style="font-size:12px;">{playa}</span></li>')
        nc = (f'<div class="sn">01 — Salidas canceladas</div>\n' +
              f'  <h2>Destinos eliminados del sorter esta semana</h2>\n' +
              f'  <div class="sd" style="margin-bottom:12px;">{len(found)} salidas · ' +
              f'{summary["rows_removed_cancelled"]:,} filas eliminadas del GRUPO_DESTINOS.</div>\n' +
              f'  <ul style="list-style:none;padding:0;margin:0;columns:2;gap:24px;">\n{items}  </ul>')
        s = html.find('<div class="sn">01 — Salidas canceladas</div>')
        e = html.find('<div class="sec">', s+10)
        if s>=0 and e>=0:
            html = html[:s] + nc + "\n</div>\n\n" + html[e:]

    return html.replace("</body>", chart + "\n</body>")


# ─── WRITE EXCEL ──────────────────────────────────────────────────────────────

def _split_destino(destino):
    """Split 10-digit destino into (id=first2, destino=last8)."""
    d = ''.join(c for c in str(destino or '') if c.isdigit())
    d10 = d.zfill(10) if d else '0000000000'
    return d10[:2], d10[2:]  # id, destino


def write_gd(output_rows, header, out_path):
    wb = Workbook(); ws = wb.active; ws.title = 'Hoja1'
    hf = PatternFill('solid', start_color='1A1A1A')
    for ci, w in enumerate([8,15,65,10,15,10,15], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ci, h in enumerate(header, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = hf; c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 18
    thin = Border(bottom=Side(style='thin', color='EBEBEB'))
    df = Font(name='Arial', size=9)
    for ri, row in enumerate(output_rows, 2):
        _, grupo, desc, tipo_zona, destino, almacen, elemento = row
        for ci, val in enumerate([f'=MID(B{ri},3,2)' if grupo else '', grupo, desc,
                                   tipo_zona, destino, almacen, elemento], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = df; c.border = thin
            if ci == 1: c.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'

    # ── POSTEX_DXC + SOREXP_DXC sheets ───────────────────────────────────────
    # Format: DESCRIPCIÓN | DESTINO(8) | ID(2) | ELEMENTO | SECUENCIA(10)
    dxc_headers = ['DESCRIPCIÓN', 'DESTINO', 'ID', 'ELEMENTO', 'SECUENCIA']
    dxc_widths  = [70, 12, 6, 18, 12]
    for sheet_name, tipo_filter in [('POSTEX_DXC', 'POSTEX'), ('SOREXP_DXC', 'SOREXP')]:
        ws2 = wb.create_sheet(sheet_name)
        hf2 = PatternFill('solid', start_color='1F4E79')
        for ci, (h, w) in enumerate(zip(dxc_headers, dxc_widths), 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill = hf2; c.alignment = Alignment(horizontal='center')
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[1].height = 18
        ws2.freeze_panes = 'A2'
        thin2 = Border(bottom=Side(style='thin', color='EBEBEB'))
        df2 = Font(name='Arial', size=9)
        ri2 = 2
        for row in output_rows:
            _, grupo, desc, tipo_zona, destino, almacen, elemento = row
            if str(tipo_zona).strip().upper() != tipo_filter: continue
            if not desc or str(desc).startswith('='): continue
            _id, _dest = _split_destino(destino)
            vals = [desc, _dest, _id, elemento, '10']  # secuencia always 10
            for ci, val in enumerate(vals, 1):
                c = ws2.cell(row=ri2, column=ci, value=val)
                c.font = df2; c.border = thin2
            ri2 += 1

    wb.save(out_path)


# ─── WRITE HTML ───────────────────────────────────────────────────────────────

def write_html(summary, semana, out_path):
    _DAY_ORDER = ['DOMINGO','LUNES','MARTES','MIERCOLES','JUEVES','VIERNES','SABADO']
    def _day_key(r): return (_DAY_ORDER.index(r['dia_new']) if r['dia_new'] in _DAY_ORDER else 99, r['playa'])

    ok      = sorted([r for r in summary['assignment_results'] if r['status'] == 'OK'],       key=_day_key)
    partial = sorted([r for r in summary['assignment_results'] if r['status'] == 'PARTIAL'],  key=_day_key)
    e2      = sorted([r for r in summary['assignment_results'] if r['status'] == 'E2_ROUTE'], key=_day_key)
    nocfg   = sorted([r for r in summary['assignment_results'] if r['status'] == 'NO_CONFIG'],key=_day_key)
    cancels = list(summary['canceladas'].items())
    n_warn  = len(partial) + len(nocfg)
    now     = datetime.now().strftime("%d/%m/%Y %H:%M")

    def b(cls, txt): return f'<span class="b {cls}">{txt}</span>'

    def cancel_rows():
        out, seen = '', defaultdict(list)
        for (dia, playa), _ in cancels: seen[dia].append(playa)
        for dia in ['DOMINGO','LUNES','MARTES','MIERCOLES','JUEVES','VIERNES','SABADO']:
            for playa in seen.get(dia, []):
                out += f'<tr><td>{b("bd",dia)}</td><td>{playa}</td>{b("bc","CANCELADA")}</td><td class="mt">—</td></tr>\n'
        return out

    def ok_rows():
        out = ''
        for r in ok:
            fb  = f' <em class="mt">(datos de {r["source_day"]})</em>' \
                  if r.get('source_day') and r['source_day'] != r['dia_orig'] else ''
            blq = f'<span class="mono">{r["bloque_new"]}</span>' if r.get('bloque_new') and r['bloque_new'] != '?' else '<span class="mt">—</span>'
            rstr = ' · '.join(f'{k}[{",".join(str(p) for p in v)}]' for k,v in sorted(r['rampas'].items()))
            out += (f'<tr><td>{b("bs",r["dia_new"])}</td><td>{r["playa"]}{fb}</td>'
                    f'<td>{b("bd",r["dia_orig"])}</td><td>{blq}</td>'
                    f'<td class="mono mt small">{rstr}</td></tr>\n')
        return out

    def warn_rows():
        out = ''
        for r in partial:
            out += (f'<tr><td>{b("bd",r["dia_orig"])}</td><td>{r["playa"]}</td>'
                    f'<td>{b("bw","→ "+r["dia_new"])}</td><td colspan="2" class="wt">'
                    f'⚠ Solo {r["n_assigned"]}/{r["n_destinos"]} posiciones libres</td></tr>\n')
        for r in e2:
            out += (f'<tr><td>{b("bd",r["dia_orig"])}</td><td>{r["playa"]}</td>'
                    f'<td>{b("bi","→ "+r["dia_new"])}</td><td colspan="2" class="it">'
                    f'🔀 Ruta E2/manual — no pasa por rampas del sorter</td></tr>\n')
        for r in nocfg:
            out += (f'<tr><td>{b("bd",r["dia_orig"])}</td><td>{r["playa"]}</td>'
                    f'<td>{b("bw","→ "+r["dia_new"])}</td><td colspan="2" class="wt">'
                    f'❌ Sin configuración GD — revisar con equipo</td></tr>\n')
        return out

    html = f'''<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Sorter VDL B2B — {semana}</title>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:"Trebuchet MS",Arial,sans-serif;font-weight:300;background:#f0f0f0;color:#1a1a1a;font-size:13px;line-height:1.5}}
.rpt{{max-width:980px;margin:0 auto;background:#fff;min-height:100vh}}
.rh{{background:#000;color:#fff;padding:36px 48px 28px}}
.ey{{font-size:10px;letter-spacing:.12em;text-transform:uppercase;color:#888;margin-bottom:10px}}
.rh h1{{font-size:26px;font-weight:300;letter-spacing:-.02em;margin-bottom:4px}}
.rh .sub{{font-size:13px;color:#888}}
.meta{{display:flex;gap:28px;margin-top:20px;padding-top:16px;border-top:1px solid #333}}
.mi label{{font-size:9px;letter-spacing:.1em;text-transform:uppercase;color:#666;display:block}}
.mi span{{font-size:12px;color:#ccc}}
.kpis{{background:#000;display:grid;grid-template-columns:repeat(5,1fr);border-top:1px solid #222}}
.kpi{{padding:18px 20px;border-right:1px solid #222}}.kpi:last-child{{border-right:none}}
.kv{{font-size:28px;font-weight:300;color:#fff;letter-spacing:-.03em;font-family:monospace;line-height:1;margin-bottom:4px}}
.kl{{font-size:9px;letter-spacing:.1em;text-transform:uppercase;color:#555}}
.kv.red{{color:#ef4444}}.kv.amb{{color:#f59e0b}}.kv.grn{{color:#22c55e}}
.cnt{{padding:0 48px 48px}}
.sec{{margin-top:34px}}
.sn{{font-size:10px;letter-spacing:.12em;text-transform:uppercase;color:#999;margin-bottom:4px}}
.sec h2{{font-size:17px;font-weight:400;letter-spacing:-.01em;margin-bottom:3px}}
.sd{{font-size:12px;color:#888;margin-bottom:14px}}
.sb{{display:grid;grid-template-columns:repeat(3,1fr);gap:1px;background:#ebebeb;border:1px solid #ebebeb;border-radius:3px;overflow:hidden;margin-bottom:24px}}
.sc{{background:#fff;padding:14px 18px;text-align:center}}
.sv{{font-size:22px;font-weight:300;color:#1a1a1a;font-family:monospace;letter-spacing:-.02em}}
.sl{{font-size:9px;letter-spacing:.1em;text-transform:uppercase;color:#999;margin-top:2px}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
thead tr{{border-bottom:1px solid #1a1a1a}}
thead th{{text-align:left;font-size:9px;letter-spacing:.08em;text-transform:uppercase;color:#999;padding:7px 10px 7px 0;font-weight:400}}
tbody tr{{border-bottom:1px solid #ebebeb}}tbody tr:last-child{{border-bottom:none}}
tbody td{{padding:8px 10px 8px 0;vertical-align:top}}
.mt{{color:#888}}.wt{{color:#b45309}}.it{{color:#1d4ed8}}
.mono{{font-family:monospace;font-size:11px}}.small{{font-size:11px}}em.mt{{font-style:normal}}
.b{{display:inline-block;font-size:9px;letter-spacing:.06em;text-transform:uppercase;padding:2px 6px;border-radius:2px;font-family:monospace;white-space:nowrap}}
.bd{{background:#f3f4f6;color:#374151}}.bc{{background:#fee2e2;color:#b91c1c}}
.bs{{background:#dbeafe;color:#1d4ed8}}.bw{{background:#fef3c7;color:#92400e}}
.bi{{background:#eff6ff;color:#1d4ed8}}
.ib{{border-left:3px solid #3b82f6;background:#eff6ff;padding:11px 15px;margin-bottom:10px;font-size:12px;color:#1e40af}}
.wb{{border-left:3px solid #f59e0b;background:#fffbeb;padding:11px 15px;margin-bottom:10px;font-size:12px}}
.wb strong{{color:#92400e}}
.rf{{background:#000;color:#555;font-size:9px;letter-spacing:.08em;text-transform:uppercase;padding:14px 48px;display:flex;justify-content:space-between;font-family:monospace}}
.np{{text-align:center;padding:22px 0 36px}}
.bp{{font-family:monospace;background:#000;color:#fff;border:none;padding:9px 28px;cursor:pointer;text-transform:uppercase;letter-spacing:.1em;font-size:10px}}
.bp:hover{{background:#333}}
@media print{{
  @page{{size:A4 portrait;margin:12mm 14mm 14mm 14mm}}
  *{{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}}
  body{{background:#fff}}.rpt{{max-width:100%}}
  .rh,.kpis,.sec{{break-inside:avoid}}tr{{break-inside:avoid}}.np{{display:none!important}}
}}
</style></head><body>
<div class="rpt">
<div class="rh">
  <div class="ey">Logística · VDL B2B</div>
  <h1>Configuración Sorter — {semana}</h1>
  <div class="sub">Semana especial · asignación con control de solapamiento de bloques horarios</div>
  <div class="meta">
    <div class="mi"><label>Generado</label><span>{now}</span></div>
    <div class="mi"><label>Semana</label><span>{semana}</span></div>
    <div class="mi"><label>Filas output GD</label><span>{summary["total_output_rows"]:,}</span></div>
    <div class="mi"><label>Canceladas</label><span>{summary["n_canceladas"]}</span></div>
    <div class="mi"><label>Cambios de día</label><span>{summary["n_especiales"]}</span></div>
  </div>
</div>
<div class="kpis">
  <div class="kpi"><div class="kv red">{summary["n_canceladas"]}</div><div class="kl">Canceladas</div></div>
  <div class="kpi"><div class="kv grn">{len(ok)}</div><div class="kl">Asignadas OK</div></div>
  <div class="kpi"><div class="kv">{len(e2)}</div><div class="kl">Rutas E2</div></div>
  <div class="kpi"><div class="kv {"amb" if n_warn else "grn"}">{n_warn}</div><div class="kl">Revisar</div></div>
  <div class="kpi"><div class="kv">{summary["total_output_rows"]:,}</div><div class="kl">Filas GD</div></div>
</div>
<div class="cnt">

<div class="sec">
  <div class="sn">00 — Resumen</div>
  <h2>Impacto en el fichero GRUPO_DESTINOS</h2>
  <div class="sd">La asignación de posiciones respeta los horarios de bloque: dos destinos pueden compartir rampa si sus ventanas de liberación/desactivación no se solapan.</div>
  <div class="sb">
    <div class="sc"><div class="sv">{summary["rows_removed_cancelled"]:,}</div><div class="sl">Filas eliminadas</div></div>
    <div class="sc"><div class="sv">{summary["rows_added_especial"]:,}</div><div class="sl">Filas añadidas</div></div>
    <div class="sc"><div class="sv">{summary["total_output_rows"]:,}</div><div class="sl">Total filas output</div></div>
  </div>
</div>

<div class="sec">
  <div class="sn">01 — Salidas canceladas</div>
  <h2>Destinos eliminados del sorter esta semana</h2>
  <div class="sd">{summary["n_canceladas"]} salidas canceladas — {summary["rows_removed_cancelled"]:,} filas eliminadas.</div>
  <table>
    <thead><tr><th>Día original</th><th>Destino</th><th>Estado</th><th>Nota</th></tr></thead>
    <tbody>{cancel_rows()}</tbody>
  </table>
</div>

<div class="sec">
  <div class="sn">02 — Asignación automática de rampas (con control de horarios)</div>
  <h2>Destinos reasignados a nuevo día</h2>
  <div class="sd">{len(ok)} destinos asignados a posiciones libres en el nuevo día.
  El bloque se deriva del ID_CLUSTER del destino en el nuevo día.
  Si el origen tenía bloque desconocido (#N/A), se indica el bloque calculado.</div>
  {'<p class="mt" style="margin:10px 0">— Sin reasignaciones automáticas —</p>' if not ok else f"""
  <table>
    <thead><tr><th>Nuevo día</th><th>Destino</th><th>Día orig.</th><th>Bloque</th><th>Rampas y posiciones</th></tr></thead>
    <tbody>{ok_rows()}</tbody>
  </table>"""}
</div>

<div class="sec">
  <div class="sn">03 — Rutas E2, sin configuración y asignaciones parciales</div>
  <h2>Casos que requieren atención</h2>
  <div class="sd">
    {len(e2)} rutas E2 (MAN/EXDOCK — sin acción en sorter) ·
    {len(nocfg)} sin config GD · {len(partial)} parciales.
  </div>
  {'<p class="mt" style="margin:10px 0">✓ Sin casos pendientes.</p>' if not (n_warn + len(e2)) else
   ''.join([f'<div class="ib">🔀 <strong>{r["playa"]}</strong> ({r["dia_orig"]} → {r["dia_new"]}): {r["msg"]}</div>' for r in e2])
   + ''.join([f'<div class="wb">❌ <strong>{r["playa"]}</strong> ({r["dia_orig"]} → {r["dia_new"]}): {r["msg"]}</div>' for r in nocfg])
   + ''.join([f'<div class="wb">⚠ <strong>{r["playa"]}</strong>: Solo {r["n_assigned"]}/{r["n_destinos"]} posiciones en {r["dia_new"]}</div>' for r in partial])
   + f"""
  <table>
    <thead><tr><th>Día orig.</th><th>Destino</th><th>Nuevo día</th><th>Tipo</th><th>Detalle</th></tr></thead>
    <tbody>{warn_rows()}</tbody>
  </table>"""}
</div>

</div>
<div class="rf">
  <span>© Logística VDL B2B · Estrictamente confidencial</span>
  <span>Generado {datetime.now().strftime("%d/%m/%Y")}</span>
</div>
</div>
<div class="np"><button class="bp" onclick="window.print()">Imprimir / Exportar PDF</button></div>
</body></html>'''

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)



def write_especiales_gd(especial_rows, header, out_path):
    """GD xlsx with only the new especial rows (rows to ADD in DXC)."""
    write_gd(especial_rows, header, out_path)


def write_canceladas_txt(canceladas, filter_days, out_path):
    """Plain-text list of cancelled playas per day (rows to DELETE from DXC)."""
    DAY_ORDER = ['DOMINGO','LUNES','MARTES','MIERCOLES','JUEVES','VIERNES','SABADO']
    by_day = defaultdict(list)
    for (dia, playa) in canceladas:
        if filter_days and CLUSTER_DAY.get(list(filter_days)[0] if filter_days else '',dia) and dia not in [CLUSTER_DAY.get(l,l) for l in filter_days]:
            continue
        by_day[dia].append(playa)
    lines = ['CANCELADAS — SALIDAS A ELIMINAR DEL SORTER', '=' * 44, '']
    for dia in DAY_ORDER:
        playas = sorted(by_day.get(dia, []))
        if playas:
            lines.append(f'{dia} ({len(playas)}):')
            for p in playas:
                lines.append(f'  - {p}')
            lines.append('')
    total = sum(len(v) for v in by_day.values())
    lines.append(f'Total: {total} salidas canceladas')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 5:
        print(__doc__); sys.exit(1)

    parrilla_path, gd_path, cap_path, sheet_name = sys.argv[1:5]
    semana   = sys.argv[5] if len(sys.argv) > 5 else sheet_name.upper()
    gd_out   = sys.argv[6] if len(sys.argv) > 6 else os.path.join(
                   tempfile.mkdtemp(), f'GRUPO_DESTINOS_{semana}.xlsx')
    html_out = sys.argv[7] if len(sys.argv) > 7 else os.path.join(
                   os.path.dirname(gd_out), f'resumen_sorter_{semana}.html')

    print(f"Cargando capacidad: {cap_path}")
    capacity = load_capacity(cap_path)
    print(f"  → {len(capacity)} rampas")

    print(f"Cargando timings de bloques desde parrilla: {parrilla_path}")
    bloque_timings = load_bloque_timings(parrilla_path)
    print(f"  → {len(bloque_timings)} bloques con horario")

    print(f"Cargando GRUPO_DESTINOS: {gd_path}")
    gd_header, tagged, by_dia_playa = load_grupo_destinos(gd_path)
    print(f"  → {len(tagged):,} filas, {len(by_dia_playa)} combinaciones día×playa")

    print(f"Cargando parrilla '{sheet_name}': {parrilla_path}")
    parrilla = load_parrilla(parrilla_path, sheet_name)
    print(f"  → {len(parrilla)} destinos")

    print("Procesando y asignando rampas (con control horario de bloques)...")
    filter_days = None
    if len(sys.argv) > 8 and sys.argv[8].strip():
        filter_days = set(d.strip().upper() for d in sys.argv[8].split(',') if d.strip())
        print(f"  Filtrando por días: {sorted(filter_days)}")
    superplaya_path = sys.argv[9] if len(sys.argv) > 9 else None
    superplaya_map  = load_superplaya(superplaya_path)
    if superplaya_map:
        print(f"  Superplayas cargadas: {len(set(superplaya_map.values()))} grupos")
    especial_bloque_map = load_especial_bloque_map(parrilla_path)
    cancelled_especiales = load_cancelled_especiales(parrilla_path)
    print(f"  Canceladas en SEMANA SANTA: {len(cancelled_especiales)} playas")
    if especial_bloque_map:
        print(f"  Bloque map (SEMANA SANTA): {len(especial_bloque_map)} entradas")
    output_rows, summary = process(parrilla, tagged, by_dia_playa, capacity, bloque_timings, filter_days, superplaya_map, especial_bloque_map, cancelled_especiales)

    ok      = [r for r in summary['assignment_results'] if r['status'] == 'OK']
    partial = [r for r in summary['assignment_results'] if r['status'] == 'PARTIAL']
    e2      = [r for r in summary['assignment_results'] if r['status'] == 'E2_ROUTE']
    nocfg   = [r for r in summary['assignment_results'] if r['status'] == 'NO_CONFIG']

    print(f"\n{'='*65}")
    print(f"RESUMEN {semana}:")
    print(f"  Canceladas:             {summary['n_canceladas']} ({summary['rows_removed_cancelled']:,} filas)")
    print(f"  Cambios de día:         {summary['n_especiales']}")
    print(f"    → Asignados OK:       {len(ok)}")
    print(f"    → Rutas E2:           {len(e2)}")
    print(f"    → Sin config GD:      {len(nocfg)}")
    print(f"    → Parciales:          {len(partial)}")
    print(f"  Filas añadidas:         {summary['rows_added_especial']:,}")
    print(f"  Filas output total:     {summary['total_output_rows']:,}")
    print(f"{'='*65}")

    if ok:
        print("\nAsignaciones completadas:")
        for r in ok:
            fb   = f" [datos {r['source_day']}]" if r.get('source_day') and r['source_day'] != r['dia_orig'] else ''
            blq  = r.get('bloque_new','?')
            rstr = ', '.join(f"{k}({len(v)}p)" for k,v in sorted(r['rampas'].items()))
            print(f"  {r['playa']:42s} {r['dia_orig']}→{r['dia_new']} [{blq}]{fb}  {rstr}")

    if e2:
        print("\nRutas E2 (sin cambio en GD):")
        for r in e2:
            print(f"  {r['playa']:42s} {r['dia_orig']}→{r['dia_new']}  🔀 E2/manual")

    if nocfg:
        print("\n❌ Sin config GD:")
        for r in nocfg:
            print(f"  {r.get('playa', r.get('orig_playa','?')):42s} sin datos en ningún día")

    if partial:
        print("\n⚠ Parciales:")
        for r in partial:
            print(f"  {r['playa']:42s} {r['n_assigned']}/{r['n_destinos']} posiciones")

    # Post-process: rename any row whose playa is in cancelled_especiales
    # This catches cases like GUARROMAN_TSA with non-standard desc format
    if cancelled_especiales:
        import re as _re_post
        def _rename_if_cancelled(rows):
            out = []
            for _r in rows:
                _desc = str(_r[2] or '')
                _renamed = _desc
                for _cp in cancelled_especiales:
                    if _cp in _desc.upper() and '_CANCELADA_SOLO_W14' not in _desc:
                        # Find exact case match
                        _m = _re_post.search(_re_post.escape(_cp), _desc, _re_post.IGNORECASE)
                        if _m:
                            _match_str = _m.group(0)
                            _renamed = _desc.replace(_match_str, _match_str + '_CANCELADA_SOLO_W14', 1)
                            break
                out.append((_r[0], _r[1], _renamed, _r[3], _r[4], _r[5], _r[6]))
            return out
        output_rows = _rename_if_cancelled(output_rows)
        if summary.get('especial_rows'):
            summary['especial_rows'] = _rename_if_cancelled(summary['especial_rows'])

    print(f"\nEscribiendo GD  → {gd_out}")
    write_gd(output_rows, gd_header, gd_out)

    # Especiales-only GD (rows to ADD in DXC)
    if summary.get('especial_rows'):
        esp_out = str(gd_out).replace('.xlsx', '_SOLO_ESPECIALES.xlsx')
        write_especiales_gd(summary['especial_rows'], gd_header, esp_out)
        print(f"  + Especiales    → {esp_out}")

    # Canceladas list (rows to DELETE)
    can_out = str(gd_out).replace('.xlsx', '_CANCELADAS.txt')
    write_canceladas_txt(summary['canceladas'], filter_days, can_out)
    print(f"  + Canceladas    → {can_out}")
    print(f"Escribiendo HTML → {html_out}")
    write_html(summary, semana, html_out)
    # Enrich with interactive chart and compact cancel list
    with open(html_out, encoding='utf-8') as _f:
        _base = _f.read()
    _enriched = _enrich_html(_base, output_rows, summary)
    with open(html_out, 'w', encoding='utf-8') as _f:
        _f.write(_enriched)

    # outputs written to gd_out and html_out (paths from argv or tempdir)
    print(f"\n✓ Listo.")

if __name__ == '__main__':
    main()
