# Version: 0.10
import streamlit as st
import subprocess, sys, tempfile, datetime as dt
from pathlib import Path

BASE_DIR = Path(__file__).parent

# ── Hojas base que NO son hojas de evento especial ───────────────────────────
_BASE_SHEET_NAMES = {
    "B2B", "B2C", "BLOQUES", "RESUMEN BLOQUES", "VOLUMENES", "APY",
}

def _is_event_sheet(sheet_name: str, headers: tuple) -> bool:
    """Return True if this sheet is an event/agenda sheet (not a base sheet)."""
    su = sheet_name.strip().upper()
    # Exclude known base sheet names
    if su in _BASE_SHEET_NAMES:
        return False
    # Exclude "Bloques *" sheets (sub-tables embedded in parrilla)
    if su.startswith("BLOQUES"):
        return False
    # Must have TIPO_SALIDA column
    if not any(str(h or "").strip().upper() == "TIPO_SALIDA" for h in headers):
        return False
    return True

def _detect_bloques_sheet(wb, event_sheet: str) -> str | None:
    """
    Find the matching Bloques sheet for a given event sheet.

    Strategy (in order):
      1. Exact suffix: 'Bloques ' + everything after AGENDA/SEMANA prefix
      2. Week-code match: if the event sheet contains a week code like 'S26',
         pick the Bloques sheet that contains the same code ('Bloques S26').
         This is critical because the parrilla has several Bloques sheets
         (Bloques, Bloques 25 de MAYO, Bloques S26) and the generic 'Bloques'
         must NOT be picked for an S26 event.
      3. Generic fallback: any 'Bloques*' sheet, but only the plain base
         'Bloques' as a last resort.
    """
    import re as _re
    sheets = wb.sheetnames
    su = event_sheet.strip()

    # 1. Exact suffix match
    for prefix in ("AGENDA ", "SEMANA SANTA ", "SEMANA "):
        if su.upper().startswith(prefix.upper()):
            suffix = su[len(prefix):].strip()
            candidate = f"Bloques {suffix}"
            if candidate in sheets:
                return candidate
            candidate_up = candidate.upper()
            for s in sheets:
                if s.upper() == candidate_up:
                    return s
            break

    # 2. Week-code match: extract S<NN> from the event sheet name
    _m = _re.search(r'\bS(\d{1,2})\b', su, _re.IGNORECASE)
    if _m:
        code = f"S{_m.group(1)}"
        code_up = code.upper()
        for s in sheets:
            su2 = s.strip().upper()
            if su2.startswith("BLOQUES") and su2 != "RESUMEN BLOQUES":
                # match the same week code as a whole token
                if _re.search(r'\bS' + _m.group(1) + r'\b', s, _re.IGNORECASE):
                    return s

    # 3. Last resort: a Bloques* sheet, but prefer one that is NOT the plain
    #    base 'Bloques' (which carries the normal-week timings).
    _specific = None
    _generic  = None
    for s in sheets:
        su2 = s.strip().upper()
        if su2 == "RESUMEN BLOQUES":
            continue
        if su2 == "BLOQUES":
            _generic = s
        elif su2.startswith("BLOQUES"):
            if _specific is None:
                _specific = s
    return _specific or _generic

def _detect_semana(sheet_name: str, parrilla_bytes: bytes | None, selected_sheet: str) -> str:
    """
    Auto-detect semana number. Priority:
    1. S+digits anywhere in the sheet name (e.g. "AGENDA S26 Sant Joan" → "S26")
    2. SEMANA column value from first data row
    3. Fallback to "S??"
    """
    import re as _re
    m = _re.search(r'\bS?(\d{2})\b', sheet_name, _re.IGNORECASE)
    if m:
        n = m.group(1)
        # Avoid matching year-like numbers (2024, 2025...)
        if int(n) <= 53:
            return f"S{n}"
    # Try reading SEMANA column from data
    if parrilla_bytes and selected_sheet:
        try:
            import io
            from openpyxl import load_workbook as _lwb
            _wb = _lwb(io.BytesIO(parrilla_bytes), read_only=True)
            if selected_sheet in _wb.sheetnames:
                _ws = _wb[selected_sheet]
                rows_iter = _ws.iter_rows(values_only=True)
                hdr = next(rows_iter, ())
                col_map = {str(h or "").strip().upper(): i for i, h in enumerate(hdr) if h}
                sem_idx = col_map.get("SEMANA")
                if sem_idx is not None:
                    for data_row in rows_iter:
                        val = data_row[sem_idx] if sem_idx < len(data_row) else None
                        if val is not None:
                            try:
                                n = int(float(str(val)))
                                if 1 <= n <= 53:
                                    return f"S{n:02d}"
                            except (ValueError, TypeError):
                                pass
                        break  # only need first data row
        except Exception:
            pass
    return "S??"


# ── Validation panel ──────────────────────────────────────────────────────────
def _run_validation(parrilla_bytes, gd_bytes=None, sheet_name=None):
    """Run validate_parrilla.validate() safely; return issues list."""
    try:
        sys.path.insert(0, str(BASE_DIR))
        from validate_parrilla import validate
        return validate(parrilla_bytes, gd_bytes, sheet_name=sheet_name)
    except Exception as e:
        return [{"severity": "warning", "category": "estructura",
                 "title": "No se pudo ejecutar la validación previa",
                 "detail": str(e), "items": [], "autocorrected": False}]

def render_validation(parrilla_file, gd_file=None, sheet_name=None):
    """
    Run validation and render the results panel in Streamlit.
    Always returns True (never blocks generation).
    """
    par_bytes = parrilla_file.read(); parrilla_file.seek(0)
    gd_bytes  = None
    if gd_file:
        gd_bytes = gd_file.read(); gd_file.seek(0)

    issues = _run_validation(par_bytes, gd_bytes, sheet_name=sheet_name)

    # Count severities
    counts = {"error": 0, "warning": 0, "info": 0, "ok": 0}
    for iss in issues:
        counts[iss["severity"]] = counts.get(iss["severity"], 0) + 1

    # Header badge
    if counts["error"] > 0:
        badge = f"🔴 {counts['error']} error(s)"
        if counts["warning"]: badge += f"  ·  ⚠ {counts['warning']} aviso(s)"
    elif counts["warning"] > 0:
        badge = f"⚠ {counts['warning']} aviso(s)"
    else:
        badge = "✅ Todo OK"

    with st.expander(f"🔍 Validación previa — {badge}", expanded=(counts["error"] > 0 or counts["warning"] > 0)):
        # Group by category
        for cat_key, cat_label in [("estructura", "Estructura del fichero"),
                                    ("contenido",  "Contenido leído"),
                                    ("cobertura",  "Cobertura especiales en GD")]:
            cat_issues = [i for i in issues if i["category"] == cat_key]
            if not cat_issues:
                continue
            st.markdown(f"**{cat_label}**")
            for iss in cat_issues:
                sev = iss["severity"]
                icon = {"error": "🔴", "warning": "⚠️", "info": "ℹ️", "ok": "✅"}.get(sev, "·")
                title = iss["title"]
                if iss.get("autocorrected"):
                    title += "  *(autocorrección aplicada)*"

                if sev == "error":
                    st.error(f"{icon} **{title}**\n\n{iss['detail']}")
                elif sev == "warning":
                    st.warning(f"{icon} **{title}**\n\n{iss['detail']}")
                elif sev == "info":
                    st.info(f"{icon} **{title}**\n\n{iss['detail']}")
                else:
                    st.success(f"{icon} {title}")

                if iss.get("items"):
                    items_md = "  \n".join(f"- `{it}`" for it in iss["items"])
                    st.markdown(items_md)

            st.markdown("")  # spacing between categories

    return True  # never blocks

def _render_output_validation(parrilla_file, gd_output_bytes: bytes, sheet_name=None):
    """
    Post-generation validation: cross-check the produced GD against the parrilla.
    Shows results in an expander. Never blocks.
    """
    try:
        sys.path.insert(0, str(BASE_DIR))
        from validate_parrilla import validate_output, summary
        par_bytes = parrilla_file.read(); parrilla_file.seek(0)
        issues = validate_output(par_bytes, gd_output_bytes, sheet_name=sheet_name)
    except Exception as e:
        st.warning(f"⚠️ No se pudo ejecutar la validación del resultado: {e}")
        return

    counts = {"error": 0, "warning": 0, "info": 0, "ok": 0}
    for iss in issues:
        counts[iss["severity"]] = counts.get(iss["severity"], 0) + 1

    if counts["error"] > 0:
        badge = f"🔴 {counts['error']} error(s)"
        if counts["warning"]: badge += f"  ·  ⚠ {counts['warning']} aviso(s)"
        expanded = True
    elif counts["warning"] > 0:
        badge = f"⚠ {counts['warning']} aviso(s)"
        expanded = True
    else:
        badge = "✅ Sort map OK"
        expanded = False

    with st.expander(f"✅ Validación del resultado — {badge}", expanded=expanded):
        for iss in issues:
            sev  = iss["severity"]
            icon = {"error": "🔴", "warning": "⚠️", "info": "ℹ️", "ok": "✅"}.get(sev, "·")
            _is_tabla = "Resumen por dia" in iss["title"]
            if sev == "error":
                st.error(f"{icon} **{iss['title']}**\n\n{iss['detail']}")
            elif sev == "warning":
                if _is_tabla:
                    st.warning(f"{icon} **{iss['title']}**")
                    st.code(iss['detail'], language=None)
                else:
                    st.warning(f"{icon} **{iss['title']}**\n\n{iss['detail']}")
            elif sev == "info":
                st.info(f"{icon} **{iss['title']}**\n\n{iss['detail']}")
            else:
                if _is_tabla:
                    st.success(f"{icon} **{iss['title']}**")
                    st.code(iss['detail'], language=None)
                else:
                    st.success(f"{icon} {iss['title']}")
            if iss.get("items"):
                st.markdown("\n".join(f"- `{it}`" for it in iss["items"]))

def gd_to_dxc_csv(xlsx_bytes):
    # Convert GD xlsx to DXC upload CSV format (POSTEX + SOREXP separately)
    import io as _io
    from openpyxl import load_workbook as _lwb
    wb = _lwb(_io.BytesIO(xlsx_bytes), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))[1:]
    postex_lines, sorexp_lines = [], []
    for r in rows:
        if len(r) < 7: continue
        desc = str(r[2] or "").strip()
        tipo = str(r[3] or "").strip().upper()
        dest_raw = str(r[4] or "").strip()
        elem = str(r[6] or "").strip()
        if not desc or tipo not in ("POSTEX","SOREXP") or not dest_raw or not elem: continue
        if desc.startswith("="): continue
        d = "".join(c for c in dest_raw if c.isdigit())
        dest8 = d.zfill(10)[-8:] if d else dest_raw[:8]
        line = f"{desc};{dest8};00;{elem}"
        if tipo == "POSTEX": postex_lines.append(line + ";20")
        else: sorexp_lines.append(line + ";10")
    bom = "\ufeff"
    def _enc(lines): return (bom + "\r\n".join(lines)).encode("utf-8")
    return _enc(postex_lines), _enc(sorexp_lines)


st.set_page_config(page_title="Sorter VDL B2B", page_icon="🏭", layout="centered")
st.markdown("<style>.block-container{max-width:780px}</style>", unsafe_allow_html=True)

st.markdown("## 🏭 Sorter VDL B2B")
st.markdown("Configurador de semanas especiales — VDL B2B")
st.divider()

# ── Session state ─────────────────────────────────────────────────────────────
for key in ["r1_gd","r1_esp","r1_can","r1_html","r2_gantt","r3_map","r1_day_filter","r1_postex_csv","r1_sorexp_csv","r1_esp_postex_csv","r1_esp_sorexp_csv"]:
    if key not in st.session_state:
        st.session_state[key] = None

# ── Inputs ────────────────────────────────────────────────────────────────────
st.markdown("### Ficheros de entrada")
col1, col2 = st.columns(2)
with col1:
    f_parrilla = st.file_uploader("Parrilla de salidas", type=["xlsx"],
                                   help="Debe incluir la hoja con TIPO_SALIDA y Resumen Bloques")

    # Dynamic sheet selector: read sheets from uploaded file
    _embedded_bloques_sheet = None  # will be set below if detected
    if f_parrilla:
        import io as _io2
        from openpyxl import load_workbook as _lwb2
        _par_bytes_tmp = f_parrilla.read()
        f_parrilla.seek(0)
        _wb_tmp = _lwb2(_io2.BytesIO(_par_bytes_tmp), read_only=True)
        _all_sheets = _wb_tmp.sheetnames

        # Filter to event sheets only (exclude base B2B, B2C, Bloques*, etc.)
        _agenda_first = []  # AGENDA/SEMANA sheets (highest priority)
        _other_event  = []  # other event sheets

        for _sh in _all_sheets:
            _ws_tmp = _wb_tmp[_sh]
            _first  = next(_ws_tmp.iter_rows(values_only=True, max_row=1), None)
            if _first is None:
                continue
            if not _is_event_sheet(_sh, _first):
                continue
            _sh_up = _sh.strip().upper()
            if _sh_up.startswith("AGENDA") or _sh_up.startswith("SEMANA SANTA") or _sh_up.startswith("SEMANA "):
                _agenda_first.append(_sh)
            else:
                _other_event.append(_sh)

        _valid = _agenda_first + _other_event

        # Fallback: if no event sheet found, show all with TIPO_SALIDA
        if not _valid:
            for _sh in _all_sheets:
                _ws_tmp = _wb_tmp[_sh]
                _first  = next(_ws_tmp.iter_rows(values_only=True, max_row=1), None)
                if _first and any(str(h or "").strip().upper() == "TIPO_SALIDA" for h in _first):
                    _valid.append(_sh)

        _options = _valid if _valid else _all_sheets
        _default_idx = 0  # Best candidate is first (AGENDA sheets are first)

        sheet = st.selectbox("Hoja de parrilla", options=_options, index=_default_idx,
                             help="Selecciona la pestaña con los datos del evento especial (AGENDA...)")

        # Detect embedded bloques sheet for the selected event
        _embedded_bloques_sheet = _detect_bloques_sheet(_wb_tmp, sheet)
    else:
        sheet = st.text_input("Nombre de hoja", value="AGENDA S26 Sant Joan",
                              help="Sube la parrilla para ver las hojas disponibles")

    # Auto-detect semana
    _par_bytes_for_sem = f_parrilla.read() if f_parrilla else None
    if f_parrilla: f_parrilla.seek(0)
    _sem_default = _detect_semana(sheet, _par_bytes_for_sem, sheet if f_parrilla else "")
    semana = st.text_input("Semana", value=_sem_default,
                           help="Se autodetecta del nombre de hoja o de la columna SEMANA")

with col2:
    f_gd  = st.file_uploader("GRUPO_DESTINOS", type=["xlsx"],
                               help="Export DXC o fichero clásico")
    f_cap = st.file_uploader("Capacidad de rampas", type=["csv"],
                               help="CSV con columnas RAMP;PALLETS")

# Bloques horarios: optional if embedded in parrilla, required otherwise
_bloques_help = "Columnas: NUEVO BLOQUE · Día LIBERACIÓN · Hora LIBERACIÓN · Día DESACTIVACIÓN · Hora DESACTIVACIÓN"
if _embedded_bloques_sheet:
    _bloques_caption = f"ℹ️ Se usará la hoja **{_embedded_bloques_sheet}** de la parrilla automáticamente. Sube otro fichero solo si quieres sobreescribir."
    f_bloques = st.file_uploader(
        f"Bloques horarios *(detectado: {_embedded_bloques_sheet} en parrilla)*",
        type=["xlsx"],
        help=_bloques_help)
    if f_bloques is None:
        st.caption(_bloques_caption)
else:
    f_bloques = st.file_uploader(
        "Bloques horarios *(necesario para Gantt y Sorter Map)*", type=["xlsx"],
        help=_bloques_help)

f_superplaya = st.file_uploader(
    "Superplayas *(opcional — mejora la agrupación de rampas)*", type=["xlsx"],
    help="Columnas: AGRUPACION_PLAYA · SUPERPLAYA — define qué destinos deben ir juntos en rampas contiguas")

st.divider()

# ── Helpers ───────────────────────────────────────────────────────────────────
ALL_DAYS = ["DOMINGO","LUNES","MARTES","MIERCOLES","JUEVES","VIERNES","SABADO"]

BLOQUE_LETRA_DAY = {
    "D": "DOMINGO",  "L": "LUNES",  "M": "MARTES",
    "X": "MIERCOLES", "J": "JUEVES", "V": "VIERNES", "S": "SABADO"
}
BLOQUE_OPTIONS = ["D (Domingo)","L (Lunes)","M (Martes)",
                  "X (Miercoles)","J (Jueves)","V (Viernes)","S (Sabado)"]

def save_uploads(tmp: Path):
    p = {}

    # Parrilla
    if f_parrilla:
        path = tmp / "parrilla.xlsx"
        path.write_bytes(f_parrilla.read())
        f_parrilla.seek(0)
        p["parrilla"] = path

    # GD
    if f_gd:
        path = tmp / "gd.xlsx"
        path.write_bytes(f_gd.read())
        f_gd.seek(0)
        p["gd"] = path

    # Cap
    if f_cap:
        path = tmp / "cap.csv"
        path.write_bytes(f_cap.read())
        f_cap.seek(0)
        p["cap"] = path

    # Bloques: explicit upload overrides embedded sheet
    if f_bloques:
        path = tmp / "bloques.xlsx"
        path.write_bytes(f_bloques.read())
        f_bloques.seek(0)
        p["bloques"] = path
    elif _embedded_bloques_sheet and "parrilla" in p:
        # Extract the embedded Bloques sheet from the parrilla file
        try:
            import io as _io_bl
            import openpyxl as _opx_bl
            _wb_bl = _opx_bl.load_workbook(str(p["parrilla"]), read_only=True)
            if _embedded_bloques_sheet in _wb_bl.sheetnames:
                # Write a new xlsx with only the Bloques sheet
                _wb_out = _opx_bl.Workbook()
                _ws_src = _wb_bl[_embedded_bloques_sheet]
                _ws_dst = _wb_out.active
                _ws_dst.title = _embedded_bloques_sheet
                for row in _ws_src.iter_rows(values_only=True):
                    _ws_dst.append([c for c in row])
                path_bl = tmp / "bloques.xlsx"
                _wb_out.save(str(path_bl))
                p["bloques"] = path_bl
        except Exception as _bl_err:
            pass  # If extraction fails, proceed without bloques

    # Superplaya
    if f_superplaya:
        path = tmp / "superplaya.xlsx"
        path.write_bytes(f_superplaya.read())
        f_superplaya.seek(0)
        p["superplaya"] = path

    return p

def run_gd(p, tmp, sc, days_arg=""):
    gd   = tmp / f"GRUPO_DESTINOS_{sc}.xlsx"
    html = tmp / f"resumen_{sc}.html"
    cmd  = [sys.executable, str(BASE_DIR / "process_parrilla.py"),
            str(p["parrilla"]), str(p["gd"]), str(p["cap"]),
            sheet.strip(), sc, str(gd), str(html)]
    cmd.append(days_arg)  # argv[8] — always pass (empty string = no filter)
    if "superplaya" in p:
        cmd.append(str(p["superplaya"]))  # argv[9]
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
    return gd, html, r

def show_log(r, expanded=False):
    lines = [l for l in (r.stdout + r.stderr).splitlines() if l.strip()]
    with st.expander("Ver log", expanded=expanded or r.returncode != 0):
        for line in lines:
            if "✓" in line:                                           st.success(line)
            elif "❌" in line:                                         st.error(line)
            elif "⚠" in line or "E2" in line or "Sin config" in line: st.warning(line)
            else:                                                       st.text(line)
    return lines

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ── Action buttons — new flow: Sort Map first, then GD ───────────────────────
st.markdown("### Acciones")

# bloques available = either uploaded or embedded in parrilla
_bloques_available = bool(f_bloques or _embedded_bloques_sheet)

base_ok      = bool(f_parrilla and f_gd and f_cap)
vis_ok       = bool(base_ok and _bloques_available)
sortmap_done = bool(st.session_state.get("r3_map"))
gd_ok        = bool(sortmap_done and base_ok)

b1, b2, b3 = st.columns(3)
with b1:
    st.markdown("**1 · Sorter Map por día**")
    st.caption("Asigna especiales · valida visualmente")
    if st.button("🗺 Generar", key="go3", type="primary",
                 disabled=not vis_ok, use_container_width=True):
        st.session_state["r3_map"] = None
        st.session_state["r3_gd_bytes"] = None
        st.session_state["_run3"] = True
    if not vis_ok:
        if not base_ok:
            st.caption("_Sube parrilla, GD y capacidad_")
        else:
            st.caption("_Requiere bloques horarios_")

with b2:
    st.markdown("**2 · Configuración DXC**")
    st.caption("Descarga el nuevo GD con los cambios")
    if st.button("⚙️ Generar", key="go1", type="primary",
                 disabled=not gd_ok, use_container_width=True):
        for k in ["r1_gd","r1_esp","r1_can","r1_html","r1_day_filter","r1_gd_filtered_bytes","r3_gd_bytes"]:
            st.session_state[k] = None
        st.session_state["_run1"] = True
    if not gd_ok:
        if not base_ok:
            st.caption("_Sube parrilla, GD y capacidad_")
        elif not sortmap_done:
            st.caption("_Genera el Sort Map primero_")

with b3:
    st.markdown("**3 · Gantt 1H**")
    st.caption("Visual bloques × rampas × hora")
    if st.button("📊 Generar", key="go2", type="primary",
                 disabled=not vis_ok, use_container_width=True):
        st.session_state["r2_gantt"] = None
        st.session_state["_run2"] = True
    if not vis_ok:
        if not base_ok:
            st.caption("_Sube parrilla, GD y capacidad_")
        else:
            st.caption("_Requiere bloques horarios_")

st.divider()

# ── Execute action 1 ──────────────────────────────────────────────────────────
if st.session_state.get("_run1"):
    st.session_state["_run1"] = False
    sc = semana.strip() or sheet.strip().upper()
    render_validation(f_parrilla, f_gd, sheet_name=sheet.strip())
    with tempfile.TemporaryDirectory() as _tmp:
        tmp = Path(_tmp)
        p   = save_uploads(tmp)
        _r3_gd = st.session_state.get("r3_gd_bytes")
        if _r3_gd:
            gd = tmp / f"GRUPO_DESTINOS_{sc}.xlsx"
            gd.write_bytes(_r3_gd)
            html = tmp / f"resumen_sorter_{sc}.html"
            with st.spinner("Generando resumen HTML…"):
                r = subprocess.run(
                    [sys.executable, str(BASE_DIR / "process_parrilla.py"),
                     str(p["parrilla"]), str(p["gd"]), str(p["cap"]),
                     sheet.strip(), sc, str(gd), str(html), ""],
                    capture_output=True, text=True, timeout=180)
        else:
            with st.spinner("Procesando parrilla y asignando rampas…"):
                gd, html, r = run_gd(p, tmp, sc)
        show_log(r)
        if r.returncode != 0 or not gd.exists():
            st.error("El proceso terminó con error.")
        else:
            esp_path = Path(str(gd).replace('.xlsx', '_SOLO_ESPECIALES.xlsx'))
            can_path = Path(str(gd).replace('.xlsx', '_CANCELADAS.txt'))
            _gd_bytes = gd.read_bytes()
            st.session_state["r1_gd"]    = (gd.name, _gd_bytes)
            st.session_state["r1_esp"]   = (esp_path.name, esp_path.read_bytes()) if esp_path.exists() else None
            _px, _sx = gd_to_dxc_csv(_gd_bytes)
            st.session_state["r1_postex_csv"] = (gd.stem + "_POSTEX.csv", _px)
            st.session_state["r1_sorexp_csv"] = (gd.stem + "_SOREXP.csv", _sx)
            if esp_path.exists():
                _epx, _esx = gd_to_dxc_csv(esp_path.read_bytes())
                st.session_state["r1_esp_postex_csv"] = (esp_path.stem + "_POSTEX.csv", _epx)
                st.session_state["r1_esp_sorexp_csv"] = (esp_path.stem + "_SOREXP.csv", _esx)
            st.session_state["r1_can"]   = (can_path.name, can_path.read_text(encoding='utf-8')) if can_path.exists() else None
            st.session_state["r1_html"]  = (html.name,  html.read_bytes()) if html.exists() else None
            st.session_state["r1_day_filter"] = None
            _render_output_validation(f_parrilla, _gd_bytes, sheet_name=sheet.strip())

# ── Execute action 2 ──────────────────────────────────────────────────────────
if st.session_state.get("_run2"):
    st.session_state["_run2"] = False
    sc = semana.strip() or sheet.strip().upper()
    render_validation(f_parrilla, f_gd, sheet_name=sheet.strip())
    with tempfile.TemporaryDirectory() as _tmp:
        tmp = Path(_tmp)
        p   = save_uploads(tmp)
        with st.spinner("Generando GD base…"):
            gd, _, r0 = run_gd(p, tmp, sc)
        if r0.returncode != 0 or not gd.exists():
            st.error("Error generando GD base.")
            show_log(r0, expanded=True)
        else:
            out = tmp / f"gantt_1h_{sc}.xlsx"
            with st.spinner("Generando Gantt 1H…"):
                r = subprocess.run(
                    [sys.executable, str(BASE_DIR / "gantt_1h.py"),
                     str(p["cap"]), str(gd), str(p["bloques"]), str(out), "Hoja1"],
                    capture_output=True, text=True, timeout=180)
            show_log(r)
            if r.returncode == 0 and out.exists():
                st.session_state["r2_gantt"] = (out.name, out.read_bytes())
            else:
                st.error("El Gantt terminó con error.")

# ── Execute action 3 ──────────────────────────────────────────────────────────
if st.session_state.get("_run3"):
    st.session_state["_run3"] = False
    sc = semana.strip() or sheet.strip().upper()
    render_validation(f_parrilla, f_gd, sheet_name=sheet.strip())
    with tempfile.TemporaryDirectory() as _tmp:
        tmp = Path(_tmp)
        p   = save_uploads(tmp)
        _filtered_bytes = st.session_state.get("r1_gd_filtered_bytes")
        _filter_days    = st.session_state.get("r1_day_filter")
        if _filtered_bytes:
            gd = tmp / f"GD_{sc}_filtered.xlsx"
            gd.write_bytes(_filtered_bytes)
            r0_ok = True
            st.info(f"Usando GD filtrado: {', '.join(_filter_days or [])}")
        else:
            with st.spinner("Generando GD base…"):
                gd, _, r0 = run_gd(p, tmp, sc)
            r0_ok = r0.returncode == 0 and gd.exists()
            if not r0_ok:
                st.error("Error generando GD base.")
                show_log(r0, expanded=True)
        if r0_ok:
            st.session_state["r3_gd_bytes"] = gd.read_bytes()
            out = tmp / f"sorter_map_{sc}.xlsx"
            with st.spinner("Generando Sorter Map…"):
                _sorter_cmd = [
                    sys.executable, str(BASE_DIR / "sorter_map_por_dia.py"),
                    str(p["cap"]), str(gd), str(p["bloques"]), str(out), "Hoja1",
                ]
                if "parrilla" in p and "gd" in p:
                    _sorter_cmd += [str(p["parrilla"]), sheet.strip(), str(p["gd"])]
                r = subprocess.run(_sorter_cmd, capture_output=True, text=True, timeout=180)
            show_log(r)
            if r.returncode == 0 and out.exists():
                st.session_state["r3_map"] = (out.name, out.read_bytes())
            else:
                st.error("El Sorter Map terminó con error.")

# ── Show results action 1 ─────────────────────────────────────────────────────
if st.session_state["r1_gd"] is not None:
    gd_name, gd_bytes = st.session_state["r1_gd"]
    sc = semana.strip() or sheet.strip().upper()
    st.success(f"✓ Configuración {sc} generada")

    with st.expander("🔍 Filtrar por bloque y regenerar especiales"):
        selected_bloques = st.multiselect(
            "Bloques a incluir en el GD filtrado",
            options=BLOQUE_OPTIONS,
            default=[],
            key="day_filter_sel",
            placeholder="Selecciona bloques…",
            help="Cada letra = familia de bloques activos ese día (D=Dom, L=Lun, M=Mar, X=Mie, J=Jue, V=Vie, S=Sab)",
        )
        selected_days = [b.split()[0] for b in selected_bloques]
        if st.button("⚙️ Regenerar con filtro", key="regen_filter",
                     disabled=not (selected_bloques and f_parrilla and f_gd and f_cap)):
            days_arg = ",".join(selected_days)
            with tempfile.TemporaryDirectory() as _tmp:
                tmp = Path(_tmp)
                p   = save_uploads(tmp)
                with st.spinner(f"Regenerando para {', '.join(selected_days)}…"):
                    gd, html, r = run_gd(p, tmp, sc, days_arg)
                if r.returncode == 0 and gd.exists():
                    esp_path = Path(str(gd).replace('.xlsx','_SOLO_ESPECIALES.xlsx'))
                    can_path = Path(str(gd).replace('.xlsx','_CANCELADAS.txt'))
                    _gd_bytes_f = gd.read_bytes()
                    _day_suffix = "_" + "+".join(b.split()[0] for b in selected_bloques)
                    _gd_name_f  = gd.stem + _day_suffix + ".xlsx"
                    _esp_name_f = esp_path.stem + _day_suffix + ".xlsx" if esp_path.exists() else None
                    st.session_state["r1_gd"]  = (_gd_name_f, _gd_bytes_f)
                    st.session_state["r1_esp"] = (_esp_name_f, esp_path.read_bytes()) if esp_path.exists() else None
                    _px_f, _sx_f = gd_to_dxc_csv(_gd_bytes_f)
                    st.session_state["r1_postex_csv"] = (gd.stem + _day_suffix + "_POSTEX.csv", _px_f)
                    st.session_state["r1_sorexp_csv"] = (gd.stem + _day_suffix + "_SOREXP.csv", _sx_f)
                    if esp_path.exists():
                        _epx_f, _esx_f = gd_to_dxc_csv(esp_path.read_bytes())
                        st.session_state["r1_esp_postex_csv"] = (esp_path.stem + _day_suffix + "_POSTEX.csv", _epx_f)
                        st.session_state["r1_esp_sorexp_csv"] = (esp_path.stem + _day_suffix + "_SOREXP.csv", _esx_f)
                    st.session_state["r1_can"]  = (can_path.name, can_path.read_text(encoding='utf-8')) if can_path.exists() else None
                    st.session_state["r1_html"] = (html.name,  html.read_bytes()) if html.exists() else None
                    st.session_state["r1_day_filter"] = selected_days
                    st.rerun()
                else:
                    st.error("Error en regeneración.")
                    show_log(r, expanded=True)

    if st.session_state["r1_day_filter"]:
        _labels = [f"Bloque {l}" for l in st.session_state['r1_day_filter']]
        st.info(f"Filtrado a: {', '.join(_labels)}")

    c1, c2 = st.columns(2)
    with c1:
        name, data = st.session_state["r1_gd"]
        st.download_button("⬇️ GD completo", data=data, file_name=name,
                           mime=XLSX_MIME, use_container_width=True)
        st.caption("GD completo — subir a DXC / MAR")
    with c2:
        if st.session_state["r1_esp"]:
            name, data = st.session_state["r1_esp"]
            st.download_button("⬇️ Solo especiales", data=data, file_name=name,
                               mime=XLSX_MIME, use_container_width=True)
            st.caption("Solo filas nuevas a añadir en DXC")

    st.markdown("**Formato CSV para importar en DXC:**")
    c5, c6, c7, c8 = st.columns(4)
    with c5:
        if st.session_state["r1_postex_csv"]:
            name, data = st.session_state["r1_postex_csv"]
            st.download_button("⬇️ POSTEX completo", data=data, file_name=name,
                               mime="text/csv", use_container_width=True)
    with c6:
        if st.session_state["r1_sorexp_csv"]:
            name, data = st.session_state["r1_sorexp_csv"]
            st.download_button("⬇️ SOREXP completo", data=data, file_name=name,
                               mime="text/csv", use_container_width=True)
    with c7:
        if st.session_state["r1_esp_postex_csv"]:
            name, data = st.session_state["r1_esp_postex_csv"]
            st.download_button("⬇️ POSTEX especiales", data=data, file_name=name,
                               mime="text/csv", use_container_width=True)
    with c8:
        if st.session_state["r1_esp_sorexp_csv"]:
            name, data = st.session_state["r1_esp_sorexp_csv"]
            st.download_button("⬇️ SOREXP especiales", data=data, file_name=name,
                               mime="text/csv", use_container_width=True)

    c3, c4 = st.columns(2)
    with c3:
        if st.session_state["r1_can"]:
            name, txt = st.session_state["r1_can"]
            st.download_button("⬇️ Canceladas.txt", data=txt, file_name=name,
                               mime="text/plain", use_container_width=True)
            with st.expander("Ver canceladas"):
                st.text(txt)
            st.caption("Salidas a eliminar del sorter")
    with c4:
        if st.session_state["r1_html"]:
            name, data = st.session_state["r1_html"]
            st.download_button("⬇️ Resumen HTML", data=data, file_name=name,
                               mime="text/html", use_container_width=True)
            st.caption("Informe con gráfico interactivo")

# ── Show results action 2 ─────────────────────────────────────────────────────
if st.session_state["r2_gantt"] is not None:
    name, data = st.session_state["r2_gantt"]
    st.success("✓ Gantt 1H generado")
    st.download_button("⬇️ Gantt 1H.xlsx", data=data, file_name=name,
                       mime=XLSX_MIME, use_container_width=True)
    st.caption("Hojas: LEYENDA · BLOQUES_DESTINOS · GANTT_VISUAL · GANTT_OPERATIVO")

# ── Show results action 3 ─────────────────────────────────────────────────────
if st.session_state["r3_map"] is not None:
    name, data = st.session_state["r3_map"]
    st.success("✓ Sorter Map generado")
    st.download_button("⬇️ Sorter Map.xlsx", data=data, file_name=name,
                       mime=XLSX_MIME, use_container_width=True)
    st.caption("Hojas: DOM · LUN · MAR · MIÉ · JUE · VIE · SÁB · LEYENDA")

st.divider()
st.caption("v0.10 · VDL B2B · Estrictamente confidencial")
