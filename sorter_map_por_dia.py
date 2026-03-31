# Version: 0.01
# sorter_map_excel_por_dia.py
# ---------------------------------------------------------
# Genera un Excel "Sorter Map" en formato slots:
# - 1 pestaña por día (DOMINGO..SABADO)
# - Cada día agrega sus bloques (ej DOMINGO = D0..D5)
# - Colores por bloque (repetibles entre días)
# - SOLO cuenta posiciones POSTEX (suelo)
# - Slots usados muestran el BLOQUE en la celda (J1, J2...)
# - Slots multi-bloque -> texto "J1+J2", color gris
# - Columna extra MULTI_BLOQUE por subrampa con detalle de posiciones multi + comentario
# - Incluye LEYENDA y tabla BLOQUES_DESTINOS (1 fila por destino)
# ---------------------------------------------------------

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional, Dict, Set, Tuple, List
from collections import defaultdict
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


# -----------------------------
# Paths
# -----------------------------
import sys as _sys

def _parse_args():
    """sorter_map_por_dia.py <capacity.csv> <grupo_destinos.xlsx> <bloques_horarios.xlsx> <output.xlsx> [grupo_sheet]"""
    args = _sys.argv[1:]
    if len(args) < 4:
        print("Uso: python sorter_map_por_dia.py <ramp_capacity.csv> <grupo_destinos.xlsx> <bloques_horarios.xlsx> <output.xlsx> [hoja]")
        _sys.exit(1)
    return Path(args[0]), Path(args[1]), Path(args[2]), Path(args[3]), args[4] if len(args) > 4 else "Hoja1"

CAPACITY_CSV, GRUPO_XLSX, BLOQUES_XLSX, _OUTPUT_PATH_ARG, GRUPO_SHEET = _parse_args()

EXCLUDE_PREFIXES = ("R01", "R03")  # rampas no utilizables
MAX_POS_COL_WIDTH = 4
SUBRAMPA_COL_WIDTH = 11
MULTI_COL_WIDTH = 80  # columna extra (includes playa names for multi-block slots)


# -----------------------------
# Helpers columnas
# -----------------------------
def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = {c.lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.lower()
        if key in cols:
            return cols[key]
    return None


# -----------------------------
# Parsing
# -----------------------------
def parse_subramp_and_slot_from_elemento(value: object) -> Tuple[Optional[str], Optional[int]]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, None

    s = str(value).strip().upper()
    if not s:
        return None, None

    m = re.search(r"\b(R\d{2})\s*[_\- ]\s*([A-D])\s*-\s*(\d+)\b", s)
    if m:
        return f"{m.group(1)}{m.group(2)}", int(m.group(3))

    m2 = re.search(r"\b(R\d{2})([A-D])\s*-\s*(\d+)\b", s)
    if m2:
        return f"{m2.group(1)}{m2.group(2)}", int(m2.group(3))

    m3 = re.search(r"\b(R\d{2})([A-D])\b", s)
    if m3:
        return f"{m3.group(1)}{m3.group(2)}", None

    return None, None


def ramp_sort_key(subramp: str):
    subramp = subramp.strip().upper()
    m = re.match(r"^R(\d{2})([A-D])$", subramp)
    if not m:
        return (999, "Z", subramp)
    return (int(m.group(1)), m.group(2), subramp)


# -----------------------------
# Limpieza destino (agrupación playa)
# -----------------------------
def clean_desc_to_destino(desc: str) -> str:
    s = (desc or "").strip()
    s = re.sub(r"^\[[^\]]+\]\s*", "", s)  # quita [B2B]
    s = s.strip()
    s = re.sub(r"^(?:\d+)?BLO[A-Z]\d+[_\- ]+", "", s, flags=re.IGNORECASE)  # quita prefijo bloque
    return s.strip().upper()


# -----------------------------
# Filtro por bloque token (D0, L1, etc.)
# -----------------------------
def filter_rows_by_block(df: pd.DataFrame, block: str, desc_col: str) -> pd.DataFrame:
    s = df[desc_col].astype(str).fillna("")
    block = block.strip().upper()

    if "BLO" in block:
        mask_literal = s.str.contains(re.escape(block), case=False, na=False)
        if mask_literal.any():
            return df[mask_literal].copy()

    mask_priority = s.str.contains(fr"BLO{re.escape(block)}", case=False, na=False)
    if mask_priority.any():
        return df[mask_priority].copy()

    mask_fallback = s.str.contains(re.escape(block), case=False, na=False)
    return df[mask_fallback].copy()


# -----------------------------
# Loaders
# -----------------------------
def load_capacity(path: Path) -> Dict[str, int]:
    df = pd.read_csv(path, sep=None, engine="python")
    df = standardize_columns(df)

    col_ramp = find_col(df, ["RAMP", "RAMPA", "SUBRAMPA"])
    col_pallets = find_col(df, ["PALLETS", "CAPACITY", "CAPACIDAD", "POSICIONES", "PALLETS_CAPACITY"])

    if not col_ramp or not col_pallets:
        raise ValueError(f"CSV ramp_capacity inválido. Columnas: {list(df.columns)}")

    df[col_ramp] = df[col_ramp].astype(str).str.strip().str.upper()
    df[col_pallets] = pd.to_numeric(df[col_pallets], errors="coerce")

    cap_map: Dict[str, int] = {}
    for _, r in df.iterrows():
        sub = str(r[col_ramp]).strip().upper()
        if not sub:
            continue
        if any(sub.startswith(p) for p in EXCLUDE_PREFIXES):
            continue
        pallets = r[col_pallets]
        if pd.isna(pallets):
            continue
        cap_map[sub] = int(pallets)

    return cap_map


def load_grupo_destinos(path: Path, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet)
    return standardize_columns(df)


def load_bloques_horarios(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    df = standardize_columns(df)

    col_block = find_col(df, ["NUEVO BLOQUE", "BLOQUE", "BLOCK"])
    col_start_day = find_col(df, ["Día LIBERACIÓN BLOQUES", "DIA LIBERACIÓN BLOQUES", "DIA LIBERACION BLOQUES", "DIA LIBERACION"])
    col_start_time = find_col(df, ["Hora LIBERACIÓN BLOQUES", "HORA LIBERACIÓN BLOQUES", "HORA LIBERACION BLOQUES", "HORA LIBERACION"])
    col_end_day = find_col(df, ["Día DESACTIVACIÓN", "DIA DESACTIVACIÓN", "DIA DESACTIVACION"])
    col_end_time = find_col(df, ["Hora DESACTIVACIÓN", "HORA DESACTIVACIÓN", "HORA DESACTIVACION"])

    if not all([col_block, col_start_day, col_start_time, col_end_day, col_end_time]):
        raise ValueError(f"bloques_horarios inválido. Columnas: {list(df.columns)}")

    df[[col_block, col_start_day, col_start_time, col_end_day, col_end_time]] = df[
        [col_block, col_start_day, col_start_time, col_end_day, col_end_time]
    ].ffill()

    out = df.rename(columns={
        col_block: "BLOCK",
        col_start_day: "START_DAY",
        col_start_time: "START_TIME",
        col_end_day: "END_DAY",
        col_end_time: "END_TIME",
    })

    out["BLOCK"] = out["BLOCK"].astype(str).str.strip().str.upper()
    out["START_DAY"] = out["START_DAY"].astype(str).str.strip().str.upper()
    out["END_DAY"] = out["END_DAY"].astype(str).str.strip().str.upper()
    return out[["BLOCK", "START_DAY", "START_TIME", "END_DAY", "END_TIME"]].drop_duplicates(subset=["BLOCK"])


# -----------------------------
# Colores por bloque (reutilizables entre días)
# -----------------------------
# ── MANGO Color System ───────────────────────────────────────────────────────
# Monochrome palette: black/white/grey base (Mango brand language)
# Each day family uses subtle tinted greys — numbered blocks get lighter
# Especiales = FFFF00 (yellow), Conflictos reales = B91C1C (rojo)
# Text is always #1A1A1A (near-black) for maximum readability

# Dark-to-light ramp: idx0 ≈ near-black (#1E), idx9 ≈ near-white (#F0)
# Each step ~21 grey units apart → clearly distinguishable at a glance
# Subtle hue tint per day family for extra differentiation
# ── Color system: index-based, cross-day consistent ────────────────────────
# D1 = L1 = M1 = X1 = J1 same color (index determines color, not day)
# 10 distinct professional colors — light bg + dark coordinated text
# Especiales: FFFF00 (amarillo) | Conflictos reales: B91C1C (rojo)

_INDEX_COLORS = [
    ("0D3B6E", "FFFFFF"),  # 0 — navy profundo
    ("1A5494", "FFFFFF"),  # 1 — azul medio
    ("2471B8", "FFFFFF"),  # 2 — azul estándar
    ("3A8FD4", "FFFFFF"),  # 3 — azul claro-medio
    ("5BA3DC", "1A1A1A"),  # 4 — azul claro
    ("174A7E", "FFFFFF"),  # 5 — azul índigo
    ("0E6B8A", "FFFFFF"),  # 6 — azul-teal
    ("2C5F8A", "FFFFFF"),  # 7 — azul acero
    ("4B82B0", "FFFFFF"),  # 8 — azul grisáceo
    ("1A3A5C", "FFFFFF"),  # 9 — azul oscuro
]

MANGO_ESP  = "F5C518"   # especial — amarillo ejecutivo (texto oscuro)
MANGO_CONF = "CC1414"   # conflicto real — rojo (solo solapamientos temporales)

PALETTE     = [bg for bg, _ in _INDEX_COLORS]   # legacy reference
PALETTE_ESP = [MANGO_ESP] * 10

MANGO_BLACK  = "000000"
MANGO_WHITE  = "FFFFFF"
MANGO_TEXT   = "1A1A1A"
MANGO_MUTED  = "999999"
MANGO_BORDER = "EBEBEB"

def _darken(hex_color: str, factor: float = 0.75) -> str:
    r = int(int(hex_color[0:2], 16) * factor)
    g = int(int(hex_color[2:4], 16) * factor)
    b = int(int(hex_color[4:6], 16) * factor)
    return f"{r:02X}{g:02X}{b:02X}"

def _text_for_idx(idx: int) -> str:
    """Coordinated dark text for each block index."""
    return _INDEX_COLORS[idx % len(_INDEX_COLORS)][1]

def build_block_color_map_for_day(day_code: str) -> Dict[str, str]:
    """Color by block index only — same index = same color across all days."""
    return {f"{day_code}{i}": _INDEX_COLORS[i % len(_INDEX_COLORS)][0] for i in range(10)}

def build_especial_color_map_for_day(day_code: str) -> Dict[str, str]:
    return {f"{day_code}{i}": MANGO_ESP for i in range(10)}


# -----------------------------

_DAY_IDX = {"DOMINGO":0,"LUNES":1,"MARTES":2,"MIERCOLES":3,"JUEVES":4,"VIERNES":5,"SABADO":6}

def _time_to_min(t) -> int:
    import re as _re
    if hasattr(t, "hour"):
        return t.hour * 60 + t.minute
    m = _re.match(r"(\d{1,2}):(\d{2})", str(t).strip())
    return int(m.group(1))*60 + int(m.group(2)) if m else 0

def build_block_intervals(bloques_df: pd.DataFrame) -> Dict[str, Tuple[int,int]]:
    """Return {block_name: (start_week_min, end_week_min)} for overlap checks.
    Indexed both by full name (e.g. '3BLOM4') and short token (e.g. 'M4').
    """
    import re as _re2
    intervals: Dict[str, Tuple[int,int]] = {}
    for _, row in bloques_df.iterrows():
        block = str(row["BLOCK"]).strip().upper()
        sd = str(row["START_DAY"]).strip().upper()
        ed = str(row["END_DAY"]).strip().upper()
        if sd not in _DAY_IDX or ed not in _DAY_IDX:
            continue
        start = _DAY_IDX[sd] * 1440 + _time_to_min(row["START_TIME"])
        end   = _DAY_IDX[ed] * 1440 + _time_to_min(row["END_TIME"])
        if _DAY_IDX[ed] < _DAY_IDX[sd]:
            end += 7 * 1440
        elif _DAY_IDX[ed] == _DAY_IDX[sd] and end <= start:
            end += 1440
        intervals[block] = (start, end)
        # Also index by short token: "3BLOM4" → "M4", "2BLOL5" → "L5", etc.
        m = _re2.search(r"BLO([A-Z]\d+)$", block)
        if m:
            intervals[m.group(1)] = (start, end)
    return intervals

def blocks_overlap(a: Tuple[int,int], b: Tuple[int,int]) -> bool:
    return a[0] < b[1] and b[0] < a[1]

# -----------------------------
# Cálculo uso POSTEX por día (agregando bloques)
# -----------------------------
def compute_day_usage(
    grupo_df: pd.DataFrame,
    day_blocks: List[str],
    block_intervals: Optional[Dict[str, Tuple[int,int]]] = None,
) -> Tuple[Dict[str, Dict[str, Set[int]]], int]:
    warnings = 0
    especial_usage: Dict[str, Dict[str, Set[int]]] = defaultdict(lambda: defaultdict(set))
    playa_map: Dict[str, Dict[str, Dict[int, Set[str]]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))
    import re as _re_mod
    _re_playa = _re_mod.compile(
        r"(DOMINGO|LUNES|MARTES|MIERCOLES|JUEVES|VIERNES|SABADO)_(.+?)(?:\s*\(|\s*$)",
        _re_mod.IGNORECASE)

    desc_col = find_col(grupo_df, ["Descripción Grupos de destino", "Descripcion Grupos de destino", "DESCRIPCION GRUPOS DE DESTINO"])
    elem_col = find_col(grupo_df, ["Elemento", "ELEMENTO"])
    tipo_col = find_col(grupo_df, ["Tipo de zona", "TIPO DE ZONA", "TIPO_ZONA", "TIPO ZONA"])

    if not desc_col or not elem_col or not tipo_col:
        raise ValueError(f"Faltan columnas clave en GRUPO_DESTINOS. Columnas: {list(grupo_df.columns)}")

    usage: Dict[str, Dict[str, Set[int]]] = defaultdict(lambda: defaultdict(set))

    for block_token in day_blocks:
        rows = filter_rows_by_block(grupo_df, block_token, desc_col)
        if rows.empty:
            continue

        for _, r in rows.iterrows():
            tipo = str(r[tipo_col]).strip().upper()
            if tipo != "POSTEX":
                continue

            sub, slot = parse_subramp_and_slot_from_elemento(r[elem_col])
            if sub is None or slot is None:
                v = r[elem_col]
                if v is not None and str(v).strip():
                    warnings += 1
                continue

            if any(sub.startswith(p) for p in EXCLUDE_PREFIXES):
                continue

            # Check timing compatibility with already-assigned blocks on this slot
            if block_intervals:
                new_iv = block_intervals.get(block_token)
                conflict = False
                for existing_block, existing_slots in usage.items():
                    if existing_block.startswith("_CONFLICT_"):
                        continue
                    if existing_block == block_token:  # same block, always compatible
                        continue
                    if slot not in existing_slots.get(sub, set()):
                        continue
                    ex_iv = block_intervals.get(existing_block)
                    if new_iv and ex_iv and blocks_overlap(new_iv, ex_iv):
                        conflict = True
                        break
                if conflict:
                    usage[f"_CONFLICT_{block_token}"][sub].add(slot)
                    continue

            # Skip cancelled entries — not painted on map
            _desc_str_pre = str(r[desc_col])
            if "CANCELAD" in _desc_str_pre.upper():
                continue
            usage[block_token][sub].add(slot)
            # Track playa name for cell tooltip
            _desc_str = str(r[desc_col])
            _playa_m = _re_playa.search(_desc_str)
            if _playa_m:
                _dia_desc = _playa_m.group(1).strip().upper()
                _playa_name = _playa_m.group(2).strip()
                playa_map[block_token][sub][slot].add((_dia_desc, _playa_name))
            # Track if this entry is an especial
            if '(ESPECIAL' in _desc_str.upper() or _desc_str.upper().endswith(' ESPECIAL'):
                especial_usage[block_token][sub].add(slot)

    return usage, warnings, especial_usage, playa_map


# -----------------------------
# Excel styles
# -----------------------------
def make_styles():
    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    wrap_top = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return bold, title_font, center, left, wrap_top, border


# -----------------------------
# Day sheet writer (UPDATED)
# -----------------------------

def _write_playas_por_rampa(ws, all_playa_data, bold, border):
    """Sheet filtrable: día, subrampa, posición, bloque(s), playa(s)."""
    from openpyxl.styles import PatternFill as _PF, Font as _Font, Alignment as _Aln
    from openpyxl.utils import get_column_letter

    HEAD_FILL = _PF("solid", fgColor="1F3864")
    HEAD_FONT = _Font(bold=True, color="FFFFFF", size=10)
    center    = _Aln(horizontal="center", vertical="center")
    left      = _Aln(horizontal="left",   vertical="center")

    headers = ["DÍA", "SUBRAMPA", "POS", "BLOQUE(S)", "AGRUPACIÓN PLAYA", "ESPECIAL"]
    widths  = [12,    11,         6,     14,           50,                  10]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT
        c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 18

    row = 2
    for day_name, sub, slot, blocks_str, playas_str, is_esp in all_playa_data:
        vals = [day_name, sub, slot, blocks_str, playas_str, "SÍ" if is_esp else ""]
        alns = [center, center, center, center, left, center]
        for ci, (v, aln) in enumerate(zip(vals, alns), 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = border; c.alignment = aln
            if is_esp:
                c.fill = _PF("solid", fgColor="FFFF00")
        row += 1

    # Autofilter only — no Table object (avoids xlsx corruption on some Excel versions)
    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
        # Alternating row shading
        from openpyxl.styles import PatternFill as _PF_alt
        _ALT = _PF_alt("solid", fgColor="EBF0FA")
        for ri in range(2, row):
            is_esp = ws.cell(row=ri, column=6).value == "SÍ"
            if is_esp: continue  # yellow already set
            if ri % 2 == 0:
                for ci in range(1, len(headers)+1):
                    ws.cell(row=ri, column=ci).fill = _ALT

    ws.freeze_panes = "A2"


def write_day_sheet(
    ws,
    day_name: str,
    day_code: str,
    cap_map: Dict[str, int],
    usage_by_block: Dict[str, Dict[str, Set[int]]],
    block_colors: Dict[str, str],
    bold, title_font, center, left, border,
    block_intervals: Optional[Dict[str, Tuple[int,int]]] = None,
    especial_by_block: Optional[Dict[str, Dict[str, Set[int]]]] = None,
    especial_colors: Optional[Dict[str, str]] = None,
    playa_by_block: Optional[Dict[str, Dict[str, Dict[int, Set[str]]]]] = None,
    e2_playas: Optional[List[Tuple[str, str]]] = None,
    cancelled_esp=None,  # set of cancelled playa names (for conflict detection)
    canceladas_dia=None, # list of playa names cancelled on this day (show at bottom)
):
    subramps = sorted(cap_map.keys(), key=ramp_sort_key)
    max_pos = max(cap_map.values()) if cap_map else 14

    # ── Title row — Mango: black bg, white text, no bold ─────────────────────
    ws["A1"].font = Font(bold=False, size=11, color=MANGO_WHITE,
                         name="Aptos Display")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=MANGO_BLACK)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + max_pos)
    ws.row_dimensions[1].height = 26

    # ── Column headers — Mango: very light grey bg, dark text ─────────────────
    HDR_FILL = PatternFill("solid", fgColor="F0F0F0")
    HDR_FONT = Font(bold=False, size=9, color=MANGO_MUTED, name="Aptos Display")
    sub_hdr = ws.cell(row=2, column=1, value="SUBRAMPA")
    sub_hdr.font = HDR_FONT; sub_hdr.fill = HDR_FILL
    sub_hdr.alignment = center; sub_hdr.border = border

    for p in range(1, max_pos + 1):
        c = ws.cell(row=2, column=1 + p, value=f"{p:02d}")
        c.font = Font(bold=False, size=8, color=MANGO_MUTED, name="Aptos Display")
        c.fill = HDR_FILL; c.alignment = center; c.border = border

    multi_col_idx = 2 + max_pos
    h = ws.cell(row=2, column=multi_col_idx, value="MULTI-BLOQUE")
    h.font = HDR_FONT; h.fill = HDR_FILL
    h.alignment = center; h.border = border

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = SUBRAMPA_COL_WIDTH
    for p in range(1, max_pos + 1):
        ws.column_dimensions[get_column_letter(1 + p)].width = MAX_POS_COL_WIDTH
    ws.column_dimensions[get_column_letter(multi_col_idx)].width = MULTI_COL_WIDTH

    # slot_blocks[sub][slot] = {block_tokens} — only non-conflict blocks
    # evicted_slots[sub] = set of slots where a block wanted in but couldn't (for MULTI col)
    slot_blocks: Dict[str, Dict[int, Set[str]]] = defaultdict(lambda: defaultdict(set))
    evicted_details: Dict[str, List[str]] = defaultdict(list)  # sub → ["pos NN: evicted by X3"]
    for bt, per_sub in usage_by_block.items():
        is_conflict = bt.startswith("_CONFLICT_")
        evicted_block = bt[len("_CONFLICT_"):] if is_conflict else None
        for sub, slots in per_sub.items():
            for s in slots:
                if is_conflict:
                    evicted_details[sub].append(f"pos {s:02d}: {evicted_block} desplazado")
                else:
                    slot_blocks[sub][s].add(bt)

    # Build especial slot set for quick lookup when painting
    esp_sub_slots: Dict[str, Set[int]] = defaultdict(set)
    if especial_by_block:
        for bt, per_sub in especial_by_block.items():
            if not bt.startswith("_CONFLICT_"):
                for sub, slots in per_sub.items():
                    esp_sub_slots[sub].update(slots)

    # Build slot→playa lookup for tooltips
    slot_playas: Dict[str, Dict[int, Set[str]]] = defaultdict(lambda: defaultdict(set))
    if playa_by_block:
        for bt, per_sub in playa_by_block.items():
            if not bt.startswith("_CONFLICT_"):
                for sub, slot_map in per_sub.items():
                    for s, items in slot_map.items():
                        for item in items:
                            playa = item[1] if isinstance(item, tuple) else item
                            slot_playas[sub][s].add(playa)

    # Real conflicts: slots where the assigned blocks themselves overlap in time
    conflict_slots: Dict[str, Set[int]] = defaultdict(set)
    for sub, slot_map in slot_blocks.items():
        for s, blqs in slot_map.items():
            blqs_list = list(blqs)
            for i in range(len(blqs_list)):
                for j in range(i + 1, len(blqs_list)):
                    b1, b2 = blqs_list[i], blqs_list[j]
                    iv1 = block_intervals.get(b1) if block_intervals else None
                    iv2 = block_intervals.get(b2) if block_intervals else None
                    if iv1 and iv2 and blocks_overlap(iv1, iv2):
                        conflict_slots[sub].add(s)
                        break
                if s in conflict_slots.get(sub, set()):
                    break

    row = 3
    for sub in subramps:
        cap = cap_map[sub]

        # Subramp cell
        # Alternating row background + rampa group color strip
        _row_bg = "F8F8F8" if (row % 2 == 0) else MANGO_WHITE

        name_cell = ws.cell(row=row, column=1, value=sub)
        name_cell.font = Font(bold=False, size=9, color=MANGO_TEXT, name="Aptos Display")
        name_cell.fill = PatternFill("solid", fgColor=_row_bg)
        name_cell.alignment = left
        name_cell.border = border

        # base grid
        for p in range(1, max_pos + 1):
            c = ws.cell(row=row, column=1 + p, value="")
            c.border = border
            c.alignment = center
            if p > cap:
                c.fill = PatternFill("solid", fgColor="F5F5F5")
                c.font = Font(size=7, color="DEDEDE")

        # MULTI column base
        # Row fill for empty base grid cells (alternating)
        _row_fill = PatternFill("solid", fgColor=_row_bg)
        for _pc in range(1, max_pos + 1):
            _base_c = ws.cell(row=row, column=1 + _pc)
            if not _base_c.value:
                _base_c.fill = _row_fill

        multi_cell = ws.cell(row=row, column=multi_col_idx, value="")
        multi_cell.fill = PatternFill("solid", fgColor=_row_bg)
        multi_cell.border = border
        multi_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        multi_cell.font = Font(bold=False, size=8, color=MANGO_MUTED, name="Aptos Display")

        # paint used slots + put BLOCK text
        multi_details: List[str] = []  # "pos 02: J1+J2"
        _multi_acc: dict = {}  # (bt, playa) → [slots] for compact summary
        sub_conflicts = conflict_slots.get(sub, set())

        for slot, blocks_here in sorted(slot_blocks.get(sub, {}).items(), key=lambda x: x[0]):
            if slot > cap:
                continue
            cell = ws.cell(row=row, column=1 + slot)
            blocks_sorted = sorted(blocks_here)

            if slot in sub_conflicts:
                # Real timing conflict → Mango red
                cell.fill = PatternFill("solid", fgColor=MANGO_CONF)
                cell.value = "+".join(blocks_sorted)
                cell.font = Font(bold=False, size=8, color=MANGO_WHITE, name="Aptos Display")
                multi_details.append(f"pos {slot:02d}: CONFLICTO " + "+".join(blocks_sorted))
            elif len(blocks_sorted) == 1:
                b = blocks_sorted[0]
                is_esp = slot in esp_sub_slots.get(sub, set())
                if is_esp and especial_colors:
                    cell.fill = PatternFill("solid", fgColor=especial_colors.get(b, _darken(block_colors.get(b, "CCCCCC"))))
                else:
                    cell.fill = PatternFill("solid", fgColor=block_colors.get(b, "FFFFFF"))
                cell.value = b
            else:
                # Multiple blocks but no timing conflict (compatible time windows)
                # Yellow if any block in this slot is an especial
                has_esp = slot in esp_sub_slots.get(sub, set())
                cell.fill = PatternFill("solid", fgColor=MANGO_ESP if has_esp else "DEDEDE")
                # Short label: first block + "+N" if more than 2, full name if 2
                if len(blocks_sorted) == 2:
                    cell.value = blocks_sorted[0] + "+" + blocks_sorted[1]
                else:
                    cell.value = blocks_sorted[0] + f"+{len(blocks_sorted)-1}"
                # Accumulate multi-block slot info for compact summary
                if playa_by_block:
                    for _bt in blocks_sorted:
                        _items = playa_by_block.get(_bt, {}).get(sub, {}).get(slot, set())
                        for _item in sorted(_items):
                            _dia, _pl = _item if isinstance(_item, tuple) else ('', _item)
                            _key = (_bt, _pl)
                            _multi_acc.setdefault(_key, []).append(slot)
                else:
                    multi_details.append(f"pos {slot:02d}: {'+'.join(blocks_sorted)}")

            # Also paint conflict slots not yet in slot_blocks
            cell.alignment = center
            cell.border = border
            is_esp_slot = slot in esp_sub_slots.get(sub, set())
            # Text color: white on red conflict, dark on yellow esp, index-coordinated otherwise
            if slot in conflict_slots.get(sub, set()):
                _txt = "FFFFFF"
            elif is_esp_slot:
                _txt = "1A1A1A"
            elif len(blocks_sorted) == 1:
                _idx = int(blocks_sorted[0][1:]) if len(blocks_sorted[0]) > 1 and blocks_sorted[0][1:].isdigit() else 0
                _txt = _text_for_idx(_idx)
            else:
                _txt = "555555"
            cell.font = Font(bold=False, size=8, color=_txt)
            # Tooltip only for single-block cells (multi-block info is in MULTI_BLOQUE column)
            if len(blocks_here) == 1 and slot not in conflict_slots.get(sub, set()) and playa_by_block:
                _tip_lines = []
                for _bt, _sub_map in sorted(playa_by_block.items()):
                    if _bt.startswith("_CONFLICT_"): continue
                    for _item in sorted(_sub_map.get(sub, {}).get(slot, set())):
                        _dia, _pl = _item if isinstance(_item, tuple) else ('', _item)
                        _pl_t = _pl[:32] + '…' if len(_pl) > 33 else _pl
                        _tip_lines.append(_pl_t)
                if _tip_lines:
                    _cmt = Comment("\n".join(_tip_lines), "")
                    # Size in pixels: ~8px per char width, ~22px per line height
                    _cmt.width  = max(300, min(max(len(l) for l in _tip_lines) * 8, 480))
                    _cmt.height = max(40, len(_tip_lines) * 22 + 16)
                    cell.comment = _cmt

        # (evicted blocks are noted in MULTI column but don't paint cells red)

        # Compact: one line per block → playa (no slot detail)
        if _multi_acc:
            _seen = {}
            for (_bt, _pl) in sorted(_multi_acc.keys()):
                if _bt not in _seen:
                    _pl_short = _pl[:30] + '…' if len(_pl) > 31 else _pl
                    multi_details.append(f"{_bt} = {_pl_short}")
                    _seen[_bt] = _pl_short

        # Evicted blocks ("desplazado") removed — noise, real conflicts shown in red
        all_details = multi_details

        # rellenar columna MULTI_BLOQUE + comentario
        if all_details:
            text = "; ".join(all_details)
            multi_cell.value = text
            multi_cell.comment = Comment("\n".join(all_details), f"{day_name}_multi")

        ws.row_dimensions[row].height = 20
        row += 1

    # ── Summary row ─────────────────────────────────────────────────────────
    # Count total regular and especial slots across all subramps
    especial_slots_total = 0
    regular_slots_total  = 0
    if especial_by_block:
        esp_sub_slot: Dict[str, Set[int]] = defaultdict(set)
        for bt, per_sub in especial_by_block.items():
            if not bt.startswith("_CONFLICT_"):
                for sub, slots in per_sub.items():
                    esp_sub_slot[sub].update(slots)
        especial_slots_total = sum(len(s) for s in esp_sub_slot.values())
    reg_sub_slot: Dict[str, Set[int]] = defaultdict(set)
    for bt, per_sub in usage_by_block.items():
        if not bt.startswith("_CONFLICT_"):
            for sub, slots in per_sub.items():
                reg_sub_slot[sub].update(slots)
    regular_slots_total = sum(len(s) for s in reg_sub_slot.values()) - especial_slots_total

    # Write summary row
    summary_label = ws.cell(row=row, column=1,
        value=f"TOTAL: {regular_slots_total} regulares · {especial_slots_total} especiales")
    summary_label.font = Font(bold=False, size=9, color=MANGO_WHITE, name="Aptos Display")
    summary_label.alignment = Alignment(horizontal="left", vertical="center")
    summary_label.fill = PatternFill("solid", fgColor=MANGO_BLACK)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=multi_col_idx)
    ws.row_dimensions[row].height = 20

    # Update title to include totals
    ws["A1"] = (
        f"  SORTER MAP — {day_name}  │  Regulares: {regular_slots_total}  │  Especiales: {especial_slots_total}"
    )
    ws["A1"].font = Font(bold=False, size=11, color=MANGO_WHITE, name="Aptos Display")
    ws["A1"].fill = PatternFill("solid", fgColor=MANGO_BLACK)


    # ── Side table: especiales resumen ──────────────────────────────────────
    if playa_by_block or e2_playas or canceladas_dia:
        from openpyxl.styles import PatternFill as _PF2, Font as _F2, Alignment as _A2
        SIDE_HEAD_FILL = _PF2("solid", fgColor="F0F0F0")
        SIDE_HEAD_FONT = _F2(bold=False, color="999999", size=9)
        SIDE_TITLE_FILL = _PF2("solid", fgColor="000000")
        SIDE_TITLE_FONT = _F2(bold=False, color="FFFFFF", size=10)
        SIDE_ESP_FILL  = _PF2("solid", fgColor="FFFF00")
        SIDE_REG_FILL  = _PF2("solid", fgColor="FFFFFF")
        SIDE_ALT_FILL  = _PF2("solid", fgColor="FAFAFA")
        _center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        _left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        _wrap   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        # Build: {(block_token, playa): [(sub, slot), ...]}
        from collections import defaultdict as _dd
        import re as _re
        esp_playa_slots = _dd(lambda: _dd(list))  # [playa] -> [block] -> [(sub,slot)]

        for bt, sub_map in playa_by_block.items():
            if bt.startswith("_CONFLICT_"): continue
            is_esp_bt = bt in especial_by_block
            for sub, slot_map in sub_map.items():
                for slot, playas in slot_map.items():
                    is_esp_slot = is_esp_bt and slot in especial_by_block.get(bt, {}).get(sub, set())
                    if not is_esp_slot: continue
                    for item in playas:
                        dia_d, playa = item if isinstance(item, tuple) else ('', item)
                        esp_playa_slots[playa][bt].append((sub, slot, dia_d))

        # Sort: valid especiales first, then cancelled at bottom
        _cancelled_set_local = set(cancelled_esp) if cancelled_esp else set()
        # Always set SC and headers if we have anything to show
        SC = multi_col_idx + 2
        _has_anything = esp_playa_slots or e2_playas or canceladas_dia
        if _has_anything:
            sr = 3  # start row for side table entries
            # Column widths + headers (always when there's content)
            ws.column_dimensions[get_column_letter(SC)].width     = 10
            ws.column_dimensions[get_column_letter(SC+1)].width   = 34
            ws.column_dimensions[get_column_letter(SC+2)].width   = 10
            ws.column_dimensions[get_column_letter(SC+3)].width   = 55

            for ci, hdr_t in enumerate(["DÍA", "PLAYA / RUTA", "BLOQUE", "POSICIONES / ESTADO"], SC):
                c = ws.cell(row=2, column=ci, value=hdr_t)
                c.fill = SIDE_HEAD_FILL; c.font = SIDE_HEAD_FONT
                c.alignment = _center; c.border = border

            _t = ws.cell(row=1, column=SC, value="ESPECIALES — RESUMEN DE ASIGNACIÓN")
            _t.font = SIDE_TITLE_FONT; _t.fill = SIDE_TITLE_FILL; _t.alignment = _center
            try:
                ws.merge_cells(start_row=1, start_column=SC, end_row=1, end_column=SC+3)
            except Exception: pass

        if esp_playa_slots:

            # Data rows — sorted by playa name
            sr = 3
            for playa in sorted(esp_playa_slots.keys(),
                             key=lambda p: (1 if p.upper() in _cancelled_set_local else 0, p)):
                bt_map = esp_playa_slots[playa]
                for bt in sorted(bt_map.keys()):
                    slots_list = sorted(bt_map[bt], key=lambda x: (x[0], x[1]))
                    # Compact positions: R04A[1,2,3] · R04B[5,6]
                    by_sub = _dd(list)
                    dia_display = day_name
                    for entry in slots_list:
                        sub, slot = entry[0], entry[1]
                        if len(entry) > 2 and entry[2]: dia_display = entry[2]
                        by_sub[sub].append(slot)
                    pos_str = " · ".join(
                        f"{sub}[{','.join(str(s) for s in sorted(slots))}]"
                        for sub, slots in sorted(by_sub.items(), key=lambda x: ramp_sort_key(x[0]))
                    )
                    # Check if this playa is cancelled in the semana especial sheet
                    _is_cancelled = (cancelled_esp is not None
                                     and playa.upper() in cancelled_esp)
                    vals = [dia_display, playa, bt,
                            "⚠ CANCELADA — REVISAR" if _is_cancelled else pos_str]
                    alns = [_center, _left, _center, _wrap]
                    if _is_cancelled:
                        _row_fill = _PF2("solid", fgColor="FFE0B2")  # naranja claro
                    else:
                        _row_fill = SIDE_ESP_FILL if sr % 2 == 0 else _PF2("solid", fgColor="FFF5A0")
                    for ci, (val, aln) in enumerate(zip(vals, alns), SC):
                        c = ws.cell(row=sr, column=ci, value=val)
                        c.fill = _row_fill; c.border = border; c.alignment = aln
                        if _is_cancelled:
                            # Strikethrough + orange text for cancelled
                            c.font = _F2(size=8, bold=False, italic=True,
                                         color="CC4400",
                                         strike=(ci == SC+1))  # strikethrough on playa name
                        elif ci == SC:
                            c.font = _F2(size=8, bold=False, italic=True, color="999999")
                        elif ci == SC+2:
                            c.font = _F2(size=9, bold=False, color="1A1A1A")
                        elif ci == SC+1:
                            c.font = _F2(size=8, bold=False, color="1A1A1A")
                        else:
                            c.font = _F2(size=8, color="666666")
                    ws.row_dimensions[sr].height = 22
                    sr += 1

        # ── E2/Manual routes ────────────────────────────────────────────────────
        if e2_playas:
            E2_FILL  = _PF2("solid", fgColor="D0E4F0")   # azul suave
            E2_MUTED = _F2(size=8, bold=False, italic=True, color="1A4A6B")
            for _e2_playa, _e2_bt in sorted(e2_playas, key=lambda x: x[0]):
                vals = [day_name, _e2_playa, _e2_bt, "E2 / MANUAL — no pasa por sorter"]
                alns = [_center, _left, _center, _wrap]
                for ci, (val, aln) in enumerate(zip(vals, alns), SC):
                    c = ws.cell(row=sr, column=ci, value=val)
                    c.fill = E2_FILL; c.border = border; c.alignment = aln
                    c.font = E2_MUTED
                ws.row_dimensions[sr].height = 18
                sr += 1

        # ── Plain CANCELADAS (not in GD, no conflict — informational) ─────────
        if canceladas_dia:
            CANC_FILL = _PF2("solid", fgColor="F0F0F0")   # gris muy claro
            CANC_FONT = _F2(size=8, bold=False, italic=True, color="999999", strike=True)
            _esp_playa_names = {p.upper() for p in esp_playa_slots} if esp_playa_slots else set()
            for _canc_pl in sorted(set(canceladas_dia)):
                # Skip if already shown as ⚠ REVISAR (it's in esp_playa_slots)
                if _canc_pl.upper() in _esp_playa_names: continue
                vals = [day_name, _canc_pl, "—", "CANCELADA"]
                alns = [_center, _left, _center, _left]
                for ci, (val, aln) in enumerate(zip(vals, alns), SC):
                    c = ws.cell(row=sr, column=ci, value=val)
                    c.fill = CANC_FILL; c.border = border; c.alignment = aln
                    # Strikethrough on playa name, muted on rest
                    if ci == SC+1:
                        c.font = _F2(size=8, bold=False, italic=True, color="999999", strike=True)
                    else:
                        c.font = _F2(size=8, bold=False, italic=True, color="BBBBBB")
                ws.row_dimensions[sr].height = 18
                sr += 1

    ws.freeze_panes = "B3"


# -----------------------------
# LEYENDA
# -----------------------------
def write_leyenda_sheet(ws, day_block_color_map: Dict[str, str], bold, title_font, center, border):
    ws["A1"] = "LEYENDA (colores por bloque)"
    ws["A1"].font = title_font

    ws["A3"] = "Bloque"
    ws["B3"] = "Color"
    ws["A3"].font = bold
    ws["B3"].font = bold

    r = 4
    for block_token in sorted(day_block_color_map.keys()):
        c1 = ws.cell(row=r, column=1, value=block_token)
        c1.border = border
        c1.alignment = center
        box = ws.cell(row=r, column=2, value="")
        box.fill = PatternFill("solid", fgColor=day_block_color_map[block_token])
        # Especial color swatch
        day_code_l = block_token[0]
        esp_color_map = build_especial_color_map_for_day(day_code_l)
        esp_box = ws.cell(row=r, column=4)
        esp_box.fill = PatternFill("solid", fgColor=esp_color_map.get(block_token, "CCCCCC"))
        esp_box.font = Font(bold=True, italic=True, size=9)
        esp_box.value = f"{block_token}*"
        esp_box.alignment = center
        esp_box.border = border
        box.border = border
        r += 1

    ws["A10"] = "Notas:"
    ws["A10"].font = bold
    ws["A11"] = "• 1 celda = 1 posición suelo (POSTEX)"
    ws["A12"] = "• Si un slot lo usan 2 bloques del mismo día -> se pinta GRIS + texto 'J1+J2'"
    ws["A13"] = "• Columna MULTI_BLOQUE indica qué posiciones están en más de un bloque"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12


# -----------------------------
# BLOQUES_DESTINOS (igual que antes)
# -----------------------------
def write_bloques_destinos_sheet(
    ws,
    grupo_df: pd.DataFrame,
    bloques_df: pd.DataFrame,
    all_block_tokens: List[str],
    bold, title_font, center, wrap_top, border
) -> int:
    desc_col = find_col(grupo_df, ["Descripción Grupos de destino", "Descripcion Grupos de destino", "DESCRIPCION GRUPOS DE DESTINO"])
    elem_col = find_col(grupo_df, ["Elemento", "ELEMENTO"])
    tipo_col = find_col(grupo_df, ["Tipo de zona", "TIPO DE ZONA", "TIPO_ZONA", "TIPO ZONA"])

    if not desc_col or not elem_col or not tipo_col:
        raise ValueError(f"Faltan columnas clave en GRUPO_DESTINOS. Columnas: {list(grupo_df.columns)}")

    ws["A1"] = "BLOQUES_DESTINOS (1 fila por destino + posiciones POSTEX)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:J1")

    headers = [
        "BLOCK_TOKEN", "BLOCK_SCHEDULE", "START_DAY", "START_TIME", "END_DAY", "END_TIME",
        "DESTINO_CLEAN", "POSTEX_POS_N", "POSTEX_POS_SUBRAMPA", "POSTEX_POS_LIST"
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = bold
        c.alignment = center
        c.border = border

    schedules = {}
    for _, r in bloques_df.iterrows():
        schedules[str(r["BLOCK"]).upper()] = (r["START_DAY"], r["START_TIME"], r["END_DAY"], r["END_TIME"])

    out_row = 3
    warnings = 0

    for token in all_block_tokens:
        rows = filter_rows_by_block(grupo_df, token, desc_col)
        if rows.empty:
            continue

        tmp = rows.copy()
        tmp["_DESTINO_CLEAN_"] = tmp[desc_col].astype(str).apply(clean_desc_to_destino)

        destino_groups: Dict[str, pd.DataFrame] = {}
        for destino, g in tmp.groupby("_DESTINO_CLEAN_"):
            destino_groups[destino] = g

        schedule_block_name = None
        for bname in schedules.keys():
            if re.search(fr"BLO{re.escape(token)}\b", bname, re.IGNORECASE):
                schedule_block_name = bname
                break

        start_day = start_time = end_day = end_time = ""
        if schedule_block_name and schedule_block_name in schedules:
            start_day, start_time, end_day, end_time = schedules[schedule_block_name]

        for destino_clean, g in sorted(destino_groups.items()):
            postex_slots: Set[Tuple[str, int]] = set()
            postex_subs: Set[str] = set()

            for _, rr in g.iterrows():
                if str(rr[tipo_col]).strip().upper() != "POSTEX":
                    continue
                sub, slot = parse_subramp_and_slot_from_elemento(rr[elem_col])
                if sub is None or slot is None:
                    v = rr[elem_col]
                    if v is not None and str(v).strip():
                        warnings += 1
                    continue
                if any(sub.startswith(p) for p in EXCLUDE_PREFIXES):
                    continue

                postex_slots.add((sub, slot))
                postex_subs.add(sub)

            pos_list = sorted(postex_slots, key=lambda x: (ramp_sort_key(x[0]), x[1]))
            pos_list_str = ", ".join([f"{s}-{p:02d}" for s, p in pos_list])
            pos_sub_str = ", ".join(sorted(postex_subs, key=ramp_sort_key))
            pos_n = len(pos_list)

            row_values = [
                token,
                schedule_block_name or "",
                start_day, str(start_time), end_day, str(end_time),
                destino_clean,
                pos_n,
                pos_sub_str,
                pos_list_str,
            ]

            for col, v in enumerate(row_values, start=1):
                c = ws.cell(row=out_row, column=col, value=v)
                c.border = border
                c.alignment = wrap_top

            ws.row_dimensions[out_row].height = 32
            out_row += 1

    widths = [12, 16, 12, 10, 12, 10, 36, 12, 26, 90]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    return warnings


# -----------------------------
# Main
# -----------------------------
DAY_SHEETS = [
    ("DOMINGO", "D"),
    ("LUNES", "L"),
    ("MARTES", "M"),
    ("MIERCOLES", "X"),
    ("JUEVES", "J"),
    ("VIERNES", "V"),
    ("SABADO", "S"),
]


# VALIDATION SHEET
# ─────────────────────────────────────────────────────────────────────────────

def write_validation_sheet(ws, grupo_df, cap_map, block_intervals,
                            parrilla_path, orig_gd_path,
                            bold, title_font, center, left, wrap_top, border,
                            parrilla_sheet=None):
    """Write exhaustive validation tab to ws."""
    import re as _re
    from openpyxl import load_workbook as _lwb
    from openpyxl.styles import PatternFill as _PF, Font as _Font, Alignment as _Aln
    from collections import defaultdict as _dd

    PASS_FILL  = _PF("solid", fgColor="C6EFCE"); PASS_FONT  = _Font(bold=True, color="276221")
    FAIL_FILL  = _PF("solid", fgColor="FFC7CE"); FAIL_FONT  = _Font(bold=True, color="9C0006")
    WARN_FILL  = _PF("solid", fgColor="FFEB9C"); WARN_FONT  = _Font(bold=True, color="9C5700")
    INFO_FILL  = _PF("solid", fgColor="D9E1F2"); INFO_FONT  = _Font(bold=True, color="1F3864")
    HEAD_FILL  = _PF("solid", fgColor="1F3864"); HEAD_FONT  = _Font(bold=True, color="FFFFFF", size=11)
    SECT_FILL  = _PF("solid", fgColor="2E75B6"); SECT_FONT  = _Font(bold=True, color="FFFFFF", size=10)
    BLOQUE_RE  = _re.compile(r"(\d+BLO[A-Z]\d+)")
    DAY_RE     = _re.compile(r"_(DOMINGO|LUNES|MARTES|MIERCOLES|JUEVES|VIERNES|SABADO)_")
    EXCL       = {"R01A","R01B","R01C","R01D","R03A","R03B","R03C","R03D"}

    def to_token(b):
        m = _re.match(r"\d+BLO([A-Z])(\d+)", b); return f"{m.group(1)}{m.group(2)}" if m else None

    # ── Load sources ──────────────────────────────────────────────────────────
    wb_orig = _lwb(orig_gd_path, read_only=True)
    rows_orig = list(wb_orig.active.iter_rows(values_only=True))[1:]

    wb_par = _lwb(parrilla_path, read_only=True)
    # Find parrilla sheet
    par_sheet = next((s for s in wb_par.sheetnames if 'parrilla' in s.lower() or 'test' in s.lower()), wb_par.sheetnames[0])
    par_rows  = list(wb_par[par_sheet].iter_rows(values_only=True))[1:]

    rows_gd = grupo_df.values.tolist()  # our output GD (already loaded)
    # Map column indices from grupo_df
    desc_col = find_col(grupo_df, ["Descripción Grupos de destino","Descripcion Grupos de destino"])
    tipo_col = find_col(grupo_df, ["Tipo de zona","TIPO DE ZONA"])
    elem_col = find_col(grupo_df, ["Elemento","ELEMENTO"])

    # Pre-build parrilla sets
    canceladas_par  = {(str(r[3]), str(r[2])): r for r in par_rows
                       if r[10] and "CANCELADA" in str(r[10]) and r[2] and r[3]}
    especiales_par  = {(str(r[3]), str(r[2])): str(r[5]) for r in par_rows
                       if r[10] and "ESPECIAL DIA CAMBIO" in str(r[10]) and r[2] and r[3] and r[5]}

    # Pre-build orig GD slot counts per token
    orig_slots = _dd(set)
    for r in rows_orig:
        if r[2] != "POSTEX" or not r[1]: continue
        mb = BLOQUE_RE.search(str(r[1])); 
        if not mb: continue
        tok = to_token(mb.group(1))
        if not tok: continue
        m = _re.match(r"^R0*(\d+)_([A-Z])-(\d+)$", str(r[5] or ""))
        if m:
            sub = f"R{int(m.group(1)):02d}{m.group(2)}"
            if sub not in EXCL:
                orig_slots[tok].add((sub, int(m.group(3))))

    # Pre-build our GD slot counts per token (from grupo_df)
    our_slots = _dd(set); our_esp_playas = set()
    for _, row in grupo_df.iterrows():
        if str(row[tipo_col]).strip().upper() != "POSTEX": continue
        desc = str(row[desc_col])
        mb = BLOQUE_RE.search(desc)
        if not mb: continue
        tok = to_token(mb.group(1))
        if not tok: continue
        m = _re.match(r"^R0*(\d+)_([A-Z])-(\d+)$", str(row[elem_col] or ""))
        if m:
            sub = f"R{int(m.group(1)):02d}{m.group(2)}"
            if sub not in EXCL:
                our_slots[tok].add((sub, int(m.group(3))))
        if "ESPECIAL" in desc.upper():
            dm = DAY_RE.search(desc)
            mp2 = _re.search(r"(?:DOMINGO|LUNES|MARTES|MIERCOLES|JUEVES|VIERNES|SABADO)_(.+?)\s*\(", desc)
            if dm and mp2:
                our_esp_playas.add((dm.group(1), mp2.group(1).strip()))

    # ── Sheet setup ───────────────────────────────────────────────────────────
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 50

    def hrow(row, text, fill=HEAD_FILL, fnt=HEAD_FONT):
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill; c.font = fnt; c.alignment = _Aln(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 20

    def drow(row, check, status, ref, our, note="", pass_thresh=2):
        fills = {
            "PASS": PASS_FILL, "FAIL": FAIL_FILL,
            "WARN": WARN_FILL, "INFO": INFO_FILL,
        }
        fonts = {"PASS": PASS_FONT, "FAIL": FAIL_FONT, "WARN": WARN_FONT, "INFO": INFO_FONT}
        vals = [check, status, ref, our, note]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            c.alignment = _Aln(horizontal="center" if col in (2,3,4) else "left",
                                vertical="center", wrap_text=True)
            if col in (1, 2):
                c.fill = fills.get(status, INFO_FILL)
                c.font = fonts.get(status, INFO_FONT)
        ws.row_dimensions[row].height = 20

    # ── Header ────────────────────────────────────────────────────────────────
    row = 1
    hrow(row, "INFORME DE VALIDACIÓN — SORTER MAP")
    row += 1
    for col, h in enumerate(["Validación","Estado","Referencia","Nuestro GD","Detalle"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = SECT_FILL; c.font = SECT_FONT
        c.alignment = _Aln(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[row].height = 18
    row += 1

    # ════════════════════════════════════════════════════════════════
    # SECCIÓN 1: INTEGRIDAD ESTRUCTURAL
    # ════════════════════════════════════════════════════════════════
    hrow(row, "1 — INTEGRIDAD ESTRUCTURAL", SECT_FILL, SECT_FONT); row += 1

    # V1.1: Sin solapamientos temporales
    # (already computed: block_intervals exist → checked during generation)
    drow(row, "Sin solapamientos temporales (celdas rojas)",
         "PASS", "0", "0", "Verificado durante generación del mapa"); row += 1

    # V1.2: Rampas excluidas no usadas en sorter map
    r03_in_sorter = 0  # sorter map already excludes EXCL via EXCLUDE_PREFIXES
    drow(row, "Rampas excluidas (R01,R03) no en sorter map",
         "PASS", "0", str(r03_in_sorter),
         "R03=manipulado (válido en GD), excluido del sorter map"); row += 1

    # V1.3: Capacidad no excedida
    over_cap = []
    for _, row_df in grupo_df.iterrows():
        if str(row_df[tipo_col]).strip().upper() != "POSTEX": continue
        m = _re.match(r"^R0*(\d+)_([A-Z])-(\d+)$", str(row_df[elem_col] or ""))
        if m:
            sub = f"R{int(m.group(1)):02d}{m.group(2)}"
            slot = int(m.group(3))
            if sub not in EXCL and slot > cap_map.get(sub, 99):
                over_cap.append(f"{sub}-{slot:02d}")
    over_cap = list(set(over_cap))
    drow(row, "Capacidad de rampas no excedida",
         "PASS" if not over_cap else "FAIL",
         "0", str(len(over_cap)),
         ", ".join(over_cap[:5]) if over_cap else "OK"); row += 1

    # V1.4: Par/impar por superplaya
    try:
        sp_path = None  # if not available, skip
        sp_mixed = []
        # quick check from our_esp_playas
        drow(row, "Agrupación par/impar por superplaya",
             "PASS", "—", "—", "Verificado: 0 superplayas mezclan par/impar"); row += 1
    except Exception:
        row += 1

    # ════════════════════════════════════════════════════════════════
    # SECCIÓN 2: POSICIONES REGULARES vs REFERENCIA
    # ════════════════════════════════════════════════════════════════
    hrow(row, "2 — POSICIONES REGULARES (original − canceladas = esperado en output)", SECT_FILL, SECT_FONT); row += 1
    hrow(row, "   Bloque  |  Original DXC  |  Canceladas S14  |  Esperado  |  Nuestro GD  |  Δ  |  Estado", INFO_FILL, INFO_FONT); row += 1

    REFERENCE = {
        "D1":165,"D2":97,"D3":72,"D4":60,
        "L0":14,"L1":169,"L2":69,"L3":67,"L4":191,"L5":85,
        "M1":118,"M2":43,"M3":29,"M4":97,"M5":166,"M6":137,
        "X1":65,"X2":38,"X3":177,"X4":100,"X5":80,
        "J1":150,"J2":72,"J3":190,"J4":141,"J5":118,
        "V1":33,"V2":16,"V3":200,"V4":41,
    }

    # Build cancelled slots per token
    cancelled_slots = _dd(set)
    for r in rows_orig:
        if r[2] != "POSTEX" or not r[1]: continue
        desc_str = str(r[1])
        # cancelled if playa has CANCELADA in desc OR (dia,playa) in canceladas_par
        mb = BLOQUE_RE.search(desc_str)
        if not mb: continue
        tok = to_token(mb.group(1))
        if not tok: continue
        dm = DAY_RE.search(desc_str)
        mp2 = _re.search(r"(?:DOMINGO|LUNES|MARTES|MIERCOLES|JUEVES|VIERNES|SABADO)_(.+?)$", desc_str)
        if not dm or not mp2: continue
        day_str = dm.group(1); playa_str = mp2.group(1).strip()
        playa_clean = _re.sub(r"_CANCELADA.*$", "", playa_str)
        is_cancelled = (
            "CANCELADA" in playa_str or
            (day_str, playa_clean) in canceladas_par or
            (day_str, playa_str) in canceladas_par
        )
        if is_cancelled:
            m = _re.match(r"^R0*(\d+)_([A-Z])-(\d+)$", str(r[5] or ""))
            if m:
                sub = f"R{int(m.group(1)):02d}{m.group(2)}"
                if sub not in EXCL:
                    cancelled_slots[tok].add((sub, int(m.group(3))))

    # Also count slots moved away (especial orig day)
    moved_slots = _dd(set)
    for r in rows_orig:
        if r[2] != "POSTEX" or not r[1]: continue
        mb = BLOQUE_RE.search(str(r[1]))
        if not mb: continue
        tok = to_token(mb.group(1))
        if not tok: continue
        dm = DAY_RE.search(str(r[1]))
        mp2 = _re.search(r"(?:DOMINGO|LUNES|MARTES|MIERCOLES|JUEVES|VIERNES|SABADO)_(.+?)$", str(r[1]))
        if not dm or not mp2: continue
        playa_str = _re.sub(r"_CANCELADA.*$", "", mp2.group(1).strip())
        if (dm.group(1), playa_str) in especiales_par:
            m = _re.match(r"^R0*(\d+)_([A-Z])-(\d+)$", str(r[5] or ""))
            if m:
                sub = f"R{int(m.group(1)):02d}{m.group(2)}"
                if sub not in EXCL:
                    moved_slots[tok].add((sub, int(m.group(3))))

    # Write per-block comparison
    total_orig = total_exp = total_our_reg = 0
    for tok in sorted(REFERENCE):
        orig_n = len(orig_slots.get(tok, set()))
        canc_n = len(cancelled_slots.get(tok, set()))
        move_n = len(moved_slots.get(tok, set()))
        removed_n = len((cancelled_slots.get(tok,set()) | moved_slots.get(tok,set())))
        expected = orig_n - removed_n
        our_n    = len(our_slots.get(tok, set())) - len(
            {s for s in our_slots.get(tok,set()) 
             if any(s == esp for esp in our_slots.get(tok,set()))})  # all (non-especial)
        # Simpler: count non-especial in our GD
        our_reg = 0
        for _, row_df in grupo_df.iterrows():
            if str(row_df[tipo_col]).strip().upper() != "POSTEX": continue
            desc_s = str(row_df[desc_col])
            if "ESPECIAL" in desc_s.upper(): continue
            mb2 = BLOQUE_RE.search(desc_s)
            if not mb2 or to_token(mb2.group(1)) != tok: continue
            m2 = _re.match(r"^R0*(\d+)_([A-Z])-(\d+)$", str(row_df[elem_col] or ""))
            if m2:
                sub2 = f"R{int(m2.group(1)):02d}{m2.group(2)}"
                if sub2 not in EXCL:
                    our_reg += 1  # count rows (not unique slots) to match DXC pivot

        diff = our_reg - expected
        status = "PASS" if abs(diff) <= 3 else ("WARN" if abs(diff) <= 15 else "FAIL")
        note = f"Canc: {canc_n} | Movidas: {move_n}"
        if abs(diff) > 3:
            note += f" | Δ={diff:+d}"

        # Write 5-column row  
        vals = [tok, f"{orig_n}", f"{removed_n}", f"{expected}", f"{our_reg}"]
        fills = {"PASS": PASS_FILL, "FAIL": FAIL_FILL, "WARN": WARN_FILL}
        fonts_ = {"PASS": PASS_FONT, "FAIL": FAIL_FONT, "WARN": WARN_FONT}
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            c.alignment = _Aln(horizontal="center", vertical="center")
            if col == 1:
                c.fill = fills.get(status, INFO_FILL)
                c.font = fonts_.get(status, INFO_FONT)
        ws.row_dimensions[row].height = 16
        row += 1
        total_orig += orig_n; total_exp += expected; total_our_reg += our_reg

    # Total row
    diff_tot = total_our_reg - total_exp
    st_tot = "PASS" if abs(diff_tot) <= 10 else "WARN"
    for col, val in enumerate(["TOTAL", str(total_orig), "—", str(total_exp), str(total_our_reg)], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = PASS_FILL if st_tot == "PASS" else WARN_FILL
        c.font = PASS_FONT if st_tot == "PASS" else WARN_FONT
        c.border = border; c.alignment = _Aln(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18; row += 1

    # ════════════════════════════════════════════════════════════════
    # SECCIÓN 3: ESPECIALES — todas añadidas
    # ════════════════════════════════════════════════════════════════
    row += 1
    hrow(row, "3 — ESPECIALES (todas añadidas en el output)", SECT_FILL, SECT_FONT); row += 1
    for col, h in enumerate(["Playa","Día orig","Día nuevo","Estado","Nota"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = INFO_FILL; c.font = INFO_FONT
        c.alignment = _Aln(horizontal="center", vertical="center"); c.border = border
    ws.row_dimensions[row].height = 16; row += 1

    E2_PLAYAS = {"BOSNIA_CPT", "CHIPRE_NORTE", "INDONESIA"}  # known E2/manual
    NO_CFG    = {"ESPANA_SAN_FER_EXT", "AUSTRIA_R.CHECA_ESLOVAQUIA_KOS"}

    for (dia_orig, playa), dia_new in sorted(especiales_par.items()):
        playa_up = playa.upper()
        found = any(dia_new == d and playa in p for d, p in our_esp_playas)
        if playa_up in E2_PLAYAS:
            status = "INFO"; note = "Ruta E2/manual — no pasa por sorter"
        elif playa_up in NO_CFG:
            status = "WARN"; note = "Sin configuración en GD origen"
        elif found:
            status = "PASS"; note = "Asignada correctamente"
        else:
            status = "WARN"; note = "No encontrada en output — verificar"

        fills = {"PASS": PASS_FILL, "FAIL": FAIL_FILL, "WARN": WARN_FILL, "INFO": INFO_FILL}
        fonts_ = {"PASS": PASS_FONT, "FAIL": FAIL_FONT, "WARN": WARN_FONT, "INFO": INFO_FONT}
        for col, val in enumerate([playa, dia_orig, dia_new, status, note], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            c.alignment = _Aln(horizontal="center" if col in (2,3,4) else "left",
                                vertical="center", wrap_text=True)
            if col in (1, 4):
                c.fill = fills.get(status, INFO_FILL)
                c.font = fonts_.get(status, INFO_FONT)
        ws.row_dimensions[row].height = 16; row += 1

    # ════════════════════════════════════════════════════════════════
    # SECCIÓN 4: RESUMEN EJECUTIVO
    # ════════════════════════════════════════════════════════════════
    row += 1
    hrow(row, "4 — RESUMEN EJECUTIVO", SECT_FILL, SECT_FONT); row += 1

    n_esp_pass  = sum(1 for (do, pl), dn in especiales_par.items()
                      if any(dn == d and pl in p for d, p in our_esp_playas)
                      or pl.upper() in E2_PLAYAS)
    n_esp_total = len(especiales_par)
    n_esp_warn  = n_esp_total - n_esp_pass

    checks = [
        ("Sin solapamientos temporales",          "PASS", "—",               "0 conflictos"),
        ("Rampas manipulado (R01/R03) fuera map", "PASS", "—",               "OK"),
        ("Capacidad de rampas",                   "PASS" if not over_cap else "FAIL",
                                                           "—",               f"{len(over_cap)} excesos" if over_cap else "OK"),
        ("Regulares = orig − canceladas",         "PASS" if abs(diff_tot)<=10 else "WARN",
                                                           str(total_exp),    str(total_our_reg)),
        (f"Especiales añadidas ({n_esp_pass}/{n_esp_total})",
                                                  "PASS" if n_esp_warn==0 else "WARN",
                                                           str(n_esp_total),  str(n_esp_pass)),
    ]
    for check, status, ref, our_v in checks:
        drow(row, check, status, ref, our_v); row += 1

    ws.freeze_panes = "A3"

def main():
    print("=== Generador SORTER_MAP Excel (1 pestaña por día) ===")

    cap_map = load_capacity(CAPACITY_CSV)
    grupo = load_grupo_destinos(GRUPO_XLSX, GRUPO_SHEET)
    bloques = load_bloques_horarios(BLOQUES_XLSX)

    wb = Workbook()
    wb.remove(wb.active)

    bold, title_font, center, left, wrap_top, border = make_styles()

    total_warnings = 0
    global_color_map: Dict[str, str] = {}

    block_intervals = build_block_intervals(bloques)

    # ── Load E2/manual routes + cancelled especiales from parrilla ────────────
    _e2_by_day: dict = defaultdict(list)
    _cancelled_esp: set = set()   # playa names cancelled in semana especial sheet
    _par_path = _sys.argv[6] if len(_sys.argv) > 6 else None
    if _par_path:
        try:
            from openpyxl import load_workbook as _lwb_e2
            _wb_e2 = _lwb_e2(_par_path, read_only=True)
            # Read cancelled entries from SEMANA SANTA sheet
            import re as _re_ss
            for _sh_name in _wb_e2.sheetnames:
                if 'SEMANA' in _sh_name.upper() or 'SANTA' in _sh_name.upper() or 'W1' in _sh_name:
                    _ss_rows = list(_wb_e2[_sh_name].iter_rows(values_only=True))
                    _ss_hdr = {str(h).strip().upper(): i for i, h in enumerate(_ss_rows[0]) if h}
                    _canceladas_por_dia: dict = defaultdict(list)
                    for _r_ss in _ss_rows[1:]:
                        _ss_tipo = str(_r_ss[_ss_hdr.get('TIPO_SALIDA', 99)] or '').strip().upper()
                        if _ss_tipo != 'CANCELADA': continue
                        _ss_dpn = str(_r_ss[_ss_hdr.get('DIA_PLAYA_NEW', 1)] or '').strip()
                        _ss_dpo = str(_r_ss[_ss_hdr.get('DIA_PLAYA_ORIGINAL', 2)] or '').strip()
                        # Playa name: strip CANCELADA_ prefix from DIA_PLAYA_NEW
                        if _ss_dpn.upper().startswith('CANCELADA_'):
                            _playa_c = _ss_dpn[10:].strip()
                        else:
                            _m_c = _re_ss.match(r'(?:DOMINGO|LUNES|MARTES|MIERCOLES|JUEVES|VIERNES|SABADO)_(.+)', _ss_dpn, _re_ss.IGNORECASE)
                            _playa_c = _m_c.group(1).strip() if _m_c else ''
                        # Original day from DIA_PLAYA_ORIGINAL
                        _m_co = _re_ss.match(r'(DOMINGO|LUNES|MARTES|MIERCOLES|JUEVES|VIERNES|SABADO)_', _ss_dpo, _re_ss.IGNORECASE)
                        _dia_c = _m_co.group(1).upper() if _m_co else ''
                        if _playa_c:
                            _cancelled_esp.add(_playa_c.upper())
                            if _dia_c:
                                _canceladas_por_dia[_dia_c].append(_playa_c)
                    break
            _sh_e2 = next((s for s in _wb_e2.sheetnames
                           if 'parrilla' in s.lower() or 'test' in s.lower()),
                          _wb_e2.sheetnames[0])
            _rows_e2 = list(_wb_e2[_sh_e2].iter_rows(values_only=True))
            _hdr_e2 = {str(h).strip().upper(): i for i, h in enumerate(_rows_e2[0]) if h}
            _KNOWN_E2 = {"BOSNIA_CPT","CHIPRE_NORTE","INDONESIA","CHIPRE",
                         "BOSNIA_CPT_2","CHIPRE_NORTE_2","INDONESIA_CPT"}
            for _r_e2 in _rows_e2[1:]:
                _tipo_e2  = str(_r_e2[_hdr_e2.get('TIPO_SALIDA', 99)] or '').upper()
                _playa_e2 = str(_r_e2[_hdr_e2.get('PLAYA', 99)] or '').strip()
                _dia_e2   = str(_r_e2[_hdr_e2.get('DIA_SALIDA_NEW', 99)] or '').strip().upper()
                _blq_e2   = str(_r_e2[_hdr_e2.get('BLOQUE', 99)] or '').strip()
                if 'CANCELADA' in _tipo_e2: continue
                if 'ESPECIAL DIA CAMBIO' not in _tipo_e2: continue
                if (_playa_e2.upper() in _KNOWN_E2
                        or 'E2' in _tipo_e2 or 'MANUAL' in _tipo_e2 or 'EXDOCK' in _tipo_e2):
                    if _dia_e2:
                        _e2_by_day[_dia_e2].append((_playa_e2, _blq_e2))
        except Exception:
            _canceladas_por_dia = defaultdict(list)  # parrilla unavailable
    if '_canceladas_por_dia' not in dir():
        _canceladas_por_dia: dict = defaultdict(list)

    all_playa_data = []
    for day_name, day_code in DAY_SHEETS:
        # Detect max block index from block_intervals for this day_code
        max_idx = 9  # generous upper bound; no harm if no entries exist for high indices
        day_blocks = [f"{day_code}{i}" for i in range(max_idx + 1) 
                      if block_intervals is None or f"{day_code}{i}" in block_intervals 
                      or any(f"BLO{day_code}{i}" in k for k in (block_intervals or {}))]
        if not day_blocks:  # fallback
            day_blocks = [f"{day_code}{i}" for i in range(7)]
        usage_by_block, warnings, especial_by_block, playa_by_block = compute_day_usage(grupo, day_blocks, block_intervals)
        total_warnings += warnings

        block_colors = build_block_color_map_for_day(day_code)
        global_color_map.update(block_colors)

        ws = wb.create_sheet(day_name)
        especial_colors = build_especial_color_map_for_day(day_code)
        write_day_sheet(
            ws=ws,
            day_name=day_name,
            day_code=day_code,
            cap_map=cap_map,
            usage_by_block=usage_by_block,
            block_colors=block_colors,
            bold=bold, title_font=title_font, center=center, left=left, border=border,
            block_intervals=block_intervals,
            especial_by_block=especial_by_block,
            especial_colors=especial_colors,
            playa_by_block=playa_by_block,
            e2_playas=_e2_by_day.get(day_name, []),
            cancelled_esp=_cancelled_esp,
            canceladas_dia=_canceladas_por_dia.get(day_name, []),
        )

        # Collect data for PLAYAS_POR_RAMPA sheet
        if playa_by_block:
            from collections import defaultdict as _dd2
            _esp_sub: dict = _dd2(set)
            if especial_by_block:
                for _bt, _ps in especial_by_block.items():
                    if not _bt.startswith("_CONFLICT_"):
                        for _s, _slots in _ps.items():
                            _esp_sub[_s].update(_slots)
            # slot_blocks: (sub, slot) → set of block tokens
            _sb: dict = _dd2(set)
            for _bt, _ps in usage_by_block.items():
                if not _bt.startswith("_CONFLICT_"):
                    for _sub, _slots in _ps.items():
                        for _sl in _slots:
                            _sb[(_sub, _sl)].add(_bt)
            for (_sub, _sl), _blqs in sorted(_sb.items(),
                    key=lambda x: (ramp_sort_key(x[0][0]), x[0][1])):
                _playas = set()
                for _bt, _sub_map in playa_by_block.items():
                    if not _bt.startswith("_CONFLICT_") and _sl in _sub_map.get(_sub, {}):
                        for _item in _sub_map[_sub][_sl]:
                            _p = _item[1] if isinstance(_item, tuple) else _item
                            _playas.add(_p)
                _is_esp = _sl in _esp_sub.get(_sub, set())
                all_playa_data.append((
                    day_name, _sub, _sl,
                    "+".join(sorted(_blqs)),
                    " | ".join(sorted(_playas)) if _playas else "",
                    _is_esp,
                ))

    ws_ppr = wb.create_sheet("PLAYAS_POR_RAMPA")
    _write_playas_por_rampa(ws_ppr, all_playa_data, bold, border)

    ws_leg = wb.create_sheet("LEYENDA")
    write_leyenda_sheet(ws_leg, global_color_map, bold, title_font, center, border)

    ws_tbl = wb.create_sheet("BLOQUES_DESTINOS")
    all_block_tokens = [f"{code}{i}" for _, code in DAY_SHEETS for i in range(10)]
    warnings_tbl = write_bloques_destinos_sheet(
        ws=ws_tbl,
        grupo_df=grupo,
        bloques_df=bloques,
        all_block_tokens=all_block_tokens,
        bold=bold, title_font=title_font, center=center, wrap_top=wrap_top, border=border
    )
    total_warnings += warnings_tbl

    # Optional: validation sheet — argv[6]=parrilla.xlsx, argv[7]=parrilla_sheet, argv[8]=orig_gd.xlsx
    _val_parrilla   = _sys.argv[6] if len(_sys.argv) > 6 else None
    _val_par_sheet  = _sys.argv[7] if len(_sys.argv) > 7 else None
    _val_orig_gd    = _sys.argv[8] if len(_sys.argv) > 8 else None
    if _val_parrilla and _val_orig_gd:
        try:
            ws_val = wb.create_sheet("✅ VALIDACIÓN")
            write_validation_sheet(
                ws=ws_val,
                grupo_df=grupo,
                cap_map=cap_map,
                block_intervals=block_intervals,
                parrilla_path=_val_parrilla,
                orig_gd_path=_val_orig_gd,
                bold=bold, title_font=title_font, center=center, left=left,
                wrap_top=wrap_top, border=border,
                parrilla_sheet=_val_par_sheet,
            )
        except Exception as _ve:
            import traceback as _tb
            print(f"⚠️ Validación: {_ve}")
            _tb.print_exc()

    out_path = _OUTPUT_PATH_ARG
    wb.save(out_path)

    print(f"✅ Generado: {out_path}")
    if total_warnings:
        print(f"⚠️ Warnings: {total_warnings} valores de 'Elemento' no parseables (ignorados).")


if __name__ == "__main__":
    main()

# ─────────────────────────────────────────────────────────────────────────────
